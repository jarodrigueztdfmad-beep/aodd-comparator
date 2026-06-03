"""
AODD Pump Comparator - Multilingual
Uso: streamlit run aodd_comparator.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# TRANSLATIONS
# ─────────────────────────────────────────────────────────────────────────────
LANGUAGES = {
    "Español":   "es",
    "English":   "en",
    "Deutsch":   "de",
    "Francais":  "fr",
    "Portugues": "pt",
    "Romana":    "ro",
    "Cestina":   "cs",
}
LANG_LABELS = {
    "Español":   "🇪🇸 Spanish",
    "English":   "🇬🇧 English",
    "Deutsch":   "🇩🇪 German",
    "Francais":  "🇫🇷 French",
    "Portugues": "🇵🇹 Portuguese",
    "Romana":    "🇷🇴 Romanian",
    "Cestina":   "🇨🇿 Czech",
}

T = {
    "app_title": {
        "es": "AODD Pump Comparator", "en": "AODD Pump Comparator",
        "de": "AODD Pumpen-Vergleich", "fr": "Comparateur Pompes AODD",
        "pt": "Comparador de Bombas AODD", "ro": "Comparator Pompe AODD",
        "cs": "Porovnani cerpadel AODD",
    },
    "app_subtitle": {
        "es": "Comparativa de eficiencia neumatica - Calculo de KPIs - Exportacion a Excel",
        "en": "Pneumatic efficiency comparison - KPI calculation - Excel export",
        "de": "Pneumatischer Effizienzvergleich - KPI-Berechnung - Excel-Export",
        "fr": "Comparaison efficacite pneumatique - Calcul KPI - Export Excel",
        "pt": "Comparacao de eficiencia pneumatica - Calculo de KPIs - Exportar Excel",
        "ro": "Comparatie eficienta pneumatica - Calcul KPI - Export Excel",
        "cs": "Porovnani pneumaticke ucinnosti - Vypocet KPI - Export Excel",
    },
    "pump_mgmt": {
        "es": "GESTION DE BOMBAS", "en": "PUMP MANAGEMENT",
        "de": "PUMPENVERWALTUNG", "fr": "GESTION DES POMPES",
        "pt": "GESTAO DE BOMBAS", "ro": "GESTIONARE POMPE",
        "cs": "SPRAVA CERPADEL",
    },
    "add_pump": {
        "es": "Anadir nueva bomba", "en": "Add new pump",
        "de": "Neue Pumpe hinzufugen", "fr": "Ajouter une pompe",
        "pt": "Adicionar nova bomba", "ro": "Adaugati pompa noua",
        "cs": "Pridat nove cerpadlo",
    },
    "active_pumps": {
        "es": "BOMBAS ACTIVAS", "en": "ACTIVE PUMPS",
        "de": "AKTIVE PUMPEN", "fr": "POMPES ACTIVES",
        "pt": "BOMBAS ATIVAS", "ro": "POMPE ACTIVE",
        "cs": "AKTIVNI CERPADLA",
    },
    "pump_name": {
        "es": "Nombre", "en": "Name", "de": "Name",
        "fr": "Nom", "pt": "Nome", "ro": "Nume", "cs": "Nazev",
    },
    "flow_lpm": {
        "es": "Caudal (LPM)", "en": "Flow rate (LPM)", "de": "Durchfluss (LPM)",
        "fr": "Debit (LPM)", "pt": "Vazao (LPM)", "ro": "Debit (LPM)", "cs": "Prutok (LPM)",
    },
    "disch_press": {
        "es": "P. descarga (bar)", "en": "Discharge pressure (bar)",
        "de": "Ausgangsdruck (bar)", "fr": "Pression refoulement (bar)",
        "pt": "Pressao descarga (bar)", "ro": "Presiune refulare (bar)",
        "cs": "Vytlacny tlak (bar)",
    },
    "air_press": {
        "es": "P. aire (bar)", "en": "Air pressure (bar)", "de": "Luftdruck (bar)",
        "fr": "Pression air (bar)", "pt": "Pressao ar (bar)",
        "ro": "Presiune aer (bar)", "cs": "Tlak vzduchu (bar)",
    },
    "air_cons": {
        "es": "Consumo aire (Nm3/h)", "en": "Air consumption (Nm3/h)",
        "de": "Luftverbrauch (Nm3/h)", "fr": "Consommation air (Nm3/h)",
        "pt": "Consumo de ar (Nm3/h)", "ro": "Consum aer (Nm3/h)",
        "cs": "Spotreba vzduchu (Nm3/h)",
    },
    "lpc": {
        "es": "Litros/ciclo", "en": "Litres/stroke", "de": "Liter/Hub",
        "fr": "Litres/cycle", "pt": "Litros/ciclo",
        "ro": "Litri/ciclu", "cs": "Litry/zdvih",
    },
    "notes_opt": {
        "es": "Notas (opcional)", "en": "Notes (optional)", "de": "Anmerkungen (optional)",
        "fr": "Notes (optionnel)", "pt": "Notas (opcional)",
        "ro": "Note (optional)", "cs": "Poznamky (volitelne)",
    },
    "add_btn": {
        "es": "Anadir bomba", "en": "Add pump", "de": "Pumpe hinzufugen",
        "fr": "Ajouter la pompe", "pt": "Adicionar bomba",
        "ro": "Adaugati pompa", "cs": "Pridat cerpadlo",
    },
    "name_req": {
        "es": "Introduce un nombre.", "en": "Please enter a name.",
        "de": "Bitte einen Namen eingeben.", "fr": "Veuillez saisir un nom.",
        "pt": "Por favor introduza um nome.", "ro": "Introduceti un nume.",
        "cs": "Zadejte prosim nazev.",
    },
    "ref_pump": {
        "es": "Bomba de referencia para deltas", "en": "Reference pump for deltas",
        "de": "Referenzpumpe fuer Deltas", "fr": "Pompe de reference pour les deltas",
        "pt": "Bomba de referencia para deltas", "ro": "Pompa de referinta pentru delta",
        "cs": "Referencni cerpadlo pro delta",
    },
    "delete_pump": {
        "es": "Eliminar", "en": "Delete", "de": "Loschen",
        "fr": "Supprimer", "pt": "Eliminar", "ro": "Sterge", "cs": "Smazat",
    },
    "restore_ex": {
        "es": "Restaurar ejemplos", "en": "Restore examples",
        "de": "Beispiele zuruecksetzen", "fr": "Restaurer les exemples",
        "pt": "Restaurar exemplos", "ro": "Restaurati exemplele",
        "cs": "Obnovit priklady",
    },
    "ftmg_title": {
        "es": "IMPORTAR SICK FTMg", "en": "IMPORT SICK FTMg",
        "de": "SICK FTMg IMPORTIEREN", "fr": "IMPORTER SICK FTMg",
        "pt": "IMPORTAR SICK FTMg", "ro": "IMPORT SICK FTMg",
        "cs": "IMPORT SICK FTMg",
    },
    "ftmg_load": {
        "es": "Cargar CSV del medidor", "en": "Load meter CSV",
        "de": "Messgeraet-CSV laden", "fr": "Charger CSV du compteur",
        "pt": "Carregar CSV do medidor", "ro": "Incarcati CSV-ul contorului",
        "cs": "Nacist CSV merice",
    },
    "ftmg_unit": {
        "es": "Unidad de caudal en el CSV:", "en": "Flow unit in CSV:",
        "de": "Durchflusseinheit in CSV:", "fr": "Unite de debit dans le CSV:",
        "pt": "Unidade de caudal no CSV:", "ro": "Unitatea de debit in CSV:",
        "cs": "Jednotka prutoku v CSV:",
    },
    "ftmg_range": {
        "es": "Filas a incluir", "en": "Rows to include",
        "de": "Zeilen einschliessen", "fr": "Lignes a inclure",
        "pt": "Linhas a incluir", "ro": "Randuri de inclus",
        "cs": "Radky k zahrnutí",
    },
    "ftmg_mean": {
        "es": "Consumo medio calculado", "en": "Calculated mean consumption",
        "de": "Berechneter mittlerer Verbrauch", "fr": "Consommation moyenne calculee",
        "pt": "Consumo medio calculado", "ro": "Consum mediu calculat",
        "cs": "Vypoctena prumerna spotreba",
    },
    "ftmg_assign": {
        "es": "Asignar a bomba:", "en": "Assign to pump:",
        "de": "Pumpe zuweisen:", "fr": "Affecter a la pompe:",
        "pt": "Atribuir a bomba:", "ro": "Atribuiti pompei:",
        "cs": "Priradit cerpadlu:",
    },
    "ftmg_apply": {
        "es": "Aplicar a la bomba", "en": "Apply to pump",
        "de": "Auf Pumpe anwenden", "fr": "Appliquer a la pompe",
        "pt": "Aplicar a bomba", "ro": "Aplicati la pompa",
        "cs": "Pouzit na cerpadlo",
    },
    "ftmg_chart": {
        "es": "Ver grafico del CSV", "en": "Show CSV chart",
        "de": "CSV-Diagramm anzeigen", "fr": "Afficher le graphique CSV",
        "pt": "Ver grafico do CSV", "ro": "Afisati graficul CSV",
        "cs": "Zobrazit graf CSV",
    },
    "kpi_section": {
        "es": "KPIs calculados", "en": "Calculated KPIs",
        "de": "Berechnete KPIs", "fr": "KPIs calcules",
        "pt": "KPIs calculados", "ro": "KPI calculati",
        "cs": "Vypoctene KPI",
    },
    "kpi_air": {
        "es": "Consumo aire", "en": "Air consumption",
        "de": "Luftverbrauch", "fr": "Conso. air",
        "pt": "Consumo ar", "ro": "Consum aer", "cs": "Spotreba vzduchu",
    },
    "kpi_cycles": {
        "es": "Ciclos/min", "en": "Strokes/min", "de": "Hube/min",
        "fr": "Cycles/min", "pt": "Ciclos/min",
        "ro": "Cicluri/min", "cs": "Zdvihy/min",
    },
    "kpi_specific": {
        "es": "Aire especifico", "en": "Specific air", "de": "Spezifische Luft",
        "fr": "Air specifique", "pt": "Ar especifico",
        "ro": "Aer specific", "cs": "Merny vzduch",
    },
    "kpi_nodata": {
        "es": "Datos insuficientes para calcular KPIs",
        "en": "Insufficient data to calculate KPIs",
        "de": "Unzureichende Daten zur KPI-Berechnung",
        "fr": "Donnees insuffisantes pour calculer les KPI",
        "pt": "Dados insuficientes para calcular KPIs",
        "ro": "Date insuficiente pentru calculul KPI",
        "cs": "Nedostatek dat pro vypocet KPI",
    },
    "kpi_explain": {
        "es": "Aire especifico (Nm3/m3) = m3 de aire consumidos por cada m3 de liquido bombeado. Indicador clave de eficiencia neumatica: cuanto mas bajo, menos aire gasta la bomba.",
        "en": "Specific air (Nm3/m3) = cubic metres of air per cubic metre of liquid pumped. Key pneumatic efficiency indicator: the lower, the less air the pump uses.",
        "de": "Spezifische Luft (Nm3/m3) = Kubikmeter Luft pro Kubikmeter Fluessigkeit. Wichtigster Effizienzindikator: je niedriger, desto effizienter.",
        "fr": "Air specifique (Nm3/m3) = m3 d'air par m3 de liquide pompe. Indicateur cle: plus la valeur est faible, plus la pompe est efficace.",
        "pt": "Ar especifico (Nm3/m3) = m3 de ar por m3 de liquido bombeado. Indicador chave: quanto menor, menos ar a bomba gasta.",
        "ro": "Aer specific (Nm3/m3) = m3 aer per m3 lichid pompat. Indicator cheie: cu cat mai mic, cu atat mai eficienta.",
        "cs": "Merny vzduch (Nm3/m3) = m3 vzduchu na m3 cerpane kapaliny. Klicovy ukazatel: cim nizsi, tim efektivnejsi.",
    },
    "tab_charts": {
        "es": "Graficos", "en": "Charts", "de": "Diagramme",
        "fr": "Graphiques", "pt": "Graficos", "ro": "Grafice", "cs": "Grafy",
    },
    "tab_table": {
        "es": "Tabla comparativa", "en": "Comparison table", "de": "Vergleichstabelle",
        "fr": "Tableau comparatif", "pt": "Tabela comparativa",
        "ro": "Tabel comparativ", "cs": "Srovnavaci tabulka",
    },
    "tab_deltas": {
        "es": "Diferencias", "en": "Differences", "de": "Unterschiede",
        "fr": "Differences", "pt": "Diferencas", "ro": "Diferente", "cs": "Rozdily",
    },
    "tab_export": {
        "es": "Exportar Excel", "en": "Export Excel", "de": "Excel exportieren",
        "fr": "Exporter Excel", "pt": "Exportar Excel",
        "ro": "Export Excel", "cs": "Export Excel",
    },
    "table_title": {
        "es": "Parametros e indicadores", "en": "Parameters and indicators",
        "de": "Parameter und Indikatoren", "fr": "Parametres et indicateurs",
        "pt": "Parametros e indicadores", "ro": "Parametri si indicatori",
        "cs": "Parametry a ukazatele",
    },
    "table_note": {
        "es": "Ciclos/min = Caudal / L/ciclo  |  Aire especifico = Consumo (Nm3/h) / Caudal (m3/h)",
        "en": "Strokes/min = Flow / L/stroke  |  Specific air = Consumption (Nm3/h) / Flow (m3/h)",
        "de": "Hube/min = Durchfluss / L/Hub  |  Spez. Luft = Verbrauch (Nm3/h) / Durchfluss (m3/h)",
        "fr": "Cycles/min = Debit / L/cycle  |  Air spec. = Consommation (Nm3/h) / Debit (m3/h)",
        "pt": "Ciclos/min = Caudal / L/ciclo  |  Ar esp. = Consumo (Nm3/h) / Caudal (m3/h)",
        "ro": "Cicluri/min = Debit / L/ciclu  |  Aer spec. = Consum (Nm3/h) / Debit (m3/h)",
        "cs": "Zdvihy/min = Prutok / L/zdvih  |  Merny vzduch = Spotreba (Nm3/h) / Prutok (m3/h)",
    },
    "deltas_title": {
        "es": "Diferencias absolutas y porcentuales",
        "en": "Absolute and percentage differences",
        "de": "Absolute und prozentuale Unterschiede",
        "fr": "Differences absolues et en pourcentage",
        "pt": "Diferencas absolutas e percentuais",
        "ro": "Diferente absolute si procentuale",
        "cs": "Absolutni a procentualni rozdily",
    },
    "ref_label": {
        "es": "Bomba de referencia:", "en": "Reference pump:",
        "de": "Referenzpumpe:", "fr": "Pompe de reference:",
        "pt": "Bomba de referencia:", "ro": "Pompa de referinta:",
        "cs": "Referencni cerpadlo:",
    },
    "need_two": {
        "es": "Anade al menos 2 bombas para ver diferencias.",
        "en": "Add at least 2 pumps to see differences.",
        "de": "Fuege mindestens 2 Pumpen hinzu um Unterschiede zu sehen.",
        "fr": "Ajoutez au moins 2 pompes pour voir les differences.",
        "pt": "Adicione pelo menos 2 bombas para ver diferencas.",
        "ro": "Adaugati cel putin 2 pompe pentru a vedea diferentele.",
        "cs": "Pridejte alespon 2 cerpadla pro zobrazeni rozdilu.",
    },
    "export_title": {
        "es": "Exportar a Excel", "en": "Export to Excel",
        "de": "Nach Excel exportieren", "fr": "Exporter vers Excel",
        "pt": "Exportar para Excel", "ro": "Export in Excel",
        "cs": "Exportovat do Excelu",
    },
    "export_btn": {
        "es": "Descargar AODD_Comparativa.xlsx", "en": "Download AODD_Comparison.xlsx",
        "de": "AODD_Vergleich.xlsx herunterladen", "fr": "Telecharger AODD_Comparatif.xlsx",
        "pt": "Descarregar AODD_Comparativa.xlsx", "ro": "Descarcati AODD_Comparativ.xlsx",
        "cs": "Stahnout AODD_Porovnani.xlsx",
    },
    "preview_title": {
        "es": "Vista previa de datos exportados", "en": "Preview of exported data",
        "de": "Vorschau der exportierten Daten", "fr": "Apercu des donnees exportees",
        "pt": "Pre-visualizacao dos dados exportados", "ro": "Previzualizare date exportate",
        "cs": "Nahled exportovanych dat",
    },
    "chart_kpi": {
        "es": "Comparativa de KPIs principales", "en": "Main KPI comparison",
        "de": "Wichtigste KPI-Vergleich", "fr": "Comparaison des KPI principaux",
        "pt": "Comparacao dos KPIs principais", "ro": "Compararea KPI-urilor principale",
        "cs": "Porovnani hlavnich KPI",
    },
    "chart_radar": {
        "es": "Perfil comparativo (valores normalizados)",
        "en": "Comparative profile (normalised values)",
        "de": "Vergleichsprofil (normalisierte Werte)",
        "fr": "Profil comparatif (valeurs normalisees)",
        "pt": "Perfil comparativo (valores normalizados)",
        "ro": "Profil comparativ (valori normalizate)",
        "cs": "Srovnavaci profil (normalizovane hodnoty)",
    },
    "chart_scatter_title": {
        "es": "Eficiencia neumatica vs. velocidad de ciclo",
        "en": "Pneumatic efficiency vs. stroke speed",
        "de": "Pneumatische Effizienz vs. Hubgeschwindigkeit",
        "fr": "Efficacite pneumatique vs. vitesse de cycle",
        "pt": "Eficiencia pneumatica vs. velocidade de ciclo",
        "ro": "Eficienta pneumatica vs. viteza ciclului",
        "cs": "Pneumaticka ucinnost vs. rychlost zdvihu",
    },
    "chart_scatter_x": {
        "es": "Aire especifico (Nm3/m3)  - menor = mas eficiente",
        "en": "Specific air (Nm3/m3)  - lower = more efficient",
        "de": "Spezifische Luft (Nm3/m3)  - niedriger = effizienter",
        "fr": "Air specifique (Nm3/m3)  - plus bas = plus efficace",
        "pt": "Ar especifico (Nm3/m3)  - menor = mais eficiente",
        "ro": "Aer specific (Nm3/m3)  - mai mic = mai eficient",
        "cs": "Merny vzduch (Nm3/m3)  - nizsi = efektivnejsi",
    },
    "chart_wf_title": {
        "es": "Diferencia de eficiencia relativa a",
        "en": "Efficiency difference relative to",
        "de": "Effizienzunterschied relativ zu",
        "fr": "Difference d efficacite par rapport a",
        "pt": "Diferenca de eficiencia relativa a",
        "ro": "Diferenta de eficienta fata de",
        "cs": "Rozdil ucinnosti relativne k",
    },
    "chart_wf_y": {
        "es": "Delta % aire especifico vs",
        "en": "Delta % specific air vs",
        "de": "Delta % spezifische Luft vs",
        "fr": "Delta % air specifique vs",
        "pt": "Delta % ar especifico vs",
        "ro": "Delta % aer specific vs",
        "cs": "Delta % merny vzduch vs",
    },
    "footer_txt": {
        "es": "AODD Comparator - Datos de ejemplo: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Los valores de ejemplo son ilustrativos",
        "en": "AODD Comparator - Example data: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Example values are illustrative",
        "de": "AODD Vergleich - Beispieldaten: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Beispielwerte sind illustrativ",
        "fr": "Comparateur AODD - Donnees exemples: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Valeurs illustratives",
        "pt": "Comparador AODD - Dados de exemplo: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Valores ilustrativos",
        "ro": "Comparator AODD - Date exemplu: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Valori ilustrative",
        "cs": "Porovnani AODD - Prikladova data: Wilden PS820 EOM / SANDPIPER S20 / SAMOA UP20 | Hodnoty jsou ilustrativni",
    },
    "ex_note": {
        "es": "Ejemplo - fuente fabricante", "en": "Example - manufacturer source",
        "de": "Beispiel - Herstellerquelle", "fr": "Exemple - source fabricant",
        "pt": "Exemplo - fonte fabricante", "ro": "Exemplu - sursa producator",
        "cs": "Priklad - zdroj vyrobce",
    },
    # ── Manual FTMg ──────────────────────────────────────────────────────────
    "tab_manual": {
        "es": "📖 Manual FTMg", "en": "📖 FTMg Manual",
        "de": "📖 FTMg Handbuch", "fr": "📖 Manuel FTMg",
        "pt": "📖 Manual FTMg", "ro": "📖 Manual FTMg", "cs": "📖 FTMg Manual",
    },
    "man_title": {
        "es": "Manual paso a paso — FTMg + Ensayo comparativo AODD",
        "en": "Step-by-step guide — FTMg + AODD Comparative Test",
        "de": "Schritt-fur-Schritt-Anleitung — FTMg + AODD Vergleichstest",
        "fr": "Guide etape par etape — FTMg + Test comparatif AODD",
        "pt": "Guia passo a passo — FTMg + Ensaio comparativo AODD",
        "ro": "Ghid pas cu pas — FTMg + Test comparativ AODD",
        "cs": "Pruvodce krok za krokem — FTMg + Srovnavaci test AODD",
    },
    "man_objective": {
        "es": "Objetivo: que cualquier tecnico pueda repetir el ensayo con resultados defendibles ante el cliente: mismo punto de operacion, mismos criterios, mismos KPI.",
        "en": "Objective: any technician should be able to repeat the test with results that can be defended to the client: same operating point, same criteria, same KPIs.",
        "de": "Ziel: Jeder Techniker soll den Test mit vor dem Kunden vertretbaren Ergebnissen wiederholen konnen: gleicher Betriebspunkt, gleiche Kriterien, gleiche KPIs.",
        "fr": "Objectif: tout technicien doit pouvoir repeter l'essai avec des resultats defensables aupres du client: meme point de fonctionnement, memes criteres, memes KPI.",
        "pt": "Objetivo: qualquer tecnico deve conseguir repetir o ensaio com resultados defesaveis perante o cliente: mesmo ponto de operacao, mesmos criterios, mesmos KPIs.",
        "ro": "Obiectiv: orice tehnician sa poata repeta testul cu rezultate justificabile in fata clientului: acelasi punct de functionare, aceleasi criterii, aceiasi KPI.",
        "cs": "Cil: kazdy technik musi byt schopen zopakovat test s vysledky obhajitelnymi pred zakaznikem: stejny provozni bod, stejna kriteria, stejne KPI.",
    },
    "man_A_title": {
        "es": "A. Antes de nada: define que significa igualdad de condiciones",
        "en": "A. First: define what equal conditions means",
        "de": "A. Zuerst: Definiere was Gleichheit der Bedingungen bedeutet",
        "fr": "A. D'abord: definir ce que signifie egalite des conditions",
        "pt": "A. Primeiro: define o que significa igualdade de condicoes",
        "ro": "A. In primul rand: defineste ce inseamna egalitatea conditiilor",
        "cs": "A. Nejdrive: definujte co znamena rovnost podminek",
    },
    "man_A1": {
        "es": "A1. Igualdad hidraulica (liquido)",
        "en": "A1. Hydraulic equality (liquid)",
        "de": "A1. Hydraulische Gleichheit (Flussigkeit)",
        "fr": "A1. Egalite hydraulique (liquide)",
        "pt": "A1. Igualdade hidraulica (liquido)",
        "ro": "A1. Egalitate hidraulica (lichid)",
        "cs": "A1. Hydraulicka rovnost (kapalina)",
    },
    "man_A1_items": {
        "es": ["Mismo producto (densidad/viscosidad) y misma temperatura",
               "Misma instalacion: tuberias, valvulas, codos, mangueras, filtros",
               "Misma condicion de aspiracion: inundada vs aspiracion (altura/depresion)",
               "Mismo punto de operacion: comparar por presion de descarga (p.ej. 2/4/6 bar g)"],
        "en": ["Same product (density/viscosity) and same temperature",
               "Same installation: pipes, valves, elbows, hoses, filters",
               "Same suction condition: flooded vs. suction (height/vacuum)",
               "Same operating point: compare by discharge pressure (e.g. 2/4/6 bar g)"],
        "de": ["Gleiches Produkt (Dichte/Viskositat) und gleiche Temperatur",
               "Gleiche Installation: Rohre, Ventile, Bogen, Schlauche, Filter",
               "Gleiche Saugbedingung: uberflutet vs. Saugbetrieb (Hohe/Vakuum)",
               "Gleicher Betriebspunkt: Vergleich nach Ausgangsdruck (z.B. 2/4/6 bar g)"],
        "fr": ["Meme produit (densite/viscosite) et meme temperature",
               "Meme installation: tuyaux, vannes, coudes, flexibles, filtres",
               "Meme condition d'aspiration: noyee vs aspiration (hauteur/depression)",
               "Meme point de fonctionnement: comparer par pression de refoulement (ex. 2/4/6 bar g)"],
        "pt": ["Mesmo produto (densidade/viscosidade) e mesma temperatura",
               "Mesma instalacao: tubagens, valvulas, cotovelos, mangueiras, filtros",
               "Mesma condicao de aspiracao: inundada vs aspiracao (altura/depressao)",
               "Mesmo ponto de operacao: comparar por pressao de descarga (ex. 2/4/6 bar g)"],
        "ro": ["Acelasi produs (densitate/viscozitate) si aceeasi temperatura",
               "Aceeasi instalatie: tevi, supape, coturi, furtunuri, filtre",
               "Aceeasi conditie de aspiratie: inundata vs aspiratie (inaltime/depresiune)",
               "Acelasi punct de functionare: comparare dupa presiunea de refulare (ex. 2/4/6 bar g)"],
        "cs": ["Stejny produkt (hustota/viskozita) a stejna teplota",
               "Stejne potrubi: trubky, ventily, kolena, hadice, filtry",
               "Stejna podminka sani: zaplavena vs saci provoz (vyska/vakuum)",
               "Stejny provozni bod: porovnani dle vytlacneho tlaku (napr. 2/4/6 bar g)"],
    },
    "man_A2": {
        "es": "A2. Igualdad neumatica (aire)",
        "en": "A2. Pneumatic equality (air)",
        "de": "A2. Pneumatische Gleichheit (Luft)",
        "fr": "A2. Egalite pneumatique (air)",
        "pt": "A2. Igualdade pneumatica (ar)",
        "ro": "A2. Egalitate pneumatica (aer)",
        "cs": "A2. Pneumaticka rovnost (vzduch)",
    },
    "man_A2_items": {
        "es": ["Mismo compresor/red si es posible",
               "Mismo regulador y filtracion",
               "Medir presion de aire REAL en la bomba, no solo en el regulador"],
        "en": ["Same compressor/network if possible",
               "Same regulator and filtration",
               "Measure REAL air pressure at the pump, not just at the regulator"],
        "de": ["Gleicher Kompressor/Netz wenn moglich",
               "Gleicher Regler und Filterung",
               "REALEN Luftdruck an der Pumpe messen, nicht nur am Regler"],
        "fr": ["Meme compresseur/reseau si possible",
               "Meme regulateur et filtration",
               "Mesurer la pression d'air REELLE a la pompe, pas seulement au regulateur"],
        "pt": ["Mesmo compressor/rede se possivel",
               "Mesmo regulador e filtracao",
               "Medir pressao de ar REAL na bomba, nao apenas no regulador"],
        "ro": ["Acelasi compresor/retea daca este posibil",
               "Acelasi regulator si filtrare",
               "Masurare presiune aer REALA la pompa, nu doar la regulator"],
        "cs": ["Stejny kompresor/sit pokud mozno",
               "Stejny regulator a filtrace",
               "Merit SKUTECNY tlak vzduchu u cerpadla, nejen u regulatoru"],
    },
    "man_A3": {
        "es": "A3. Igualdad de bomba y configuracion",
        "en": "A3. Pump and configuration equality",
        "de": "A3. Pumpen- und Konfigurationsgleichheit",
        "fr": "A3. Egalite de pompe et configuration",
        "pt": "A3. Igualdade de bomba e configuracao",
        "ro": "A3. Egalitate pompa si configuratie",
        "cs": "A3. Rovnost cerpadla a konfigurace",
    },
    "man_A3_items": {
        "es": ["Anotar modelo exacto, version de valvula de aire, materiales (diafragmas/bolas/asientos), silenciador",
               "Si una bomba es mas grande o tiene distinta valvuleria, NO es equivalente"],
        "en": ["Record exact model, air valve version, materials (diaphragms/balls/seats), muffler",
               "If one pump is larger or has different valving, it is NOT equivalent"],
        "de": ["Genaues Modell, Luftventilversion, Materialien (Membranen/Kugeln/Sitze), Schalldampfer aufzeichnen",
               "Wenn eine Pumpe grosser ist oder andere Ventile hat, ist sie NICHT aquivalent"],
        "fr": ["Noter le modele exact, version de vanne d'air, materiaux (membranes/billes/sieges), silencieux",
               "Si une pompe est plus grande ou a une robinetterie differente, elle n'est PAS equivalente"],
        "pt": ["Registar modelo exato, versao de valvula de ar, materiais (diafragmas/bolas/assentos), silenciador",
               "Se uma bomba for maior ou tiver valvularia diferente, NAO e equivalente"],
        "ro": ["Inregistrati modelul exact, versiunea supapei de aer, materiale (membrane/bile/scaune), amortizor",
               "Daca o pompa este mai mare sau are supape diferite, NU este echivalenta"],
        "cs": ["Zaznamenat presny model, verzi vzduchoveho ventilu, materialy (membrany/kulicky/sedla), tlumice",
               "Pokud je jedno cerpadlo vetsi nebo ma jine armatury, NENI ekvivalentni"],
    },
    "man_B_title": {
        "es": "B. Montaje correcto del SICK FTMg (evita el 80% de errores)",
        "en": "B. Correct SICK FTMg installation (avoids 80% of errors)",
        "de": "B. Korrekte SICK FTMg Montage (vermeidet 80% der Fehler)",
        "fr": "B. Montage correct du SICK FTMg (evite 80% des erreurs)",
        "pt": "B. Montagem correta do SICK FTMg (evita 80% dos erros)",
        "ro": "B. Montaj corect SICK FTMg (evita 80% din erori)",
        "cs": "B. Spravna montaz SICK FTMg (vyhne se 80% chyb)",
    },
    "man_B_items": {
        "es": ["Instala el medidor en la linea de aire: Red/compresor -> filtro/regulador -> FTMg -> bomba (lo mas cerca posible de la bomba, sin vibracion excesiva)",
               "Revisa orientacion IN/OUT",
               "El sistema debe estar limpio (filtros adecuados) para evitar errores por suciedad",
               "El FTMg mide flow/pressure/temperature en sistemas neumaticos y registra datos para analisis de consumo energetico"],
        "en": ["Install the meter in the air line: Network/compressor -> filter/regulator -> FTMg -> pump (as close as possible to the pump, without excessive vibration)",
               "Check IN/OUT orientation",
               "The system must be clean (adequate filters) to avoid errors from contamination",
               "The FTMg measures flow/pressure/temperature in pneumatic systems and logs data for energy consumption analysis"],
        "de": ["Zahler in der Luftleitung installieren: Netz/Kompressor -> Filter/Regler -> FTMg -> Pumpe (so nah wie moglich an der Pumpe, ohne ubermassige Vibration)",
               "IN/OUT-Ausrichtung prufen",
               "Das System muss sauber sein (geeignete Filter), um Fehler durch Verschmutzung zu vermeiden",
               "Der FTMg misst Durchfluss/Druck/Temperatur in pneumatischen Systemen und protokolliert Daten fur die Energieverbrauchsanalyse"],
        "fr": ["Installer le compteur dans la ligne d'air: Reseau/compresseur -> filtre/regulateur -> FTMg -> pompe (le plus pres possible de la pompe, sans vibration excessive)",
               "Verifier l'orientation IN/OUT",
               "Le systeme doit etre propre (filtres adequats) pour eviter les erreurs de contamination",
               "Le FTMg mesure debit/pression/temperature dans les systemes pneumatiques et enregistre des donnees pour l'analyse de consommation energetique"],
        "pt": ["Instalar o medidor na linha de ar: Rede/compressor -> filtro/regulador -> FTMg -> bomba (o mais proximo possivel da bomba, sem vibracao excessiva)",
               "Verificar orientacao IN/OUT",
               "O sistema deve estar limpo (filtros adequados) para evitar erros por contaminacao",
               "O FTMg mede caudal/pressao/temperatura em sistemas pneumaticos e regista dados para analise de consumo energetico"],
        "ro": ["Instalati contorul in linia de aer: Retea/compresor -> filtru/regulator -> FTMg -> pompa (cat mai aproape de pompa, fara vibratii excesive)",
               "Verificati orientarea IN/OUT",
               "Sistemul trebuie sa fie curat (filtre adecvate) pentru a evita erorile cauzate de contaminare",
               "FTMg masoara debit/presiune/temperatura in sisteme pneumatice si inregistreaza date pentru analiza consumului de energie"],
        "cs": ["Nainstalujte merak do vzduchovodu: Sit/kompresor -> filtr/regulator -> FTMg -> cerpadlo (co nejblize cerpadlu, bez nadmernych vibraci)",
               "Zkontrolujte orientaci IN/OUT",
               "System musi byt cisty (vhodne filtry), aby nedochazelo k chybam zpusobenym znecistenim",
               "FTMg meri prutok/tlak/teplotu v pneumatickych systemech a zaznamenava data pro analyzu spotreby energie"],
    },
    "man_C_title": {
        "es": "C. Configuracion y puesta en marcha del FTMg (paso a paso)",
        "en": "C. FTMg configuration and commissioning (step by step)",
        "de": "C. FTMg Konfiguration und Inbetriebnahme (Schritt fur Schritt)",
        "fr": "C. Configuration et mise en service du FTMg (etape par etape)",
        "pt": "C. Configuracao e colocacao em servico do FTMg (passo a passo)",
        "ro": "C. Configurarea si punerea in functiune a FTMg (pas cu pas)",
        "cs": "C. Konfigurace a uvadeni FTMg do provozu (krok za krokem)",
    },
    "man_C1": {
        "es": "C1. Conexion al sensor por IP",
        "en": "C1. IP connection to the sensor",
        "de": "C1. IP-Verbindung zum Sensor",
        "fr": "C1. Connexion IP au capteur",
        "pt": "C1. Ligacao IP ao sensor",
        "ro": "C1. Conexiune IP la senzor",
        "cs": "C1. IP pripojeni k senzoru",
    },
    "man_C1_items": {
        "es": ["En el sensor, lee IP y Submask (boton izquierdo)",
               "En el PC, configura una IP del mismo rango y la misma submask (cambia el ultimo octeto). Ejemplo: sensor 192.168.0.10 -> PC 192.168.0.8",
               "Abre el navegador y entra a la IP del sensor para abrir la interfaz web",
               "Login 'maintenance' con password: airflowsensor (acceso a opciones ampliadas)"],
        "en": ["On the sensor, read IP and Submask (left button)",
               "On the PC, configure an IP in the same range and same submask (change the last octet). Example: sensor 192.168.0.10 -> PC 192.168.0.8",
               "Open a browser and go to the sensor IP to open the web interface",
               "Login 'maintenance' with password: airflowsensor (access to extended options)"],
        "de": ["Am Sensor IP und Submask ablesen (linke Taste)",
               "Am PC eine IP im gleichen Bereich und gleicher Submask konfigurieren (letztes Oktet andern). Beispiel: Sensor 192.168.0.10 -> PC 192.168.0.8",
               "Browser offnen und Sensor-IP aufrufen um Weboberflache zu offnen",
               "Login 'maintenance' mit Passwort: airflowsensor (Zugang zu erweiterten Optionen)"],
        "fr": ["Sur le capteur, lire l'IP et le Submask (bouton gauche)",
               "Sur le PC, configurer une IP de la meme plage et le meme sous-masque (changer le dernier octet). Exemple: capteur 192.168.0.10 -> PC 192.168.0.8",
               "Ouvrir un navigateur et aller a l'IP du capteur pour ouvrir l'interface web",
               "Login 'maintenance' avec mot de passe: airflowsensor (acces aux options etendues)"],
        "pt": ["No sensor, ler IP e Submask (botao esquerdo)",
               "No PC, configurar um IP do mesmo intervalo e a mesma submask (mudar o ultimo octeto). Exemplo: sensor 192.168.0.10 -> PC 192.168.0.8",
               "Abrir browser e ir ao IP do sensor para abrir a interface web",
               "Login 'maintenance' com password: airflowsensor (acesso a opcoes alargadas)"],
        "ro": ["Pe senzor, cititi IP si Submask (buton stang)",
               "Pe PC, configurati un IP din acelasi interval si acelasi submask (schimbati ultimul octet). Exemplu: senzor 192.168.0.10 -> PC 192.168.0.8",
               "Deschideti un browser si accesati IP-ul senzorului pentru a deschide interfata web",
               "Login 'maintenance' cu parola: airflowsensor (acces la optiuni extinse)"],
        "cs": ["Na senzoru precist IP a Submask (leve tlacitko)",
               "Na PC nakonfigurovat IP ve stejnem rozsahu a se stejnou submask (zmenit posledni oktet). Priklad: senzor 192.168.0.10 -> PC 192.168.0.8",
               "Otevrit prohlizec a prejit na IP senzoru pro otevreni weboveho rozhrani",
               "Login 'maintenance' s heslem: airflowsensor (pristup k rozsirenym moznostem)"],
    },
    "man_C2": {
        "es": "C2. Seleccion de variables y logger",
        "en": "C2. Variable selection and logger",
        "de": "C2. Variablenauswahl und Logger",
        "fr": "C2. Selection des variables et logger",
        "pt": "C2. Selecao de variaveis e logger",
        "ro": "C2. Selectia variabilelor si logger",
        "cs": "C2. Vyber promennych a logger",
    },
    "man_C2_items": {
        "es": ["En Process data veras valores online en tiempo real",
               "En Logger define que grabar y cada cuanto tiempo",
               "Para comparativa AODD: Flow + Pressure + Temperature minimo (y energia si esta habilitada)",
               "Intervalo recomendado: 1-5 s para correlacionar con ciclos/min | 10 s para mayor estabilidad",
               "Inicia registro: Start recording",
               "Descarga log: Download last logfile a Excel"],
        "en": ["In Process data you will see online real-time values",
               "In Logger define what to record and how often",
               "For AODD comparison: Flow + Pressure + Temperature minimum (and energy if enabled)",
               "Recommended interval: 1-5 s to correlate with strokes/min | 10 s for greater stability",
               "Start recording: Start recording",
               "Download log: Download last logfile to Excel"],
        "de": ["In Process data sehen Sie Online-Echtzeitwerte",
               "Im Logger definieren was und wie oft aufgezeichnet werden soll",
               "Fur AODD-Vergleich: Durchfluss + Druck + Temperatur mindestens (und Energie wenn aktiviert)",
               "Empfohlenes Intervall: 1-5 s fur Korrelation mit Huben/min | 10 s fur mehr Stabilitat",
               "Aufzeichnung starten: Start recording",
               "Log herunterladen: Download last logfile nach Excel"],
        "fr": ["Dans Process data vous verrez les valeurs en ligne en temps reel",
               "Dans Logger definir quoi enregistrer et a quelle frequence",
               "Pour comparatif AODD: Debit + Pression + Temperature minimum (et energie si activee)",
               "Intervalle recommande: 1-5 s pour correler avec cycles/min | 10 s pour plus de stabilite",
               "Demarrer l'enregistrement: Start recording",
               "Telecharger le log: Download last logfile vers Excel"],
        "pt": ["Em Process data vera valores online em tempo real",
               "Em Logger define o que gravar e com que frequencia",
               "Para comparativa AODD: Flow + Pressure + Temperature minimo (e energia se habilitada)",
               "Intervalo recomendado: 1-5 s para correlacionar com ciclos/min | 10 s para maior estabilidade",
               "Iniciar registo: Start recording",
               "Descarregar log: Download last logfile para Excel"],
        "ro": ["In Process data veti vedea valorile online in timp real",
               "In Logger definiti ce sa inregistrati si cat de des",
               "Pentru comparativ AODD: Flow + Pressure + Temperature minim (si energie daca este activata)",
               "Interval recomandat: 1-5 s pentru corelarea cu cicluri/min | 10 s pentru stabilitate mai mare",
               "Porniti inregistrarea: Start recording",
               "Descarcati log: Download last logfile in Excel"],
        "cs": ["V Process data uvidite online hodnoty v realnem case",
               "V Logger definujte co zaznamenavat a jak casto",
               "Pro porovnani AODD: Flow + Pressure + Temperature minimum (a energie pokud je povolena)",
               "Doporuceny interval: 1-5 s pro korelaci se zdvihy/min | 10 s pro vetsi stabilitu",
               "Spustit zaznam: Start recording",
               "Stahnout log: Download last logfile do Excelu"],
    },
    "man_E_title": {
        "es": "E. Protocolo de ensayo comparativo",
        "en": "E. Comparative test protocol",
        "de": "E. Vergleichstest-Protokoll",
        "fr": "E. Protocole d'essai comparatif",
        "pt": "E. Protocolo de ensaio comparativo",
        "ro": "E. Protocol de testare comparativa",
        "cs": "E. Protokol srovnaваciho testu",
    },
    "man_E_matrix": {
        "es": "E1. Define la matriz de puntos: 3 presiones de descarga (2/4/6 bar g). En cada punto: mantener estable 3-5 minutos y grabar log.",
        "en": "E1. Define the test matrix: 3 discharge pressures (2/4/6 bar g). At each point: maintain stable for 3-5 minutes and record log.",
        "de": "E1. Testmatrix definieren: 3 Ausgangsdrucke (2/4/6 bar g). An jedem Punkt: 3-5 Minuten stabil halten und Log aufzeichnen.",
        "fr": "E1. Definir la matrice de points: 3 pressions de refoulement (2/4/6 bar g). A chaque point: maintenir stable 3-5 minutes et enregistrer le log.",
        "pt": "E1. Define a matriz de pontos: 3 pressoes de descarga (2/4/6 bar g). Em cada ponto: manter estavle 3-5 minutos e gravar log.",
        "ro": "E1. Definiti matricea de puncte: 3 presiuni de refulare (2/4/6 bar g). La fiecare punct: mentineti stabila 3-5 minute si inregistrati log.",
        "cs": "E1. Definujte matici bodu: 3 vytlacne tlaky (2/4/6 bar g). V kazdem bodu: udrzovat stabilni 3-5 minut a zaznamenat log.",
    },
    "man_E_record": {
        "es": "E2. Para cada punto (Run), registra SIEMPRE:",
        "en": "E2. For each point (Run), ALWAYS record:",
        "de": "E2. Fur jeden Punkt (Run) IMMER aufzeichnen:",
        "fr": "E2. Pour chaque point (Run), TOUJOURS enregistrer:",
        "pt": "E2. Para cada ponto (Run), registar SEMPRE:",
        "ro": "E2. Pentru fiecare punct (Run), inregistrati INTOTDEAUNA:",
        "cs": "E2. Pro kazdy bod (Run) VZDY zaznamenat:",
    },
    "man_E_record_items": {
        "es": ["Modelo exacto bomba A y B",
               "Presion de descarga (bar g) — manometro en linea de liquido",
               "Presion de aire en la bomba (bar g) — lo mas cerca posible del inlet",
               "Presion en el sensor FTMg (bar g)",
               "Caudal de liquido (L/min)",
               "Ciclos/min (emboladas/min)",
               "Caudal de aire (Nl/min ideal para comparativa universal)",
               "Duracion, estabilidad, incidencias"],
        "en": ["Exact model of pump A and B",
               "Discharge pressure (bar g) — gauge on liquid line",
               "Air pressure at pump (bar g) — as close as possible to pump inlet",
               "Pressure at FTMg sensor (bar g)",
               "Liquid flow rate (L/min)",
               "Strokes/min",
               "Air flow rate (Nl/min ideal for universal comparison)",
               "Duration, stability, incidents"],
        "de": ["Genaues Modell von Pumpe A und B",
               "Ausgangsdruck (bar g) — Manometer in der Flussigkeitsleitung",
               "Luftdruck an der Pumpe (bar g) — so nah wie moglich am Pumpeneinlass",
               "Druck am FTMg-Sensor (bar g)",
               "Flussigkeitsdurchfluss (L/min)",
               "Hube/min",
               "Luftdurchfluss (Nl/min ideal fur universellen Vergleich)",
               "Dauer, Stabilitat, Vorkommnisse"],
        "fr": ["Modele exact de la pompe A et B",
               "Pression de refoulement (bar g) — manometre sur ligne liquide",
               "Pression d'air a la pompe (bar g) — le plus pres possible de l'entree de la pompe",
               "Pression au capteur FTMg (bar g)",
               "Debit liquide (L/min)",
               "Cycles/min",
               "Debit d'air (Nl/min ideal pour comparaison universelle)",
               "Duree, stabilite, incidents"],
        "pt": ["Modelo exato da bomba A e B",
               "Pressao de descarga (bar g) — manometro na linha de liquido",
               "Pressao de ar na bomba (bar g) — o mais proximo possivel da entrada da bomba",
               "Pressao no sensor FTMg (bar g)",
               "Caudal de liquido (L/min)",
               "Ciclos/min",
               "Caudal de ar (Nl/min ideal para comparacao universal)",
               "Duracao, estabilidade, incidencias"],
        "ro": ["Modelul exact al pompei A si B",
               "Presiunea de refulare (bar g) — manometru pe linia de lichid",
               "Presiunea aerului la pompa (bar g) — cat mai aproape de intrarea pompei",
               "Presiunea la senzorul FTMg (bar g)",
               "Debitul de lichid (L/min)",
               "Cicluri/min",
               "Debitul de aer (Nl/min ideal pentru comparatie universala)",
               "Durata, stabilitate, incidente"],
        "cs": ["Presny model cerpadla A a B",
               "Vytlacny tlak (bar g) — manometr na potrubi kapaliny",
               "Tlak vzduchu u cerpadla (bar g) — co nejblize vstupu cerpadla",
               "Tlak u senzoru FTMg (bar g)",
               "Prutok kapaliny (L/min)",
               "Zdvihy/min",
               "Prutok vzduchu (Nl/min ideal pro univerzalni porovnani)",
               "Delka, stabilita, incidenty"],
    },
    "man_E_kpi": {
        "es": "E3. KPIs irrefutables (los 3 que debe usar):",
        "en": "E3. Irrefutable KPIs (the 3 you must use):",
        "de": "E3. Unwiderlegbare KPIs (die 3 die Sie verwenden mussen):",
        "fr": "E3. KPI irrefutables (les 3 a utiliser):",
        "pt": "E3. KPIs irrefutaveis (os 3 que deve usar):",
        "ro": "E3. KPI irefutabili (cei 3 pe care trebuie sa ii folositi):",
        "cs": "E3. Nevyvratitelne KPI (3 ktere musíte pouzit):",
    },
    "man_kpi_items": {
        "es": ["L/embolada = Caudal liquido (L/min) / Ciclos/min",
               "Nl/embolada = Caudal aire (Nl/min) / Ciclos/min",
               "Aire especifico (Nl/L) = Caudal aire / Caudal liquido  ← KPI principal de esta app"],
        "en": ["L/stroke = Liquid flow (L/min) / Strokes/min",
               "Nl/stroke = Air flow (Nl/min) / Strokes/min",
               "Specific air (Nl/L) = Air flow / Liquid flow  ← Main KPI of this app"],
        "de": ["L/Hub = Flussigkeitsdurchfluss (L/min) / Hube/min",
               "Nl/Hub = Luftdurchfluss (Nl/min) / Hube/min",
               "Spezifische Luft (Nl/L) = Luftdurchfluss / Flussigkeitsdurchfluss  <- Haupt-KPI dieser App"],
        "fr": ["L/cycle = Debit liquide (L/min) / Cycles/min",
               "Nl/cycle = Debit air (Nl/min) / Cycles/min",
               "Air specifique (Nl/L) = Debit air / Debit liquide  <- KPI principal de cette app"],
        "pt": ["L/ciclo = Caudal liquido (L/min) / Ciclos/min",
               "Nl/ciclo = Caudal ar (Nl/min) / Ciclos/min",
               "Ar especifico (Nl/L) = Caudal ar / Caudal liquido  <- KPI principal desta app"],
        "ro": ["L/ciclu = Debit lichid (L/min) / Cicluri/min",
               "Nl/ciclu = Debit aer (Nl/min) / Cicluri/min",
               "Aer specific (Nl/L) = Debit aer / Debit lichid  <- KPI principal al acestei aplicatii"],
        "cs": ["L/zdvih = Prutok kapaliny (L/min) / Zdvihy/min",
               "Nl/zdvih = Prutok vzduchu (Nl/min) / Zdvihy/min",
               "Merny vzduch (Nl/L) = Prutok vzduchu / Prutok kapaliny  <- Hlavni KPI teto aplikace"],
    },
    "man_F_title": {
        "es": "F. Reglas de oro — para resultados defendibles",
        "en": "F. Golden rules — for defensible results",
        "de": "F. Goldene Regeln — fur vertretbare Ergebnisse",
        "fr": "F. Regles d'or — pour des resultats defensables",
        "pt": "F. Regras de ouro — para resultados defendiveis",
        "ro": "F. Reguli de aur — pentru rezultate justificabile",
        "cs": "F. Zlata pravidla — pro obhajitelne vysledky",
    },
    "man_F_items": {
        "es": ["NO compares si no tienes modelo exacto de ambas bombas y configuracion",
               "NO compares si no tienes presion de descarga documentada y estable",
               "NO compares si no tienes ciclos/min (sin eso no existe 'aire por embolada')",
               "Usa mismo criterio: o comparas a misma presion de descarga (recomendado) o a mismo caudal de liquido — pero no cambies el criterio entre bombas",
               "El FTMg NO mide ciclos/min — hay que medirlos externamente (conteo manual 60s o sensor externo)"],
        "en": ["DO NOT compare if you do not have the exact model of both pumps and configuration",
               "DO NOT compare if you do not have documented and stable discharge pressure",
               "DO NOT compare if you do not have strokes/min (without it 'air per stroke' does not exist)",
               "Use the same criterion: either compare at the same discharge pressure (recommended) or at the same liquid flow — but do not change the criterion between pumps",
               "The FTMg does NOT measure strokes/min — these must be measured externally (manual count 60s or external sensor)"],
        "de": ["NICHT vergleichen wenn nicht das genaue Modell beider Pumpen und Konfiguration vorliegt",
               "NICHT vergleichen wenn kein dokumentierter und stabiler Ausgangsdruck vorliegt",
               "NICHT vergleichen wenn keine Hube/min vorliegen (ohne diese gibt es keinen 'Luft pro Hub')",
               "Gleiches Kriterium verwenden: entweder gleicher Ausgangsdruck (empfohlen) oder gleicher Flussigkeitsdurchfluss — aber Kriterium nicht zwischen Pumpen wechseln",
               "Der FTMg misst KEINE Hube/min — diese mussen extern gemessen werden (manuelles Zahlen 60s oder externer Sensor)"],
        "fr": ["NE PAS comparer si vous n'avez pas le modele exact des deux pompes et la configuration",
               "NE PAS comparer si vous n'avez pas de pression de refoulement documentee et stable",
               "NE PAS comparer si vous n'avez pas de cycles/min (sans cela l'air par cycle n'existe pas)",
               "Utiliser le meme critere: soit comparer a meme pression de refoulement (recommande) soit a meme debit liquide — mais ne pas changer de critere entre les pompes",
               "Le FTMg NE mesure PAS les cycles/min — ils doivent etre mesures en externe (comptage manuel 60s ou capteur externe)"],
        "pt": ["NAO comparar se nao tiver o modelo exato de ambas as bombas e configuracao",
               "NAO comparar se nao tiver pressao de descarga documentada e estavle",
               "NAO comparar se nao tiver ciclos/min (sem isso nao existe 'ar por ciclo')",
               "Usar o mesmo criterio: ou comparar a mesma pressao de descarga (recomendado) ou ao mesmo caudal de liquido — mas nao mudar o criterio entre bombas",
               "O FTMg NAO mede ciclos/min — estes devem ser medidos externamente (contagem manual 60s ou sensor externo)"],
        "ro": ["NU comparati daca nu aveti modelul exact al ambelor pompe si configuratia",
               "NU comparati daca nu aveti presiunea de refulare documentata si stabila",
               "NU comparati daca nu aveti cicluri/min (fara acestea 'aerul per ciclu' nu exista)",
               "Folositi acelasi criteriu: fie comparati la aceeasi presiune de refulare (recomandat) fie la acelasi debit de lichid — dar nu schimbati criteriul intre pompe",
               "FTMg NU masoara cicluri/min — acestea trebuie masurate extern (numarare manuala 60s sau senzor extern)"],
        "cs": ["NEPOROVNAVEJTE pokud nemáte presny model obou cerpadel a konfiguraci",
               "NEPOROVNAVEJTE pokud nemáte zdokumentovany a stabilni vytlacny tlak",
               "NEPOROVNAVEJTE pokud nemáte zdvihy/min (bez toho 'vzduch na zdvih' neexistuje)",
               "Pouzijte stejne kriterium: bud porovnavejte pri stejnem vytlacnem tlaku (doporuceno) nebo pri stejnem prutoku kapaliny — ale nemeňte kriterium mezi cerpadly",
               "FTMg NEMERI zdvihy/min — ty musi byt mereny externě (rucni pocitani 60s nebo externi senzor)"],
    },
    "man_cycles_title": {
        "es": "Como medir ciclos/min correctamente",
        "en": "How to measure strokes/min correctly",
        "de": "Hube/min korrekt messen",
        "fr": "Comment mesurer les cycles/min correctement",
        "pt": "Como medir ciclos/min corretamente",
        "ro": "Cum se masoara cicluri/min corect",
        "cs": "Jak spravne merit zdvihy/min",
    },
    "man_cycles_def": {
        "es": "Un ciclo = un movimiento completo de ida y vuelta de los diafragmas. Error comun: contar golpes individuales (media carrera). Un ciclo = dos medias carreras.",
        "en": "One stroke = one complete back-and-forth movement of the diaphragms. Common error: counting individual beats (half stroke). One stroke = two half-strokes.",
        "de": "Ein Hub = eine vollstandige Hin-und-Her-Bewegung der Membranen. Haufiger Fehler: einzelne Schlage zahlen (halber Hub). Ein Hub = zwei halbe Hube.",
        "fr": "Un cycle = un mouvement complet aller-retour des membranes. Erreur courante: compter les coups individuels (demi-course). Un cycle = deux demi-courses.",
        "pt": "Um ciclo = um movimento completo de ida e volta dos diafragmas. Erro comum: contar batidas individuais (meia carreira). Um ciclo = duas meias carreiras.",
        "ro": "Un ciclu = o miscare completa dus-intors a membranelor. Eroare frecventa: numararea batailor individuale (semi-cursa). Un ciclu = doua semi-curse.",
        "cs": "Jeden zdvih = jeden uplny pohyb membrán tam a zpet. Bezna chyba: pocitani jednotlivych uderu (pul zdvihu). Jeden zdvih = dva pul zdvihy.",
    },
    "man_cycles_method": {
        "es": "Metodo recomendado — Conteo manual cronometrado: espera a que la bomba este estable, usa cronometro, cuenta ciclos completos durante 60 segundos (minimo 30 s), repite 2-3 veces y usa la media. Formula: Ciclos/min = ciclos contados / tiempo(s) x 60",
        "en": "Recommended method — Timed manual count: wait for the pump to stabilise, use a stopwatch, count complete strokes for 60 seconds (minimum 30 s), repeat 2-3 times and use the average. Formula: Strokes/min = counted strokes / time(s) x 60",
        "de": "Empfohlene Methode — Zeitgesteuertes manuelles Zahlen: Warten bis die Pumpe stabil ist, Stoppuhr verwenden, vollstandige Hube 60 Sekunden zahlen (mindestens 30 s), 2-3 mal wiederholen und Durchschnitt nehmen. Formel: Hube/min = gezahlte Hube / Zeit(s) x 60",
        "fr": "Methode recommandee — Comptage manuel chronoetre: attendre que la pompe soit stable, utiliser un chronometre, compter les cycles complets pendant 60 secondes (minimum 30 s), repeter 2-3 fois et prendre la moyenne. Formule: Cycles/min = cycles comptes / temps(s) x 60",
        "pt": "Metodo recomendado — Contagem manual cronometrada: aguardar que a bomba esteja estavle, usar cronometro, contar ciclos completos durante 60 segundos (minimo 30 s), repetir 2-3 vezes e usar a media. Formula: Ciclos/min = ciclos contados / tempo(s) x 60",
        "ro": "Metoda recomandata — Numarare manuala cronometrata: asteptati ca pompa sa se stabilizeze, folositi un cronometru, numarati cicluri complete timp de 60 secunde (minimum 30 s), repetati de 2-3 ori si folositi media. Formula: Cicluri/min = cicluri numarate / timp(s) x 60",
        "cs": "Doporucena metoda — Casovane rucni pocitani: pockeite az se cerpadlo stabilizuje, pouzijte stopky, pocitejte uplne zdvihy po dobu 60 sekund (minimum 30 s), opakujte 2-3 krat a pouzijte prumer. Vzorec: Zdvihy/min = pocitane zdvihy / cas(s) x 60",
    },
    # ── Column headers ────────────────────────────────────────────────────────
    "col_pump":    {"es":"Bomba",             "en":"Pump",              "de":"Pumpe",          "fr":"Pompe",           "pt":"Bomba",              "ro":"Pompa",          "cs":"Cerpadlo"},
    "col_flow":    {"es":"Caudal (LPM)",      "en":"Flow (LPM)",        "de":"Durchfluss (LPM)","fr":"Debit (LPM)",    "pt":"Caudal (LPM)",       "ro":"Debit (LPM)",    "cs":"Prutok (LPM)"},
    "col_pdesc":   {"es":"P. descarga (bar)", "en":"Disch. press. (bar)","de":"Ausg.druck (bar)","fr":"P. refoul. (bar)","pt":"P. desc. (bar)",   "ro":"P. ref. (bar)",  "cs":"Vyt. tlak (bar)"},
    "col_pair":    {"es":"P. aire (bar)",     "en":"Air press. (bar)",  "de":"Luftdruck (bar)", "fr":"P. air (bar)",   "pt":"P. ar (bar)",        "ro":"P. aer (bar)",   "cs":"Tlak vzd. (bar)"},
    "col_cons":    {"es":"Consumo (Nm3/h)",   "en":"Consumption (Nm3/h)","de":"Verbrauch (Nm3/h)","fr":"Conso. (Nm3/h)","pt":"Consumo (Nm3/h)", "ro":"Consum (Nm3/h)", "cs":"Spotreba (Nm3/h)"},
    "col_lpc":     {"es":"L/ciclo",           "en":"L/stroke",          "de":"L/Hub",           "fr":"L/cycle",        "pt":"L/ciclo",            "ro":"L/ciclu",        "cs":"L/zdvih"},
    "col_cpm":     {"es":"Ciclos/min",        "en":"Strokes/min",       "de":"Hube/min",        "fr":"Cycles/min",     "pt":"Ciclos/min",         "ro":"Cicluri/min",    "cs":"Zdvihy/min"},
    "col_aesp":    {"es":"Aire esp. (Nm3/m3)","en":"Spec. air (Nm3/m3)","de":"Spez. Luft (Nm3/m3)","fr":"Air spec. (Nm3/m3)","pt":"Ar esp. (Nm3/m3)","ro":"Aer spec. (Nm3/m3)","cs":"Merny vzd. (Nm3/m3)"},
    "col_aciclo":  {"es":"Aire/ciclo",        "en":"Air/stroke",        "de":"Luft/Hub",        "fr":"Air/cycle",      "pt":"Ar/ciclo",           "ro":"Aer/ciclu",      "cs":"Vzduch/zdvih"},
    "col_dcons":   {"es":"Delta Consumo",     "en":"Delta Consumption", "de":"Delta Verbrauch", "fr":"Delta Conso.",   "pt":"Delta Consumo",      "ro":"Delta Consum",   "cs":"Delta Spotreba"},
    "col_daesp":   {"es":"Delta Aire esp. %", "en":"Delta Spec. air %", "de":"Delta Spez. Luft %","fr":"Delta Air spec. %","pt":"Delta Ar esp. %","ro":"Delta Aer spec. %","cs":"Delta Merny vzd. %"},
    # ── Excel sheet names ─────────────────────────────────────────────────────
    "xl_sheet1":   {"es":"Comparativa",       "en":"Comparison",        "de":"Vergleich",       "fr":"Comparatif",     "pt":"Comparativa",        "ro":"Comparativ",     "cs":"Porovnani"},
    "xl_sheet2":   {"es":"Datos entrada",     "en":"Input data",        "de":"Eingabedaten",    "fr":"Donnees entree", "pt":"Dados entrada",      "ro":"Date intrare",   "cs":"Vstupni data"},
    "xl_sheet3":   {"es":"KPIs calculados",   "en":"Calculated KPIs",   "de":"Berechnete KPIs", "fr":"KPIs calcules",  "pt":"KPIs calculados",    "ro":"KPI calculati",  "cs":"Vypoctene KPI"},
    "xl_sheet4":   {"es":"Metodologia",       "en":"Methodology",       "de":"Methodik",        "fr":"Methodologie",   "pt":"Metodologia",        "ro":"Metodologie",    "cs":"Metodika"},
    "xl_title1":   {"es":"COMPARATIVA AODD",  "en":"AODD COMPARISON",   "de":"AODD VERGLEICH",  "fr":"COMPARATIF AODD","pt":"COMPARATIVA AODD",  "ro":"COMPARATIV AODD","cs":"POROVNANI AODD"},
    "xl_title2":   {"es":"DATOS DE ENTRADA",  "en":"INPUT DATA",        "de":"EINGABEDATEN",    "fr":"DONNEES ENTREE", "pt":"DADOS ENTRADA",      "ro":"DATE INTRARE",   "cs":"VSTUPNI DATA"},
    "xl_title3":   {"es":"KPIs CALCULADOS",   "en":"CALCULATED KPIs",   "de":"BERECHNETE KPIs", "fr":"KPIs CALCULES",  "pt":"KPIs CALCULADOS",    "ro":"KPI CALCULATI",  "cs":"VYPOCTENE KPI"},
    "xl_method":   {"es":"METODOLOGIA",       "en":"METHODOLOGY",       "de":"METHODIK",        "fr":"METHODOLOGIE",   "pt":"METODOLOGIA",        "ro":"METODOLOGIE",    "cs":"METODIKA"},
    "xl_formulas": {"es":"Formulas utilizadas:","en":"Formulas used:",   "de":"Verwendete Formeln:","fr":"Formules utilisees:","pt":"Formulas utilizadas:","ro":"Formule utilizate:","cs":"Pouzite vzorce:"},
    "xl_units":    {"es":"Unidades:",         "en":"Units:",            "de":"Einheiten:",      "fr":"Unites:",        "pt":"Unidades:",           "ro":"Unitati:",       "cs":"Jednotky:"},
    "xl_sources":  {"es":"Fuentes datos ejemplo:","en":"Example data sources:","de":"Beispieldatenquellen:","fr":"Sources donnees exemple:","pt":"Fontes dados exemplo:","ro":"Surse date exemplu:","cs":"Zdroje prikladu:"},
    "xl_warning":  {"es":"AVISO: Valores de ejemplo ilustrativos. Verificar con medicion real.",
                    "en":"WARNING: Example values are illustrative. Verify with real measurement.",
                    "de":"HINWEIS: Beispielwerte sind illustrativ. Mit Echtmessung verifizieren.",
                    "fr":"AVERTISSEMENT: Valeurs exemples illustratives. Verifier avec mesure reelle.",
                    "pt":"AVISO: Valores de exemplo ilustrativos. Verificar com medicao real.",
                    "ro":"AVERTISMENT: Valorile exemplu sunt ilustrative. Verificati cu masurare reala.",
                    "cs":"UPOZORNENI: Prikladove hodnoty jsou ilustrativni. Overte skutecnym merenim."},
}

def t(key, lang):
    d = T.get(key, {})
    return d.get(lang, d.get("en", key))

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AODD Pump Comparator",
    page_icon="pump",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.main-header {
    background: linear-gradient(135deg, #0f2027 0%, #1a3a4a 50%, #0d3b55 100%);
    padding: 1.5rem 2rem; border-radius: 12px; margin-bottom: 1rem;
    border-left: 5px solid #00d4ff;
}
.main-header h1 { color: #fff; font-family: 'IBM Plex Mono', monospace; font-size: 1.7rem; margin: 0 0 0.3rem 0; }
.main-header p { color: #8ecfdf; font-size: 0.9rem; margin: 0; }
.header-badge {
    display: inline-block; background: rgba(0,212,255,0.15); color: #00d4ff;
    font-family: 'IBM Plex Mono', monospace; font-size: 0.68rem;
    padding: 2px 10px; border-radius: 20px; border: 1px solid rgba(0,212,255,0.3);
    margin-bottom: 0.6rem; letter-spacing: 1px;
}
.kpi-card {
    background: linear-gradient(145deg, #1e2d3d, #162535);
    border: 1px solid #1e3a4f; border-radius: 10px; padding: 1rem 1.2rem;
    text-align: center; border-top: 3px solid #00d4ff;
}
.kpi-label { color: #7a9db5; font-size: 0.7rem; font-family: 'IBM Plex Mono', monospace;
             text-transform: uppercase; letter-spacing: 1px; margin-bottom: 0.4rem; }
.kpi-value { color: #e8f4f8; font-size: 1.5rem; font-weight: 700; font-family: 'IBM Plex Mono', monospace; }
.kpi-unit  { color: #00d4ff; font-size: 0.72rem; font-family: 'IBM Plex Mono', monospace; margin-top: 0.2rem; }
.pump-card {
    background: linear-gradient(145deg, #1a2c3d, #142230);
    border: 1px solid #1e3a4f; border-radius: 10px; padding: 1rem 1.1rem; margin-bottom: 0.8rem;
}
.pump-card-title { font-family: 'IBM Plex Mono', monospace; font-weight: 600; font-size: 0.85rem;
                   color: #e8f4f8; margin-bottom: 0.7rem; }
.section-title { font-family: 'IBM Plex Mono', monospace; font-size: 0.73rem; color: #00d4ff;
                 text-transform: uppercase; letter-spacing: 2px; border-bottom: 1px solid #1e3a4f;
                 padding-bottom: 0.4rem; margin-bottom: 1rem; }
.info-box {
    background: rgba(0,212,255,0.05); border: 1px solid rgba(0,212,255,0.2);
    border-left: 4px solid #00d4ff; border-radius: 6px; padding: 0.7rem 1rem;
    font-size: 0.84rem; color: #b0cad6; margin: 0.5rem 0;
}
.warn-box {
    background: rgba(255,190,0,0.05); border: 1px solid rgba(255,190,0,0.2);
    border-left: 4px solid #ffbe00; border-radius: 6px; padding: 0.7rem 1rem;
    font-size: 0.84rem; color: #d4b87a; margin: 0.5rem 0;
}
.lang-btn button { font-family: 'IBM Plex Mono', monospace !important; font-size: 0.72rem !important; }
div[data-testid="stSidebar"] { background: #0d1c28; border-right: 1px solid #1e3a4f; }
div[data-testid="stSidebar"] label { color: #b0cad6 !important; }
div[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #006a8e, #004d6b) !important;
    color: white !important; border: 1px solid #00a8cc !important;
    font-family: 'IBM Plex Mono', monospace !important; font-size: 0.8rem !important;
    border-radius: 6px !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
IMG_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAoHCAkIBgoJCAkMCwoMDxoRDw4ODx8WGBMaJSEnJiQhJCMpLjsyKSw4LCMkM0Y0OD0/QkNCKDFITUhATTtBQj//2wBDAQsMDA8NDx4RER4/KiQqPz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz8/Pz//wgARCASwBLADAREAAhEBAxEB/8QAHAABAAIDAQEBAAAAAAAAAAAAAAEDAgQFBgcI/8QAGgEBAAMBAQEAAAAAAAAAAAAAAAECAwQFBv/aAAwDAQACEAMQAAAA+zAAAAAAAAGEsZCYCQTAACSQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACACAAAQJRIZQygAAAAAAAAAAAAAAAAMZYkgEwAAkEgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgEAAECUmUAAAAAAAAAAAAAAAABiRICRAASSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACCAACDKUwAAAAAAAAAAAAAAAAGJEhMAJBIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAAIBJIAAAAAAAAAAAAAAAAIIABJIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABBBJIAAAAAAAAAAAAAAAAIIAMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQCQAAAAAAAAAAAAAAAAQACQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADWprxeb0Obj14Ra2ab+vN2unztrTIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQcnHr8zwe1yOb0JImttqzMV1vlLudPmeu9DwrrUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFcTy8url4dWvTXCLQCYJCEohIlASAYiYJlAom3H4fXsM75X9nmZdHFlfC2Ypw7aOT1Ix6dzbl2tOSIQCSJiQkgASBMIJiQW2puXx6m3L0dcJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABByMuryPB6/C4vTim10xdauVoEoiJmYAkgkhIlGMTMxgmJjFFcW01ejeL4ts9Xm2d3jeYtnyVa5bMxuJ70Xs4vYz5vRwia8riEwCSBCZjGJxi0ERMkzEAGMTXW1au1vy+n7fJ9X2+VsWoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABq1v4Dzvd8753sW0vfZu9GHQ35d7bmttnCRJnNcYmZjJCUoymMlYicVhlMRKDyfB6nMx6OjNt3r83zff4HjrV5yKpjEtNk9FE+rtO753vWY9lqM1RMxballqSiEyESQCy1RKMYtBRTXnYdenhvTS1UM9+f23p/P+r7PMkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHIz6PmfkfS0cvbdFrDY0z298L755zTOayCEgjOaxEpSiSImZjFIxmcCLV8rxd3Qm+7fLm+v8ALfP740lExUVmJBbD3ifR16N3y/o8jGkjImYyQlEJIJIBESEsazCcYmimtOeldLYVnCYqvT0PoeN9E9PwLJgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcfPo+V+P8AT5YdWynOLSSgRExWUwiZmBBJBEShMxCUBCYlRatGmS9apr3l93p8/wAL6nznJNdGsa5UVgJ6lX0izpeb9Dbl1aOG8RMRJMRMQJIRIiSEJCGMSJAIiYhhEYTNVqdv0PH+mep89nMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADTrf5D5H0+HP27a2VL5Qv1zvvlrZba+O0GVoEEiSEzEoxiwytUIYpiY1tKa2mONqxW3oVt3u8bxXd42qaaNM0zVRVEgWJ+lHY4fb3Of0NamlVbZJAkEoAiJiUhBMhERNdb4UtKMazXS9cKyrTP1fqeB9B9DxQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIPlvB7nnfN9nYz1spayt79KdPs4+p0cWzpjoYdXK5u3Vz3gppaZSEJWWrhW0EaUxmsxOdL4S1tMtC+WFqWU19JW/R7/D8x3eRrmgjmnOidIpRiZJmH0aXd4Pc3MO8a1b5ylAyRIkBjEyjKYymIShRnrr469Dfl6XTx11vzufr0cOiml6KTXaKd8PqXsfLd7fkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHn8+n5R4/01uO+xXS6JypfY1z7Pb53ou7yc5ryefu4Hn+pp8/Rfrns9PPxeTugmYkztGFZAEygwmurrjo6ZRarLX1dNN/r8zgen8/zzTOYaENKGuhMyWn02Z6/l/R2Z9OnW9MSTKJMpIACCUwiSU45Xxyv0Onl6XTx7+3NhFudh1c3m7dTDailqyrTPo9vmfWfW+azmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTE/G/M+k0uTu24vtLZxM1vkjf35et1cN98tDHq4/D6Gtz7YUnf6ubb7ubj8/bjBKUEggSQYWrraZaumNCNTm29LXXp2pzvb+S89NNKWnDThqypBJ1ofR69G95X0la3Lz1iQkEkwASQAgnO9sTki++e9vz7uvPjE6mW+hh1a2OtdZrhhaKdcfZ+p4Hve7yAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8/wCb0vBeZ7uzTTdm9yc62zi2NJiFk1ymuFbU46Y0tFRN3Tlh0VorbGEyImUoTBJGM1o1z1r5UTXSw16db9+mu53+L8/9HxOejXhrlMqyIWy+inf8/wB/Pm7+ZTXUCZRCZRNZJEIlMEwzvEospbOGScjOYzmsowreutq6Wria4VzGF6V9HL9j9n5TfvmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMD4X5n0dPN2bq+yvbW1kWyiYrNeV8aWymuVoQwpbDO0QiZq0rG1aKkxMiCEhMsbVr0zovnVamvWdzSvocOjcVep89846OLmRFMxgYwvl7A9fxeznxetSngZXhMQBJAmJRMTCJRNkrts5qtrNyc5ImIlWZhlZMxjE1Vmqs1ortWnXH2np+B9A7fKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHnqb/G/K+i2Ka703ure2L2ROVbTE1ZXxpa/TPb6MMYnVw3ow0xpMJpmtGlar0qvSZSTMARMYyrtWu1MJrhMdDSnqts9fg9BSdvr87keh4vJ25pR14t0MerLh9mMOzG0cSc9StoiRESBCCUIIrOUTbeJtW/O2wm+ZslMTETjS2Zs657OuNNb6eO2vS1MK7Vq0y6HZ5n2X1PnQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB4LDu+a+b7e1Gm7N7K3tre2LZRMxbDKyJ3ujn7HX5+U14nH6XP5OqrG0QrvXV2xz68bOnmyVzTfMZzEERKREoyLUYm1aMq6UcnZjjsRs3y2L469daY0St35tTp48L1qMSImCUZzGc1qTr1tVFtHLSnDbKJ2K2289Nha+y1OUTjW0UtdenR6Obo78mpTbmc/ZpY6a9FVoq0yr6OT7x6/wAtnMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfNef0fBef7Gy03FrKaWxe2s5pmtq8rzDpdXL6Tv8geb4fY5nJ1Uc98Jiq9Y7eT6B6vhev6PLzVAkAAAAxT5zLv8xh6Wph1Y47456RE5Wp1+jh9N2eT3OjzpAAAAAIPPTt8a4va1+bo3cdtmm2ynYmboZLRW2Gdti9Ox18XT34tWm/F5e/nY60Umqa0650b8v3b1vltq1QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmfN6Xg+H19mNNxaymltbWxbNM1thndWen1cno+/ycYnz/F6vJ5OunC+KKtKbPocP2f2fkc0AAAAAAAYnNr0aNdoNucuhbHYmgAAAAAAA+PZen4/g9Pcw6Nmuu0tsJuTnExS2FLX6U7HVxdbo4dSm/D5fQ0MddXOa5rVtlrb8v3j1fmNq1AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmXN6XheH19iNNta2mlsXsibIma2rzvMT0unl9F3eSPP8Xq8vj69fC8WirSm33+f9m9n5KQAAAAAAAAAAAAAAAAAD5Jn6PifO9bbx6Nqt9pfYTcnOJilsKWvvTsdfF1+jh06b8Lk9DQx11s7VWrXtlr78v3j1fltm1QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmPN6fheH1tiNNxa2ullb2xOcTlW1ed5rPT6+T0Xb5Q87xerzOTr18L4zFelNz0fO+y+v8nIAAAAAAAAAAAAAAAAAB8ky9HxPn+tt477VddpbYTctZDGl687X3p2Ovi6/Twa1NuByelzsNtfNXatW2VHRy/ePV+W2bVAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+Y83p+F4fXvX3FraaWVvbE2RbKtqsrzDp9XL6Tu8jFPA4/U5PH20YWxmKtKb/o+b9j9f5SQAAAAAAAAAAAAAAAAAD5Jl6XifP9Xax32q67S2ym1ayEVtXne+9e118HX6eDUz38/yelzufbXpNd6V65Ub8v3n1vlti1QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmPN6fhOH17lt1eymltb2xayJyravK8xPT6uT0Pb5UzHn+L1eXydmvjbBFd6dD0vN+xev8AKSAAAAAAAAAAAAAAAAAAfJMfT8VwepsZb7ddNmL7KbVs4lS2FLXWr2uzg7HTwatN/O8fp8/n116TXalW2NO/N969b5a+1QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmXN6fguH17ptuL2Z6XVvbFrUzW1WN5rPV7OT0Pb5WFbcLk9LmcvZq89sJiu+fS9LzvsXr/KSAAAAAAAAAAAAAAAAAAfI8fT8Vwers5a7MX2YvsRe5NkSrfDO2xenY6+Lr9HBq024HJ6XOw2185qtSnfGrbm+9+t8tfaoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHEz6PKcnq8vk9Dk5dOrMY9GGWWtueu1ZYtnWasdModbr4+72eYTwuP0ubydmpz31r50bYdHt5fb9Xmb9s7pjb15+ttxAAAAAAAAAAAAAADCJwTrRfynn+553Ls43N00xFG/PdF9mmt61yVbRS2xevY6uLpbcevXXhcvo8/n21spo1z0+nlq15/p+/JbEoXXy9h3eNIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPE4d/zvi9mcunocfbnMadYp7uOzPS6muynOJyhMLppvb8/Q35KqX5fN3afN0a+V+dvhodPJu3narv3r8/V05t3bl9Z3eLbNQAAAAAAAAAAAABrRf8+eV9HoWp0Zel5fUypanK+rRrdnFbFtmumwtYm1ExNlo2b53Wyprpo4dGrlpTWdbXLjd3n1Xy9Ly+jspxpbNn9b9743atQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcPPo8zzehzOfv5WHbbTTKlrrVoy0WrmZJvtnTnrZat18r75ZzFNNKqXqrcJYzGFq0yzLU5Qu0y6W3J6Hr8r0XX5oAAAAAAAAAAAAqifLYeh824vZ18tOnj07Gd9nWm7tz6WO+CbZrlEyWWrffOy1M7ViJrraSuLYVnCLYTFOmVJZF7DYvnq1vs6c/X05ut0ed6Xo4gAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8k4fd4XN3dLHq2MtYiZlEMYmQmuloELL1mYERMRIFlq2XolIRnasmU173Z5ujl0TOP0L1fmwAAAAAAAAAAAB865vX5WPZXW/C5Ozr572Z33t+e/TLXprt6Y7OuNdbUV1zmtlq5K2WrhFpRJKuEWhbKa4xMROUwIMItlavKz1u1876j6XhAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQfG/O+i5mHV0sum+mkxIxrOMTlME4UtjE2XplaIrOdoiEJxhFZztFumcxMSQmYmY2dMut08NcW161+i+r82AAAAAAAAAAABB815fa1c+imL+f4+zrZdFmWl+me/0csyhN9skqK6WzRJEzNZJQlETlMZTUnCJxixGcxMxKIOZTaNOD6r6XgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQfG/O+i5fP1dKnTs0tKUMK3xiZhZetVLZSGUxFUym0ZTFGd8YmTOYytXKUostWDc1w39uXBODL6H6nzoAAAAAAAAAAAA+bcvtUV3orbzPH2dfDqnHbf6OftdnnYLaGPTlMSYxN98rLVrpbO9ZIiYM5qlhW2EWwi2CYhbauSucxbbPk03u1836f6PhgAAAAAAAAAAAAAAAAAAAAAAAAAAAADw3N6fgub0dfDp3qaWRaKzCQTMMYtEIJghCcpIIklCZiZJSiy9ZMazsXpZevV24vpHrfMgAAAAAAAAAAAQfNub2tWmuEX85w9+7jvZW2zvjffKIshnaqREzAxSQExCbZpnNcYsEsIkgTMTMVHIzt9a9T5vrXxAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHncuv51xe1q83bv01rw1kJxrMQRMRJCZipJWUAhJMpmM9IzvWIXXzqz0xTtaZdTs4eThv9b9740AAAAAAAAAAADUi/gOb2N61KFvFeR7fYrrlS3S6+PCJymNnTK62eU1GcxXF7JpCRBjE5TGUxkrJZauMSBBMOfTo8jy9P2n2/kNm1AAAAAAAAAAAAAAAAAAAAAAAAAAAAABxM+rwnJ63F4vT6FNMMdcazJjEwlUTAArMCE2ICITEzaBBsbZ05aDe6ObqdvDyOfo+t+98aAAAAAAAAAAABpxp4rn9TatWhbwvj+92K6WUb/VzZTXFNtqXXpnauauUxKEpQJMYkDJXKWU1lEJkiGMTzc+ryvP0fafb+O2LVAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHFz6fFc3qcDi9PcrpjneYnMiLRWcU41mImIEqzEEJkkESgQiYTsdGUEy29cNvfn0qT9Y9v5AAAAAAAAAAAADSjTwPL7G1al16eA8b6DsU1ypPS6uSZWWpTW+5rgEhKIiYJkhhFs5qhEBMxBikjFOETpU35FX2r2/j7ZqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAME6ddKYvxc+nx3P370V9Fty9bbmk8lz9/geP0s5jImJxicazjWYIiSUJlIiYmMpj0vVx+h35PQb8mzbPcthsTQAAAAAAAAAAADBPOptrL8iu3jcOujPbm010ufpziLJTMZTCYlATEylEzXErrbBPavn9K7fK8tj18bPf1W3H0NceDz9V9q+x6vNumt80sQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPAcfq/M+L19y8ZJtJiaKXyRBEsIYxMFc0xvnE1ypfpeh5Ew3Fuzpl7jTm7NsAAAAAAAAAAAABzqbeVx9DyGenPVqKLV1OX0LYtbfMdCbbMskCIkZQymurWebjtVFvR78f2f1PngAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB4Hj9X5rw+xu6LK3vMotqc+owiEq4iu9dfXCroxqpCIlOp1cOee+zTbZs9dNfrnofOyAAAAAAAAAAADCJ8lh6c2j5553tcGKYIovlz+nj6XN2LRMPWxfpa1RbGJwic0wZq83PXm4a1RPoOrg+w+r84AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPmHB7XieX092021vetlWdLm2xhgjGYotTV359Po5ty7Pk6rM7Jau3PpdPH3eL1dmt+hW/1r2/kN/XnAAAAAAAAAAAHNpvwMu261PO8fqeZ871NW+davnfR8br8fp5xYn0WmPoerDCJwiYiZTjCTkYdHMw1ql1eji+z+v8wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPl/F7Hi+X1N5ayml62UTocm2CK7VpvTV0w098MdM+jF+55/pbnPuMbV8X7Hz/tvF+iylnpX3ns/Mej7fJAAAAAAAAAAAHnsuzTjWyY52fV5f5/6WIIeb9Hyt/m6unz9czGNo6npebsa552qhKc0WTXzfJ3c3DWu0bnTw/bfW+YAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHy/k9fxvJ6e7GlldL4sOfy61TWi1NXTHV2xptT0fH6GOuXXy228drtM4PnvpeJ7vx/oAl6D0fJ9r6nzu1fIAAAAAAAAAAasX+S8fsd+1NXHs5lbT5vq2ROdq8y2XD6eL1nF6mdq1Z32uvl6XVxtKgWK32p5Pi9Ln8+ld69Dr4PtHqfNAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD5fy+t47j9TdaZ00thEubz6a9qaumOtplVbOZn1vJ6O/neL06U1ymMpr867PL9xwewz1iE6V7XZ5nvvU+bymAAAAAAAAABWn4txe55nHp6ee2VNNKael5+nopsvTT59/D+n4nv+D15mKc9KIm3ox2dMuj18gvmls18lxelz+fXC9Oj2ed9m9P5wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfL+X1vH8fqbkaZ1vnLFHJynWtnraZ02zvi3p+fujDbl3yy0y9lW8zEQ8J2+Z6Tj9Lex6IgvHD1p7bq833/p/OZTAAAAAAAAArT8n5/W8nz9uxj0cPTDKttymnew6etjtlNas9PDer4PufN9u0widROprnfM79s+v28N9s75p4/i9TQw0xtTo9nnfZvT+cAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHzPm9XxnF6u3F5i+UxVanIya00qvTdpr6LDt1Mb6V8ufvzWLe7x36XTx62e3j+nht5e3vYdUzFdb0bV5019x2+T7v0vnpAAAAAAABB8px9Px/P3349PGtlRNPQY9WNbbeenfx1ymKs7+M9HxvT8Pq7tdLb01YtyLVsmL5ncmve7vO2r5+P4fT0cNMb06nb5v2P0vnQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB4Dn9HwHD7Oysiy1abZ8eihHTz27eHZpY21NM9PbDXvnZE/Qs9vQd3l6OPT5K+enWfXY9co189dK2uKNA+men836Tr8sAAAAAAAeGp2fMeb1LsOrlKUzXv5dGGuephtt0t7m2dlqauW/jdeToZb92m+UxRW/nonKVkxdad2Kel9DzPI8npaXPpjpTudvlfW/R8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeA5/R8Bw+zsLYC9KLZ8jOepl0dXDr08Z1rU1dsdbTOu0ZQ+r6Y97r8+DxuPVwLY/Rl663rrfzXL6lc2xqm+H1v2fkLrUAAAAAAGmt8M5Pbqx6tGtteaehx6atKU3z1+fe4+o7cN2metnt4bDop05/eWtETp5dHkufprM5WSuluTTVvGnjeLV7nb5X1v0vAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHz/n9HwPD7N8zUiL0pvTLm6elzdepjai1NXXPXvlTekTBP1fTD0PXwbF8vM49Xh5y+taUklHifO+g5+fRMxXSe/wB3j9Dp4MpgQAiZiUTMWTXKa2I9JfCxHzrPv+d8vqYZb6M072PRneNO1K5pVz722j610eZ0ujmqrf5/53s+d34PsfVyU105PN3eM4/Q1KzMLLrJXWYUnXVi1e52+X9c9L58AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfPOb0vCcXr3WUmN8789trj7dTG1Nqa2mevpnTamFq4ky+q6c/pOni2tMeFj1/M7Y/adMrbUzmnhfN+i4uHdKMk6tY5nRyQmUxNYmudq36ZrUrmKJrrKVTH23fyvQXx+D8vuc/Dr0onfz16caads9S1KlYy02LPrG/m+g6uDCLeC8v6LwPT5n3Tp5Mpjj83f4ng9XlZ3mGRZdbaLpnXoptTt9nmfXvT+eAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHzHl9bx/H6dl4pRdF9vk7dTG1KtF6UXpTfOq1cbRBJ9R05vbd3lUV10c9/jyn2/XPYtlZNPFef7/neX0spjKWGVvOdnnbkqa6YopmkTWZjq74aBpK6c1TX71v5e3ePhXB72Ge1MPQ5dNF66k56s1pmmETu1v8ATteL1/b5dVb+I8r6P5t1+V9y0xv0z5HP3+R4fS4eOglOUrbRbZYnVmnQ6eD7N6vzQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+Xcnr+R5PSsvXCtt/n7NXDSmK03rRbOm9arUrtETAvtHvlff9/jVRppZdHxiK/btK7F8c0eS4fZ81yennLOYk5M46++GKMZjl3wvrcW64Z6V0ojVmMkffOjyuO3+U8Ps11vv5bbq+rbPWmlE1otXBG5np9Dth7Xt8vBbxvkfQ/OO/wAf7NFtzbHi8noed5O/g5WTAyTnaM5XXZF3T5/2f1fmQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8t5PX8jyelbeNvn66OfWhWmYpmlN61WpXauNovtGxaEW9Tlb33f5VFNKa6fGmf2KNt3XnmY89x+l5rk9LOVloytGlWnO6eTGYi1OBry7uPRknDbnuvTUhroytX9Ab+X83z9HzfJ6CZ6nN2V2rrzTXtSi1KpjGK7FNfbVj2/Z5+MT5fzPa+f+h431PPfoa5cvn7eHydvCohCZGUs7LZi+7p7cX1n1/lgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8s5PY8jyejvV1x5ujXqqtWma0zWq1a7Vz0rs3rjW0VtJ6Ol/W68116Vxb5Fpy/TeP1d/flmY5eHX5/n7ciy8WWiqscjr4ImKdMfOa83S5uuyJ19+Wy9NVNSM5r996PM+NYetlx919NdmNNaa0WpRNKrVxmLprnS/o8NvX7YScfi7/Deh43vcevo3a+W3L5eriyqiMIDIysztF1mzaPqXrfK9bfkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEHyLi9vhcvobOPVq5qUVTSq0VTXG1djWsxONLWRaTKY72WvWidvbCD5b0ed77zfd3NcLL1189Mtc6Vubj1ZSQ5HX5+N89Dbm4Nsepy9tidbbmx1xoia5jM+29Hn/GOb1Ozx9+xlvN2valE1iaWzFkTVVEuzjv6Wsk6mG/j/Q8f0+XT0sukjUw35NoqRXEYEkyzssmNi8d7fi+nev8vIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB5unR8l4fb3ebu1srUQqmtUxVNZvGxeMKTZW9sW2ottxbfpM1mwv0ywi/zvu8b23k+9faM7VprpStlNen18bSOXh18bXjw2x4u/Fy5r1OXtvW09eajfnqicCxH1vfi+Tc3o9Xk7+jn0YWhamUxNLKWyicLV17U6WW/fx1xrOFLeY9DyOnTo6/L2yVZ3s6MaItz87UWpUiZTKyy20bNo+m+p8x3enhAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEHxfk9nR5fRYb69VExWipFl4tmFL31vtxa9aikacRp3y6ka97Hbe6eWimvz3t8j13k+7fW8oiZm8LMEojZ0x89rnpdPJxenh00dTm7t5bnXw1+nihbBMQ9ptzeJw6+hzdnf5+2y0Y0tNZyraUYSptTUvn0ctu3zbxWYPP8AoeXZXTvcfeIrPB7uCmttumuzTWlGuotOUxZLYu9D0cH031vmJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPO06PjnF7fS5uzTopRVMVwtvGac6X3YvenVo1K1otTV1x1NsPQcvd6Hj7dvp566X8L6Hj+o8z2tnDSZRLGZBMRONJ5nZx8fv4OZ0cc017nL3XHI25tHq4M05mRN89XLo2cOn1fmesEhijG0VzWq9K702c9e3ydKA4nd5tN6ej4fTispec9DzNWaWVvZW+5Ta+musjCYzldZtWj6t6/yfU25gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPkfN63D5PTqx1oiKZiuItszi21XTbW1qNOtappVpTXvnFqxMej5+3s8fXESl4/1PE9D53rbuG0kS5dq7MXvTNSJ0unm5fdw0aZ2YdPYz31YjU35dXbCETKJiu2dNb40v6jzPWsTjaImIljETaKdM9O+eWWnpePrQm0cnt4eZfD0/mevlpVLjdnDz9MKprlE5RN9b72XVgtUiyY2bvZ9nke89HwQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABz40+IcPvbHP1aVYpRWZy2V92l66zpViqaVWpTelN6VXp3uLu7efRVvj0OLrxiZl5X0/F7HD6fQw6APOdfHRem1jvu477VbbNqb/AGckEkqpQYxMSxlNq4I1LPM6c/W4PTzi2M1omKEa2mettjrXytz02s9PWcPfFJz1podfHwNOX1XleznrXO9bNsdfDo8t6HkU2rMTnW2cW3ab3V1piL7T1NMfr3s/I5TAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHhMe/wHF7FGd9ZWsHTjW2l9KsUK1WpVelNqU3phNcZjvcXodfn6q+vk63B3wkcD0PK3eXs6WHTMoNPbPzvVwZVtlXXaw36+WvT6MN3XniUBNdbYpWriLV51dNO1ORtzdDk78otzJpr6ZU6ZUXzqtnXane5O4n0/B3qzbrSro5vMbcfpfK9nPSt2ueGmdl3kejg1GYlOcWzrfarrt02wTty+p+r8v2ujjAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEHxvk9rlcvfoRWlERPUjVS2nFarUqmlV602pVamExkm5bs8fdu476/Vyeh870iYPOen4+5ydvYw6pkOFrlo6c8SmLK26ee29nvhWbNKzMLRExZemUxFowtS3XHhxHDnHqc3ZRetGmdVs6ZpTalc0iY6vP2Yns/P9KazdrnV083itvP9l5PvW60v2yq0zm0+S14tS2UIJmJzi2a21TTcpvYt7vu8b2/f4gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA0ov8T4PoNXPTTiqHUrtTVrKVWpVatV6UzUm1a6LStjDLO2vpj0suj1Pm+pIPOep4+zy9na5uyzSsHOvAERbGsxjbc6K4IytGVoGVq4lcxVMYWzt1x4kRyWPQ5+2+Yu6MqbUqtnVamFq4oxhnz9PsfN9SyJv1yjp5vnu/m+58j6G3Sl+2XM35uLty7Pn+lqa8+l0c1d6InImLWRbcrru129Jry/UfW+YAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHk8+v5hxe1zqxXDp57a6tE0qtnVatMxnFtiL5L4VnGWUxnauVqkX4dHpvL9VBaPP+n41nN193l7rtsxzMdaIImZZFmVtnuxiUzBC1VqYJyRXKi1da+PJivPvj28Oze0iu+evamojXmtU0Q2M9dvl6vR4dVtZ2tsbOnm+Y7+Z7/yPo7tc9jXLnxHIx1mkxMY2rp9HNo9HLhasmcTbXTo136sT9m9n5LKYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+cYel4bk9LnxHRz1ommvbKq1aZrK2zXTKukRbKWVqym+ltmL3UvvZ63Yas9MKWz0rwfS8bHn6fRcvfsb44xbk46K2ilqrRFWzS93dhEs5rlfPmdfneS6/K6KPTcvq583frTGjfLlVrp6YdnLq6l7UXyqmlVq41m/n6Nim2vNNfK3o897otvb82108vyPbz/ofj/R7G2WzrjNo08tOVz702rWitFemXO6uPW1xksi2zTXsV3+yep8xv64AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADRW+XcvreU5+q+mlamrfOq+daL66XU1iul6dmL31vdW9lbVzWqa1KVTW+unb4+7DO2eleF6Xj0YdHqsO2/XLCJ4NZ1EW1vsYa5LW53nrwm8X3pHRw+C9T5zkowO1XX1fne+pto2z5NY19MOtn0dadNbTGtE8vTllrTpTVtQTlf0+e2xFuh0cu91cnxvbh+j+R9HfrjsaZ6Fs/ORXOs2V1upplS9UxRpny+ji09ufJN1NOnXf613eJ6bp4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAANFb4hxe5XnrRWcojR0wptngnaz1uz2267bddLovWrUilWqaY2rhNa5jCYvrp6rzfXVWbU4noeRp4b+s5+3Z3xxrPm0akTMTMTflru56z0Z5Xi/TK3q8j5p6HjcIzN+Le34vauw7efNOXWKNMOlTfstNRFfL0Y2iq1Kr0rtREZZ6+qw6ttPU6eLe6eX4vrx/R/J+j2dMLr041s/M2xmJygTZFr4vfTUnl78nO35concpt0abfUe7xfVdXnAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADjNfjfm/QVRbGK87Tm175YnRw697Lp2YvhMU1ipGMxExirhZhMYq4yg2Kaes831s7Vs1pxu/yebhv6/n7dzo56q381EadbImSYttZabl5y0rdfN0cHjvT+f8ijE9JXT0/B7qnRoKcykUa4b9N+pntrUvFormldq1zTC1cEWZ6et5e3dW6/X5+/08vxO/L9G8n6PZ0y2tceIjy7FICQInOZ2YvKebty1o62Hb6i/N9U9j5kAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcNt8s836HSpPM05tLXnuz26vP2ZxNaKYhLFEJgmUymUTEzGUpmLaX9D5nq5It1px+/yuXjr7DDt3N8K624ueulWayuYxM6T0WmV4tvRplz+nzOF1+XbMdnl9TPn9DWmOerzqxTphtZ77uPRneMFcLVxtVNZtXKYimnpOPt3Yt1eni6nVxfENOb6F430uzplv9HNzauBVSilFVowmuMwkkLDbWwvSzn6uzW32P3vjZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOI2+Y+Z9FpZ6aFspqtTSrUjEgzTamxbOZiUIwRjMY2jIylZS/oPM9SUZ2jleh5HKpf2WHds6441tz8uinPSJYoxlr4TbvSNKTKy1cLUTXCYpThMads9KaadFV8tvLo7GfRdrXGaYXpXauFqQqlhlr3+Lt3ovsWy9D6Pk/EL8/v/E+mv0p0Ojk0a24nPpiYmEsbUpvSjSlFs8ZJWS3JdDLp2cN/snv/FWWqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABxZ1+Y+X9Hz870TWmc6JpUrKbYtatamxOa0JiJxljEYzC1ZEs6273menlLO1eX3+Tyaz7Hm9HY1xwi2jltTTWIE4RKs5b4Y2jEytXKYpMLVwRNqad8+arp1iu+W/j0+gnbC1cJrVelc0xmuMwmK8te5w929XSZj1XpeJ8S25foHi/TbWuWzpjoUvxubaCERLG1cb0wvWq9NW2dF87bRetuZ7d3Lo+p+18ft6ZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcib/L/L+m1MtOfOWpfOu+UxOdL5RaxbYmbkzFhCcDFGEwmszAzi/d8z00M7RzPR8jlQ9jzejbfPFOnnrRnrEBhEzS1nRgtEETGJiVWiZpRamrfPRmmlVVfHo49fenTC1cJpXatdq4TWEJirPTt8PfvVvJ6b0fG+LdHF9H8n6HZ1z2L5c/PXh8XTBExBjesaVw0pXNMJrTNJ0zvW3qa+ltX6B6fzW9rgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABxZ0+ZeV9NqY766tOmVeuGvSylkLptszNhkmEwYmCMZiJiZgZ1t2/M9OYmZjleh5XPtT2PP6ERMJ1a2qppBEIrM1tbvjr3plKCCuEWiLVommrpnoKa1Y1tMOnh2d+18LRXNMJrXeuM1hArpp2uHv3qWS73b5nx7s8z6J5H0e1pnsaZaGenF4umERKLRjaMb1x0pVNMpKz2L07HRjbE6eW30f1flt/XAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADnTb5T5f0+pnfmIoUptTKJziUTcnZstmZiSYMTBGMxEk1IsrfueZ6edqYxbi9vm6t8vW8/oyiItqVvVS4iERM0tZ0YxpTExJRWmJrjaKbU075aCmrWNfXHp4dfetpXNcJrhauFq4TSAV007XB3dKY2NsLdM/lHX5P0fy/oti+fS6OLjc3bx+PoiRGNoxmMb1x1pXNMYLVnTPOVtNenTT637XyO5piAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABxmvyzzfodJGrneqJrmuKJrOS1krpWzOSZicTGYxREsZqQLK37nmenKbdKef6vP0Yr6ji9W7WkQ1zCtoSiYhNZt6MIvFU1hMowliUzWq1NW2WpamnEU6YdHDr70612rhaldowmmE1TEGGena4e7dy2THT6uT5T6HifRfI+ktvS61dOk8rl2gxmIsia4XrF64zWq2evrlletpt016dNfsHtfIX2oAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABxWvyjzvoNeJ5al+WsxatEIiEpustmZSEoRCMJiJgJW0t3fN9OYtdrTzvV52or6fg9bK0Ja6MIkmSSMr29PPZpSuYitqYmLVwmKb5wV65ak58usa98d/Hp78dGGmVVq12rWrjNYmIlGWvb4e/cy1mXQ6eX5X6HhfRfJ+ksmLbV06TzOfSJRMRKJrFow0praY0Xzi9cpjYppuxfpRH2T2/kZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOM1+S+d7+pnfl3z1b49DHo28OnGYrmuKIkM5ZySlGNowmBlJK2lvQ+d6cVmy9fO9Xn0TT03F6qJWim1cDCJwqUtnnbPrwysiJzmuuL1rmtd89eGGmejOehWK9MdvLo7HP12dOGrNa7Vm1F6SYkZ693g9Dcz1mY3OjD5d6HgfRPH+lstF16aVLc3nvEiMLxVbOjXOu+cWjG1ZmNit75Zw9Fph9X9b5kAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAaK3x3h96imnNpPPvjRbO+t+tzd2xTStFdq4TGMxjMVoxmExNozlJlK+l/R8HpRWc7R5vq87CXpeT0kSmORFdeqDOLbGd7yzqxzsxhnMTaKkYWrFq6s01dM9KaaFIr0x26bdDm69iL6mmOrpnRbOuYytW2VuWvpuTv3a6ZI2N8vmHofP/AETyPpbJiy9dGjlZTVeld6V3rFqpJiUZGzXTO1dXTPbh9F7PJ9p2eUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIOLG/z3m9Ty/N1aSutbKq1B1suro5dNuelVVVq12rXNarVrmtdowViQ2qaeu4vSJzmPNdfmon0nL6MiXCpSuJspa6l7qWy0rZ1ZZ2jCs4pwMEVWptdPDsThx515kW0KxVpjv116LWzg67FtTbHU1x19Mqb0iYvw29by9+5Gkou1z+Y+h8/wDRvK+kzRZevLV4jKBKZgZyzhbFrq3o2x1L42Rf6T2+R7ns8vJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+e8/qfPuX0a6zo2zpmlc0iY3abbme+9XXaz1jO1BValVq03pTbOm1cEbddPXcnpIZ2jzHV5udb+j5/QlCXH55VmyLWRYh0Z53rlLCJgrRRNebvydvbg9bhTlXnw2nXz6ML5b8a9FphE3cHZfXTW1y1dMdXXLG1c8dfU49e5NkFq/N/S+d+j+X9HlIrzJrxpyyMk5wzi11bWVtXrTn78lc1zi+yt9s9T5np64AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeZz6/lPF7ehRZWdC+eE0qmuM12K32q6zXTcrferpfTWM7VlKuvplp6ZWJ9Vy+hMs5jyvV5mxTX0OHdMwiedybTLKGNq47Z52RMWSxicUUTXU1y5unNu6cf0HDPXmfnuvbzqxVbLqR0bszhNYM+Xp3OfpwRp3zrvWKvQtttYYzX5z6Xz30fzPopQOapzopbF7Imyl84mrXPnb8urpjnF84nKZ6Uvt/rfK22qAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABqxf4xw/Q8XOys3J0JzrVrtSuYyNuultdMKzYnai+5XTbppZnfKLLRtV0GUx5Pr8vey6O7j2zMDS4t4mK9KNs8kyJZEGJTamrrjr2zs15uhOWvXfnrc2K618evl1beiua1opRFL7+HRuZ7UxGujur7S0Fdq/PvQ8D6J5v0MokqREV1efevSmtfPWvlqaZYzTNaU5JuT6TXD6563zEgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEHyvl9vyHN16kLonYi2gpVNa5rXNcJrknajSyLYVmEzNb4narpu569nHsGUx4/s8nqYdnby65REuXhOFosmc5RSa8rXdWecorOKML0ptXC1IvXGa4TFFs+bOerOXQw677M9sq5imIpRUia23s9rsN/Q06NlIovTwXd4X0PzvfklGpNeaprRFcRhatV85mZickzLOW7NvadPm/Q/R8EAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADyGXd8z4/Y51JqRs1ttxfQitM1rmtU0wtXCYS2Im6LzBnpqaZdDLf2PN6UiY8d2+R2efu7GXVKIOPWIicIV1RE7dov3pnIRLGYxRFq4WrjMVzSi9OfOXNnHd5u1KnXFaE1xmE1lC0Wc3V7Pl9G6LDV0x8T2eN9A8/3pRKOQpz81WuVV80xNbSmUzZkbS+zL6V6Hz/AKjr80AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADWW+N8P0PEy0phWjYrbpV10aqFKrVqtSq1a5pXNYmJT3eH0de1N61vS4dsiXj+3x+3y+h1KdGSMU+eY6s0EpupffWy2zySkImMUYlc1xtFc1ptTWtnyrc/T5u6yZ1ppra41TWqa1zTCY6ePTFLe55vRtWI09MPJ9Xl+34fakHPq08p5/XxYk1tKZJtMo2IvsWi8+x+z8jt3zAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHzrn9Xw/L6WhRWVxGwt2I21MbaytV6UWpRbOq1MJrjaPT8no7i9MR28uqSZeQ7fH73J6W/TaSJjkznyr4Im2tt3LevXHW6+PJaREiZYoxRjaMJrXammrrRHW5+7YljWdas622FF8qb0wmna5e6hX2nP6GaSNHXDze3B7Lk9aUyjTL5ednmpgJSMrNiL32jKJ9D0cX1P1vmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAObGvyHh9/j43phWYm1M9WusZaaVFFs9e+etplVauM19PzehdLOturl0zKTxPoeH6jg9fcrrJhMeZtz4Wrs5a71dMDS6ePR7OJIYwmYiWKMEYlU172/J5fDp2Mt+py9+aZtXIjLSiaaXRza189vLa2k+0w9HII5O/JxdOX1/J6skmrMWp87GFaAlknai9tokss+iej4HrezygAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB87w9XwvF6ejWcEVoxNpO/XXbppr520lNa+WtpnTfOYTL0nP2dvLrmQ8J6Pg+w8/29uNBWjjYrydK1ROMRp93BqdPLBCcETMVFJgjEymv0Hp4PmvP2X49HY5O+UzMTMJjJNqaprqznbht6ynXIOH0+foTn6rl9OSZjiTloViYBLOJ266zehNlo2tKfYfZ+Q2LUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHicfQ+fcPsc6tqYjBGMsYjYTtRpt1vsU01Mras01dMtTbCm1PRc/Z6DPqySPAel4Ht/O92+LSmua86rGEpttF6OF3+ZRpniYlZM1phQjAwtWxH2Pq8z4zzels5bdPHrsTVWcUwhMJgjIvx29Jl0ged6/Nwpf0vP3EjzbKuqZSX1vsJrtQtZMWXew7PI+g+l4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHk8u35d53uTM6FWBWjFGJbLci+Vb71b3U01cp0rU0N+bu59HaptMyh8+9LwPe+b71iRUc+J2Jjp7YbeuWnTXxnT590omMCsTFUqIilFV8ron7v0+V8L5/W289XP1Y1tsWi2YriaqolE1hG9lv6THqA8v2+VtY9Hfx7ISPP51qRbE7Vbtc6YZRbK0Z2jYtH1v2fkujrzgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeRy7fl/m+7sWnSq1YYlcxjEYoyluTM0vnFt+L7OeteGm1vnsxeSIfPfT+f+geb9BlAYy2tsMrV1Kba9L52pxe3y7JYzGJUJjBWmVaMEWTH3no8j5DT0tPDq0+bp16XwRExfaNi0RE0RGJ1Mt/QYdQHj+/xuzy9/Wz6QhpxGFJrNe9MaWzTlKy7OY9b2eV9E9T50AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeNy7vmfne5sTMJ52c1RGExgjFGExjMbMzsrV0tkblb9im+xWwhPz30vn/AKD5vv5JC1dWl8E2WibV9H1eb7H0fns5iDFNMTajKYkwIL1dya/MsvR+f8vp7FNNbHWpFU1wmsIttGzaMYt0ubp7+XSmET4n0fD9Pw+tv11gGnbPSoqpbKJzTknK0Z2i2X0L1vmvWdXmgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADxGPf848729iyE41tz6TWitGFq4mCuMwlt2WxNdZ72PT06bJYxPz70vn/oXm+/ITSa8WstGVosmvvPR+e9N1+SAAAAAPIV6/knJ6+nS27h061YqtFc0xmMUYTXYjTv8/T26bpRDwfpeD7Xzva2GkRNKvHrnNbZROS0mUznaM7Rjan0X1PnPXdXnAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADyWfZ8u833tiWK0lWemlnFcxXauExXNcTFWJJjbs9Jz9nQprJVE+B9LwPoXne7KRTM0xbKYymNzTL6b6vye5pgAAAAAOU0+I8fu8nO3Rz3pzvTNappXMQiu1dmmvosejsV2Qxifn3peB7/wA/3c4lDURzaxlEzWckzLOzKWUxr2p9N9T5v1fV54AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA8rn1/LPN96+bRE4TFsTrY66kRVNa5iua4WriiJjGRHsMO3oRcayvi+/xve8HtyLKYnCLTMZzHa6OH6P6ny0gAAAAAg+Jc/t+ew6Kaz0MunUVpVqtXFGExtU19Jl0dONFVcT8/wDR8H6Jw+5jWUNJGnnMxKJmZztGcs5VVV3p9Q9b5n1HTwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADyufX8r833tibRE4zC1bKaamGutEVWpVNa7VwmMZiEQe0x7NyLpaFsvN9Pnez4fYytC0UEReUbmmPpezyPZdviAAAAAADw1O35py+ryaOlTbGmmurRNUwrNy/pst9yugoiPD93ifQOP264SaNY18rymTOWdozsrpOFJw3z+o+r8x6fp4QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB5TPr+V+Z7+zaUTgL1qVux318dNaa02pVatU1xtEII9pj27FbTLka83M15vV8nqZIi9cU1xa+1L9cfo/pfL72mAAAAAAApT8U5vb4ePRpVjr59GqjVRZMMdM9I9Nj03RJPOth5zp872XH7EIiY1Kq89JM5Z2jIJqytiivbL6l6/wAz6fo4QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB89y7/AJ95/s7C+REwmNZGjnPXrtTlpqq0WpVamFoxmucT6/HsticpcHo4a0em5/QkwtW+6qltnXn7vT53uO7wQAAAAAAB56u/yLl9vk52wiOnn0c9GU1ZaXXj0WHTkmIcXbj1bZeh5fSxQTURWZmckZysmItGrz6YzFe+X0P0vnvcdfmgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAearv8h4vaox3yi196zMaNJ1oiqaxNfQ5ddeOunatNs6r0wRan1HP2WmdnmuvzdrLbvYdqYiXR2z1bYbB9N9P5i+2YAAAAAAAHiqdnzXk9jkVjE6GW+siaX27R2ubrxmIR5zr83cw6elz9iExMXqTkmyYTHMiiLSRasXpbvz/YvR+e7WmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGsn4lxe5z+frzrfbszItXUrOjnNcxWql26dEZ30ZprXzqtXYi3oebrtsymPJ93kdnk9Dq03xsztTa1yiI+lel8509eUAAAAAAAADxlOz5pyexxa1wT1c96qW6Vm1ydUEHlfQ8bvcPqX4bxMZSztGczmXWjPXniGEW5uOuEIvXHSnT6uD7R6Hz9kwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPn+PoeC4vVwz22pm1aYmM9NPK2telc0wmMLVxmvQjXZy15qmprjtVv28OnOyTxvpeF6rzvaum2UxdrlfNPpPo/PdfbjAAAAAAAAAA8PTt+d8fr8SobeeveWr4+yEYxPkvT8L1Pl+5MSmM7Lpm0utW3XPKSF0xt3y4fPvzstI1pXpj9I9HwfbdHAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABifEeH3tDDrurfZWmt8s9NXDWtWjSlFs67UwmMEV3rtxbsZdOhFbr1vw2ysymPH+l4XrvN9zOyJjZ1z+h9/g+m6fNAAAAAAAAAAA87Xo8xz+h4Tk9Pnzl6XPpo5uiEYxHl/S8T0/me5hCZjbW6G1MpjpdHNdaJhMxJZMbFsvE+f6mrDHTLo9Xn/aPS+ekAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHLafE/L+iU23Zmyt5rdlprZXwRValFqU3pValatd64zXr4dd8sqWyzvExlePL+l4nqPO9qyZztX3foeJ7Hr8m2agAAAAAAAAAAasX14v5bD0NCtvJ+d7dOV4Rq7c2jtzdji9KCSxN1om1LpnOYykJJlka2OmrS2Oude3N9t9X5jfvmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPLU6vk/me/bF9qbZ0vlW9kTr5X1qKpiqa06Z1WrValdqVWrdW9sujzdN9NMEa3Rz87r8/u8fqWSuvX696ny/S05wAAAAAAAAAAIKotoU3smONTp+V+P9JhEIcnt8ze5u3cy2SmY06xnW2aZhMyRKZMzZmYpfFGOlK98Pq/qfNek25QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB4/Lt+W+b7ty21N8875xfOJshKdTG9GVsEYTGF61zFdqUXzu3z2+fpzTXDm9/lbEa7uW9969TTD6/6HzcgAAAAAAAAAAEA5lOmTwHH7HjeL0MJhDld/ldrzvYsTJuo39KImQTLJFk1tlq57aOG1Gc4zGF6V9GH0/wBT5z1u/EAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAB47Hv8Al3m+5dNtpOddM63si2cTaijHTX59BExEkoIRleJlfrnpw5fr/P8ATy7Nqa7t8/aa+f6zXhAAAAAAAAAAAEAoi/Ay7fmnB7ejDUN7fj5cadfi9HappYbErbRlMADKYzmuMTq56auV6qTjNcL0r6MPp3p/Oeu6OIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQIkJjx+Pf8r833b07c2zrfOl7E5xa018dKefSZibM7VwpaIkmSzWm90Y87G/0/wB35Dq68tkxkjIAAAAAAAAAAAAEAhOETVE/Def1055c/ftcnodGLVUvlMZzEzEAEzGcxENal9bK9WbCYw0pV0YfUfT+c9b0cQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgRI5lN/kvB7vGx32JbVrWU0zreyJsi1qNfDWnHTJFutKstIiQhMr98tnaunlP1L3fj+vtyAAAAAAAAAAAAAACASD4Nxe5E0yw7tnl7t6LyYxOUwlKJEwM5jE1876uV6qMJjC9Kd8O518H1bu8PevkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABESBwse35R5v0XMUutG5Ns6XsrpZE2Ra018NKsNMpjO8YVmK2CE2i7bPZ2pqZ3+pe58d19uQAAAAAAAAAAAAAAQCQfBeD3ptScO3Z5u7arewgymEpQlKBnMYmvnfVyvTRhMV3zp6MdyLfT/S+f8AT9XmgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACBEgefx7vk/l/Sc62N1o3JnOullL2Ra2JtTr46Vc+kyyRCYiQkRsb536V16z9Q9v4/sbcgAAAAAAAAAAAAAAEEg+B8Puzas8/btc/ds56WTGUxMiJmAExYiDVpfVyvTnNc1r0zp3y7EvovoeF6fq8wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACAIkDg5dnxjy/o9ec77xtzNtdLKXti9kLE04a1Y3RMzEVmCSRZfrnfrXWpP1L3PkOztxgAAAAAAAAAAAAAAAD5Vz+pwa31ub0Nrn7trO9sspiZgTMAEZzCY1c9NPK9OauYrvSjbK9b6p6Hges7PMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgxTq10wifN59XkMuvx+HVOWuxZuWmyullL21tYtYU46V46IImICZIJXbZ7OldTO3033Pku1tx2zGaAAAAAAAAAAAAAINaL1p59dOHl09jXLjxr894vU2OXu26aWzGVoEkzUSiJZzA1M7aeWlGatFelaNc/QdXB9T7vG6V8LZgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACAcXPp4GfZZMc+t6DgZ7ea5eu+Z27WtrpnS9kWsibE1ZXpw0VkAmQJZ2jpdPPy8dPX+x4HS15t6cvQ6cdkwAAAAAAAAAAAAOfGvncuzetlrxfRrrs3plenyvg9fb5e3dppuXpbaMITMSjJEShGUxjE6dL6OOlGc4wwvFGufvPR8f0G/LXFulbHs35bZgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQBE+bz7eNn2bE1ztXmUvw+T0PPZaYWpnFrE2xbOLZRZWcKXxztMCQEpIiUJvXXtX0PZyd7p8/0XT590061ufNAAAAAAAAAAAAHPrr5XL0L1cpZzHnqdWjDzPB6VGN7632romFbZImYymCJlMxBTWa6zjCJRauF6+47POutHH5PQ6GnL6Ps87pWw6+nNmgAAAAAAAAAAAAAAAAAAAAAAAAAAAAQIkebz7fGcvq6nN2bcXpztr00hGETjEomImImKyiYhMShJNkVEpiUjO8ETaMrV2dM8Zd/s8n6D6XzwAAAAAAAAAAAGlGnksfS2VOLn2eY8/wBPci9WWuUETnLJXOxMSRMCEAiJgTITMCS6Yms69LW2rs6U1a23tuL6f6vzQAAAAAAAAAAAAAAAAAAAAAAAAAAAAiBI87Tr8Lx+vzuT0dvLWvLSAmKzEoiUEBAEAEAkJBNoCUl16pjv93j/AET1fmwAAAAAAAAAAANeL+Qw9LoWy4OfX4zyvc6cXZ2mVdZkxhdpWCQTIiCUCZTMJSCYSICUyjXrN2/F9d9v5MAAAAAAAAAAAAAAAAAAAAAAAAAAAAIQD5Zye1z8+rW5PQ2sta8tIETESBAAAEEkBNoymISRKYhBlKy1UvQd3j/Q/U+bAAAAEFUWksmsgAAAAEHnce7lZdnPjTx3l+z04vlW0yqpIxrN+tSEkJlBJBKBMwRMpBlMSRCUyjXrN+3H9c9v5MAAAAAAAAAAAAAAAAAAAAAAAAAAAABCD5vy+35PHpc3dvZ7xncAhEgSJCCIAASJhEgYwkmWUxN6ek7fH+g+l8+AAAOTXfzmXdyc+rXpqrO7fPvbcPotvP2poAAABBxcuznZ9XnMO/yXB6HTrrYnK0VUmInO1cpSgJCAjGEyIEyIEi0ZIggkoq3t+H6p63zl80AAAAAAAAAAAAAAAAAAAAAAAAAAAAAHzbk9vyvL33Y9VmOsRMzEBMQiBOMBJEwhEiZAhMxMxKcYlCZiZZ2hD1XoeL770fAAAGCfLU7fL49/O5uzV5uvGloiYkRua5ey7fG9j1+RZMAAAQefy7vNc/o7No+fcHpb2W9lL7OmcJrrKUkzEomYSEoTAAkGSEpVhJEJQJ16W9h6Hhe57vIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHzjk9ny/L6N+PVbltFbQIQSCImAAAARDK0SQZSxiZRJBbauETn08foPS8aZoJTMTEW1K6cqm/Px6NbDfGESghEykymO10cne6OKVZtCa4mMMU0xPmMevh4dHfu5nL1dWm2xS994iGMWAITGUxIlkgDKYEBMomYyQIRCYhkaNZ9X6Hg+67vIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHlse753w+3bh125bRW0QgkSiAERICSAAmYQmUCAmUyttWus86K63RhExMsoZxOUTMTESEEhiEQATIRMRMYTFc1ptWi2etalKM6Wsrpv479trsUvdYRBESmCZRMxMhKCRlMSgSQSTMAgCTRrPrPR+f9x2eUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIPJc/o8Hm9Hz/ADehGeuUTlMVzHpduLh0306a4xOEM5ZUvETgjAkkykmJRiQZmabJgItRnNMxdvndMXTFqL5iSTJGMSlETgYxauEkQxTCa4VRNRXCqa696al8qL5xE7ldMqzvRpt00tSmKoRMCAiCJhaMi2t85TaiYzmvfvz2TGJxM+jXrfImJymvb15pnP1Pb43e35AAAAAAAAAAAAAAAAAAAAAAAAAAAAAABB8u5fY83wezdnpEvYd3i/QO7xPneXf4Dm9KuLYRMG7x+hsVtpWyw1zslkQikwmImKIns5b7db4Y6V1nb2reiqs43i2YumLJiyYzRETlMZIhMJxicYnBMVQnCs4mEWrhXE4GCKLU1rZ0aZX61yicJrXnfLHX0mkzLzLNfPExMEYIqvSm+eNZ7uHb1JnmTlXpjvaYfc+/wM5gefp1/H/M+j3M9bs7+27/AA/b9/iyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeez6vP8AJ6eOe8zX0/X5Xb34uBXb47ye1TW8Jpq3fL9W/PWvXLf9Djumsy1Yc6J1IiqERPVy36EaTlfIs0r0dsefjrrZaZymGSZiZJBIQJSCAsQBESgiYgwtWma6lqa81i+dRFZtz07d5zloXzi1IRRE1wxtFd89K+VWd+z53pbddcNsMOrk9f0+b9Q7fHA1Y08hy+njTS+s9/q87q7cgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAxPivN7PGx6sa2rRv+b6t2WuNq7fdzZq7N60Q5tZ04imGcTsVvvRpsyyMiInKSEpyQibJiQnKYiGRExMTMomuFbZzGUxWlDO1ckYFsxFmKKpjVRoxGpWRcnaldMRLFArMSua6emXMvlVlftcHpbtNY1xx6+L6z6Hz/e25QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPG59Xy7k9iuukzTd4PUt59l4t6MaotdatkxWjAxhBdW9sTKUJBETkDMiYttWURW6YyETjMTFiM7Uxi2VqhEiEWWrVFs5TMQiJYWriisQzTmERKURLCFRqzHJvlz744Z37PJ3b+W0aZ9z0PG+reh4MgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGJ8l5fY8rz92zfPZ4fTsw2wTXrSus3WiyYkEkxOUTKZIJEJJEJlCM04zFkTiiYmTGYyiSYvXKs42hDK0KzgTMQnJCUETGtNMpSjBGJYlKIm4SxKkaNqce2FOuMZ6dnm7OtTZbP6x7fx/U0xAAAAAAAAAAAgEgAAAAAAAAAAAAAAAAAAAAAAAAAAAA1Iv8X8/wB/ZvXW4PUtx3prOlauELZWytlbKyYyMYmUzMSBCSZIZJTWAZxbGa5ROUsEZxMmEwi2MxKM5CsQEiYiUEQggggiEykstTA06zro598ubrhjfPOl+zz9fartlNfsvvfE2TAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA8Nz+pyc+vzvn+rZhvrVcrTGJWVtnExpW/PTZiejpXVWiEiYyJSRJMJkCBktCJCSAIJCYRAIEImEIliQQjFOJjDGsxDO9YFLTaukjWvnpa4aHRzEX569fDq79OjKY+u+58ZuXyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEEHjOf1ORn1+d8/1M8ujUrXgdHFiZxau1dXXPu8Ho7FL7N4x0qmJMpZQmYlMwmWQJBMACREJlABEBiQEYkESiIxliYoxhgYxOEMdscM77PPvXevO6OWjTPU259XXDOttnLo6mXT6XHqzmPqfs/IdXbmAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAgAgHLpv5PD1NWunn+D07MejSrHnuzzNPXDBGrrldTf0nmextUvbtnRrlkSnIkkyJTJIhkAkBEkQJQAgEwiCCJjEghGJjMYmEK4VlM1x0ytw33+bqwmvB9Dy9XXHW0xxtS2mm1l09DDp9Xl2WH0f1/lO70cUgEgEgEEgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGIIBBiYkEGJicynR5vm9XVi/C4vSzw6NSsavRycbbm0defT1wtpt6Xy/a3KaXdfLX08MpklMkrTCSYEiIEwYRaqLYwEAEkzFs1smJmIIREoRExCIMCoxidatqEa81ti9+HT0OPshHN7vN5++GlpjNq41nay6NzLo9Zh3bEvb+n8z6Lr83IkkyMiSQCSSQAAAAAAAAAAAAAAAAAAAAAAAAAAAACDEwIIMDArMCsqME6Vb+f4/dwTxePvyx6NWrldPDoXy0tcNTbC3PX0PnevvU19n6fhem9Dw8gSSZmZkSCSQDXi3kOb0/L83paVNYSJRuWy9J0ef6/p83dvlJBAIIMDAxMQcynRo4ejEvLYdWhj0U+d6EnM7fP52/Lq65YWoi2zlvs57es5+7dT6T0fn/AEXb5FhmZlhaZkkkkmRkSAAAAAAAAAAAAAAAAAAAAAAAAAAAADExMTAxKyopKig1ysrOJxe3FOnjcvZOPTq1cPr8/CajCYhbew6dmLem7/J7Pd4UkkkmZYWGRJkWEkgEGpTbVroNq2W1fLIkkgwMDAxMTArMDEA0ce7fy7edn0edy35XD2ZxOp0c2rtz12pXaKpia2vz29Pz9nSm3V7fD7/oeJeWlpcXlpkZkkmZkSAAAAAAAAAAAAAAAAAAAAAAAAAAAACDEwMTAqKCgpKTXKTE52Hbq8vscXn6sceiiI1dcMZqkgMk2J6GuPb9X5eESSSZGZmZmRmWGZJkSAQASSCDErKysxMDArMQAnErw7p5fX0q35GOuplphpnhplhMISQmymnRy337x0OjzfS+n89eXFxsF5cWGZJkZEkgAAAAAAAAAAAAAAAAAAAAAAAAAAAAxMDExMCspKSoqKSsgpi3n+D38MunmcfbMwmIKqXus0dMteadPPbYtXv+p87tdPmySZGRmZmZkZFhJJIAIJAJJIIMCsxMTErKzAggGKcTlc/fyuL2cotq01YbW2rXC+mmjeuhfHrxa+LZy2r5+g9Dwer1ebcWlxcWmRmSZGRkSAAAAAAAAAAAAAAAAAAAAAAAAAAAADExMCDExMDExMDEwMCtPFw6/N+d9DTx9+UJmZRRMc/TPla82zTT0OHXbavovS8Hveh4WRJJkZGRJJJJIIBAABJIAIIIIIMTAwBBiQYpwieLl0+X4Pdpz32OXs27Rro0KuNtz0q+vz6tq0IQnf6fO9n6/wAvmZmRmZEkmRJkZEgAAAAAAAAAAAAAAAAAAAAAAAAAAAAgggEEEAEEEGJXE4J8Jw+zxfL9+YtKMpYaY4dnm8+2dHN29vm9C6a+r9LwvU+h4UgEkgkAAAhIAAIkAAAEAggEEEGITjE8ymvz3zvenh9W6ZqVz6/N5umOtTTs8fq7MgR0+nz/AH/sfLZAkAkkEkkgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAAAAgEGMTBrxPHpuQTJsTTp3zg51b6ldLkdbTHYmCBIAAAAISAAAJCJAAIAAIATBBESCdaHGz3mJzmva0wTNUOPl0SAnctl19cMgAAASAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACCAmBAJIJIAAJSCUAAAQmUAAAAAAAAAAAAAQmBAAAEiAAAEZSlEgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGISEEgBAJABKAACSACZRAAAAAAAAAAAABCQAIAECQAQBMiJJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIBCQAAAAJCAAAAAAJIAAAAAAAAAAAISAABEACSBMiJBIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIJIAAAAAAAAAAAAAAAAJIAAAAABJAAISAAAAAJRIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIAAAAAAAAABIAAAAAAABBJBJABIAABAATCAJAJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIJAAAAAAAAAAAAAAAAAAAAAAAAABAABIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIABIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAAJAAAAAAAAAAAAAAAAAAAAAAP//EACkQAAICAQMEAgIDAQEBAAAAAAECAwQABQYREBITFAdgFRYgQFAwgHD/2gAIAQEAAQIA/ov04/48fW+P+Sf4b/8ALj7Qv+G3/Efah/hn+Q+2j/DP3Mf4Z+5j/DOcf+97V6zup92fs0W66266l/8A+G6jr93chOLGImh7UajuWrZ+5SzWtyy72O9zvf8Adxvb93/d/wB2/d/3b92/dv3b92/dv3Y72/d/3j95G+v3q7ufjtSNK5aPUTqRLpxp+ovvP92/dv3f93/d/wB3/d/3f93/AHf94/d/3f8Ad/3gb4/ef3mPfEO86uuA/WdV3Ff3jLOsawLVFMVPT9MU/UFI0xSFL0vS9I0xSNH0fRkrgxKohinN3XZdcS6ktTWIbQwCSL1fU9QVPT9P0/S9Q1DX8HhEHqeoKpqtXMJSlqmm71p3fqt+9re7QESOOCCGqlZa5gNcQCv4BB63rCsKvqGqaYrGsKvq7isU40FaG9vjUtSfCSa119X0XWopInyIZwqqixmLxCIQmAw+uK/gFc1vXmrTQyIyMtW1oW7gfqWva7qWpoioixBJEtC6NQ/Ii/751Aaj+SOp/kxqY1X8r+W/KflDqp1U61YkrJCm9NRdWDA4esE22dRQoWIl83mWx7Rte17Yue57gutfN86g2o++11p2Yk4wYbd3PBN9Q3Frdq3FGqADO7v8nkEvf5fL5PIJO/v7+/vL95d5HmhflBp661I8EkciMCOuiWlkrE4T3F+4v3d3dzzzzzh6cAZxhBDBl2zuCKT6dql/VNShRc71MdYaVLSK8AEccccEMgGFQvBx8kEjwtUJyIy13imgljlRh1iOjyUDjxBfGIzH4zF4hD4vEI/D4PCIDD4WGCAwshBxxsnXvp29tbQIVxRGKkcNYmytipw0RYN3ckyNyWDs5YuXmeeSJtPY5C06SJYSdZxKCORi5t7KWc5yGB557ueSxznkkN3Sz+StpaQTrYhdWDBsObS1j6Zu/WBiKiBAIc0pq4dbkM+O1Wzb0tl6cuSOAO3jhsmEqx5WdTQayZ8meZpskw9BkKaTDUUmSfyrMJhMJvMZfMZxP5vMJxO06LWyuyBssZNjlsYMu3NVhl+lWrGuaomRIqgcI1OxBqP5CzcnkJ503UNSrEdRnP8AB8kxxJitpU1aXcdeaWR5XkaQ9Bm3aSLGLEndzzyDyDzzyDzz3IBgyJ4p1sNYmsSuxOHDjjYet/SvkfWEyERhSOoZZ/ZeVj1htO7v5fKJPJ39/f3s7s5fGzR7GSR6hG8zu7MT0rwbe0tIma9N39/f39/k8nk8vk8gkDDFbnAQ6z+dpGJw4/RsSXRL/wBImk1W9HkIUrg6HGwFc7e0g9DhcsSGD9/k8vlMpleVpC8mackJp2dV0nVNNYsTnENPQNFiTmzK0vk8nf39/kD9/f396N3h0KkdOe4EZ2srBsONj58bX/pG+boyER4CmAgnG6QqkEld1bOSZcLPJ5BJ3hueSSSTnac00+Fgj+e3sZ/jeL41qfHabeZlLyXr/PPd3d3f39/kMnlEiyeWIoylTh6KIo1geOQPhxsfNq2/pHypMuQYhUoVIOHFFSOvVnq2kkznmQnI6/41dNTThQGl/iTpf4oaSNHXRBtxNtojlhgaK9+ZlvEhgkumnQm0Y6UNK/FDR10Abdl0H8NJpM0PeJYmjaMoVwEkGLK8aR2IpxJjYxfIpIW+j/KpGQ4mLi4uDoxU0cp5bF3JcOHHMUW1ttxV+3sC9OO3t46yR6jtS7Q7ChThIqG26G2AvXjt4/hwV1/auo14zDkZjKFeh6QZXKZbFjJMcnGLZpp+j/KuDIMXFxMXF6MY8oiobbXcmwk4+aYkSf8Ad0n0B9ortGHbcFL+h8mQRZFkeR4mLgw9IcqZGLYtZLjY2PjZpf0j5UwZDiYuJgxejYgoZUFzLuS4cOPmjj/G+U8jyIpkeJi9DhNfKuIbJs5MWxsbDmm/SPlTBkOJiYmLg6NkeUMqZby4ZcOEvmj/AOP8p5HkOR5HiYuDD0hyoIBdy1kuNjY+HNO+kfKnSHExcTFwdGyPKJq5cy1kuHo+aJ/j/KWJkWRmPEwEYekWU8Vb2W8lw42N0ofSPlTpDiEFMGDDjZFlDKuWxdybDhxs0L/H+UcTI8iyPEK4MOEwZTxWuZaEuHGxsGUfpHyr0hxMBQqVJxjHlLKzWWtmbDhJOhH/ABvlHEyPIzGUKEYc5r5UxDcy2ZcbGLkGj9G3Nrqb9169+OkrKQ8ciEdGyLKQhyzlwTB8kczaTZO+6G5X1zS9U/vNL7N/VH3drhl0uWopRkMbA4Mgyq0b22t5JjZIZJkmrb1G9Dvervf6D8h6pLYrEMBKqwCrHUWAReIU49MrURLNbneQNDag8VUA6DbfUdv6hDP/AHNQkt3nsaMosCWdZUWrHWSsldawpikleOVdRlvzSurQzw2YooqcqySJQyC//va5uGbf2t6gunCsa/c1Za4QmNu9ZYrqax+ZfU3umbuwxGqdMXTxSEEOm1L229f/ALU9jW90nQ/xUWnxwJTWrLpcdFK3jDpMk6yiR8EA09tNagaB0p9A/XI9D/EQ1H0jSqc+7tr7l/3NcnGQ/wATihgeikDCDg6piAFsGKe6JasG4Y9jt/a3ZJ2nHyDEyGRnBhwZIMXEk7y3OBy7YcGNJgDNzyp1cbDf/bOXXOQ4OrHFJwgBRxhztCkcJ0XGxW5Bgyk+v5sf+3r5deJMhxXjKoUE0czuxjzhOgIY9OQwJBYSpnaFMWrjY/8AuHLi8w5wAVbCOFBzgEHBgJxsDchgxbDirAK+azJsQf2tYWYKsxhK5WSGOaN0QHDkOSpH0OKFxc4d/IZ/N3LJ5BMjasdgp/uavtq4IYo4RGU8Yh8HZ4ynYEC8dOwxhBEsRjAERhTImtUdmXf7JOs2TclcVkoRVhS4CeqKH4/8ZHQ9RKHpGk9FqMOnfjTpg0j8J+GbRBpH4waXa0+ObRdG/wBze+XooMjBHd3lienHHHTj+HckncpsOJzLprbiGkf2tUIoxwakKGJlReiiuysMULk47FiETRNHGvHaFXO3jCIl1nJc07/d3WmupWEeEBCnHHH8uMAzjBnOSddKzcp0j+1qprOuaktDEzT8fAi5UwYCCrFUPVgM5GAr/A4DrBlygP8Ac3SdXqwIrdwfyc/8B/IDoxRfHph121osf9nWcpOJLsNV43qusyyKacqv5PMLAt+373ve+19bxvC/+SGqHVfyv5b8udVv3+yuv+5JFJpY0TdFK0mgbXobK/Xl0DfKG0HBwkv5vY9v3RfF/wDIHUhqo1TbmhU9ix7Uj2/Xpf2ZY10WxpN3bU+jW9InrJfGqLq41gat+U/Je/7vtCbvaU2vc2zKdp7n0ZaW2du6js5Nh0diLoiabHB9G+VQQoVVR0Ve3t7O3s7HlNiGylob1Te8O/NN37oGt/2tT1DVNZub2/d9S3lX3dBeN03o7FSsKnp+p6wrLD4rUboU+PU+ofKmEKEVFZSDnBw4WknnCjtC6lHSZVgO2NVU/wBiV5ZHi1ui8Pit1aDdoVc0CP1xA8KxtGIjFbR1K7A+o/KD4gQIGxuhwl3mlDyZDW9H09So6Akenx6dXr6JY/sa1MoXNz1JqDacaF6vBW9MU9rMY+wxiPsSFkuhww2g/wBQ+T8GIFxMfDhLM7yySPEtdaFTs8csG2z2gKNlzf2NamVeNUBAHbu+toTiPxSQQWYbHaM4iS0l/Gw5t0/UPk7BiYmLjl2ZnkkldzlIBKxEawPXmATtzbGqUL/9a5f3ZurQtc17WdG1iPFjMOtVdpTeN0ypAKOccQi5l/Gw5tgfUPk3FxcXAZDIzvI7scijoU0vrqCpgbdgoFlAbNp60rf1Jpd5aosUbpqLXNNdMJYaS4zubJZntU7+DIcuZeJxs2t9R+S8XFwYDK0rO7scpUqdKezasaZZSRs531Ht+yTyxry7X3Kj/wBKefd26RHLLPM8lelorghmzVBWlDPlwKyPSvI0OXzew4c2t9R+S4Ri4D3TNKzMx07TFEs080r0GrLZpwR/IcG0nUdnFisjbR3B/RJ3ruZFnmmkRKdO3NozqnjlTd0ekOFkS5GCjI9O3Vlu5dw4c2lH9Q+T8GDAeZjKxOmaW8kkkkkrnKxoJ4Zam/62wR6xheDJ4Im2jrH9D5C3BIWeeWOKpUd3OlOtcVrsG9IdpSBQJ0wFWR6NuaW70ObG+o/J+DoSTOZM0vTZZneR5Hc5GdsIBYi39B8aD1PWNGwgM8WkXqlj/tq1/UrkYszIlSrK6lsqGnH4NSr7ti2A7VfWlWyAVIKPUu22PTY/1H5O6HCSZs03T5pWZ3d3YnFzZ5jkze8fxk6IIUh1pMBmj0PdP78N+nfv78N/jf43/wDvw39Dv399rb5gmz5L1VctTHNOqSvK8ROQnb+GKxDr0Wx5TDLVtVtUXkEMpUyL02R9R+TejYcOVq8sjM7uzsx6DNgYqvm5k+PHiEaoN1pg6WUlsCQv5PLFOcMxtSXTcW5sMs2v3wZG0irK8jSMsjNFmynKsuow7XZcfLo19MGAqysDKmbK+o/JDDGw5XhkdmZnZ2YnqM+M5LFc5rY2Y8To8bb0AIIyYapDPbJJmlieO15Zcs4LEB2Dm67crTNDGokeRnJzuibYEpZjdNI0pHa224h0GAoVbiRdrfUfkaMY2RxkuzMzMWLHrDW2E3c2akdEeORXSTeeAgg5cU53NlySJnOmzTPIzmI7Bz5P1EmU6VFI8jOWLdK5+Ppu8y3TdGiy2MvPrf8AAYCpRpF21D9Q+RiMIjR3ZmZixYk5FVVDNs10mmlsyLiTiZZt2EEFTmoo5JbHdC2VWld2JiOwM35fkxq8KuzlyxbpWzY0xm817Nwjbck893NSiP8AAYChjO2j9Q+RUXII5ZGLMxYnCYq6oWCBdpyC35ZHv4kyzrPrp4BUqbanCtlkaLDimR3YmI7GktWniUyGQuWJwRisF2nJ7HsXDu1dqWZbNme5I2MnUFShjfbmqfT976jWimdmLMWJIVK7OqBOwQaHVWSOWRtcEUgm81iyLNnRGRclDDNTeLK+ANjuxJiOkzSNp0iZIzl8WFK2OsQ0JvKZZjvEbRmkkkkkMsGMnQEYpjba+p/Tt7awTjsxYsxMcIPasaxx04YItRFkKqO25crsr+WUtBBOmqS6PNFIpGrSRZVxBNjYTkRnthtJaPGKwLFwIvE6AaSSxZjvJdpSyN3ZFZtu0EiYMGDEKHaWpfTCd6avSSV2ZmZiYYWZEjijqZLdltS2tCkidRZG6xSIILEHlo4ntapcaS1eljynkcdkFySqDcdyDNOkrV+wqIuztOSYDp+c85vAbSbFOXWSaK6Glg4wYpjbamo/TN7apEszszMSTFExjjiql5bryvJNOZNuuMillk3gNNPQ4M5wnjWBKZ8rU6mk+K6AyuJjNc1FMgGdvHQlscR1ax6DN2rs9yRgzUyGVkkitsjxDAUaNtu3/pe/9Vpxyu7MSYYuYavMtxpGd5XkgryUNLr5zm880g9DguR2g/TUq50t9FraXWMkEmmnQxo34Y6M+gHSokgOHoR28SXJb9Rj0TN0LtFsGDNXpeqCGDQzxTSRYhhbYV76Vrl5jYd2ZiYY0jjhmtvIXaR5C5bbJt06j4c53lFomDrPXV4rcNuOavB+unbQ20NsfrD7ZXbZ20duLtpttPoNipAuMXsPae5JZaSvFAp6RjcEW1BkecNKEspyCGV4LToMjbS7aN9J+StRqLK7MWiSKu8ss5ZnZ2dnLc7cy49B+u630Fh1DHAQYZ6c9fWRuj9nO5v2V9xJr53Edx/stzVvyOo2aTzzyztKzvI0hfQnlsDpCL50KbI8AZA+qjoGBDV7JAyE7Tt/Sd43bDOzFFq1prTOzMzOzFuViSloi3n0qQdd7Ztop1aY1/B4wNOeRUwYQcJLBPD4J6l/J3ozyytjYwYFPHQkR4+kA1YbYkxMUTkZqhznkEMGrWDkLbAtfSNYtKbMjNFHDFNOWZmdnZljSmsflLLPPd0Vx13xm1imcANH4/EYWggjlxVUZwIjEysDiWdQeU1VghMLQtCYjhkM8Ag6Vl1OPbBGIEGvzxaq0ctKSLnAQytVnTNoWfpHyFbBkaNIklmZmdnLRwLWLnAoRYvF4dDqDAc37m0zHnAyeI4G8nlh6HAOHHkLJLIpyZrjSHRAUbJBKrRmIQxab69XpSS9FtxlCKi65EE6Ms9N0wEMjVZaE6N9H+SZ7rgQxyys7OzxxJCX7RGIhFHRTTUoVqz9Rm/s2fkIAAsSrK0RUCLO7FyEa5rL7mobmSVicnM5fNAyXDjYQIYqGSW3ajgzTVsJoOLiYglglqPC0JQ5JHPWwFGhkrPoU30axq27LVt60M0zOz8xQdwVIY6iUlg73tNda6b9HUegzfubMNYAAalnKTpKI0wEZGtUavqc1uJ9vXa0kpnabBmhvLj9FBlktOyZ3aaQNJycaNiBAma9b94XluJOQYJILVbFZGoy7O1P6LrU9yxCK8dmZ3ZwI1SOGhFVJa21xrBcuXLk0DiADfw2XlLO0DVeoMU0TjFEeaSN2VsjzR8qZMZzJgzRTIZG7mdizEiRX0vBmjCbNNxMXI83R15BSZLq2Hjtaa6I1WQvsXWvom55JIw8jSSO6CCnBU5a29gt3dxYuWLd3NAnIsGb/wA2XlDAM1cdQa0gIKGlJusNgzblCusxsF8GaUZrGEsWYkkmHNI6aIJMp5GVyNd0DrzzzzHOluWGajHkJ2RY+ibkzxyxWXfIKFesZHsFu7uLchPB63riER8V8IjHO/c2VlDOM1Gs1Jqvi7Yih5jMqX9NfZdbaHhAmyxjYuUyMOEFPB6wqiuYtI6aI86whMhZDr6tWao1Q1yOee4NHNHLLFVGmSD6Jr+SiWWSKGsXeUuXBWJaywgYcI7e0DIAcXBm+82RlLoMu9e0xyQqO4FXdgXcl5JJZ5e4ZQWCsYypBB6tmkY50+UYuLkJTNbJw4c4KvC9dl55UxSU5AK7fQ9wLPkykSztIZFCopUhQvHGEYBnEOHFIzfWbGyqXPNzqc5AwhSq9pwY2GNori8KNLDYcOHDh6cNmlElTRL5EiRRjWR1PQ5w6vBkeKK7Tx6W30PXnlxkmheMwrHi4mIBg6kH+MWNgwHfT7EyvjMDd6noOgUgliQ7PkmOLPRXolsOHDhw9XzSugzTHkymQRmudD/E42cSRR1FqVKGqwaIfoe443wpIwd26DI8TB/A4f4xYSmd29jsXAxYG2eh6DEwkKB0OMpxzOxx8oYcOMTh/g+aXhIzRpnNQghtZPU4cOHHxSg0lg8U2qWND+iavjmxLJjAqrYMjKEfwOHD/CLAY3Z95ZsZzOsvksn+AxMbEHHPfhLmV5sOEUAcbocP8GzS8ryrbNuLIArxWdXJ6noejgB3jkBgwvpA+h62J2csGVl7RgChWVucJJ/hzDjYgObzzZmRYTk568jIsYkFQuSDiYnLOHBmnHGwkkknnl80voDVdMrnAdVPU9D0IeQlGSSsZMpL9D1/JDKZLMMrR9vHCukwXgP3ckkg9IMcp03rmysjzlcmPQARnIHckyjO1wSrSrZMzJmmFS2EHCeeeTml9BkeRZD0VtUP8maSY9FESVcroo+h7lSXJGnUNWvcNGVK8DBOLIfCSQ3cGrHgAZvTNkgDo4LGwbInVos4xsIIZipyQ2MOJlM1ZLE5te37AbjjNLwYoVoMj6KNU/gWadp8PRUih4ik2vB9E1lbcTCyjhlr3a9nsaIxlCpUrgmFoWhMDVA6DN7ZsrBnHFyYygqsaIIcPSOU2PN5vPZyTLGHEyosBkilheJlOLOtxZ9J6LhFPI8GA6o5smyZienHascUDs0lVNh6Z9FtbY3PQGTxsrLG9W7FP4XgMRjaMoyFSCONN6jN8ZssKOOLahFREVRkfVuhJkipjQr1efGyPNPwrDJJFJDJE8ZTtgGkjp36biYMGa4evAURrGkIjntExDZu240+jfIVa4hDoylVaGeO1BqAw13gaFoHgeEpxpuDABm+c2WB1lRYljVGxBHh6HCZmksUL0Uer6bMGEY05cYQTNHJDJC8IhWPSh0Y6RiA4M1vOAoRY1hSERzWZbIAyum3Ivo++qt5FaaJlKlcgtBY5odShv8AJgNZq8lSWtpyDF6b6zZuDqVCgM6jI8OcnGM5lSlJXmvzT44jNInDnMU6PLXkhEbR6YOkuaGEw4M1dViWFYVh4lnnucADFFCKtF9H1OtcicQyzwMpUqVSSO0U5jni1KLU0s8CpDSwdN9HaAH8CxlXoMQnocYSKYqCR3rduYSA5QBxi0hmFqtqnMkEuad1sHboTGwZLXOl8STSXZLDOFAwZCm2av0nd1K1Gwq2J6zIVKFSqTJa7COQ8d2PW9I1k4vTfebVA6vMEWPgkPgJJDgxvF28ESLKGFaSKU46PA0Bijlg1aWah1unbGJjdL9v8tJZ4aQ4ABgxFgTYFb6Tv+jdRsYVLU0DxtGyFSpCvHdWUxCGc6EnXfmbZC9DhsG0bZsiaJVPPJw5wQSSxfLCMY1IMgsm37vui57EUtLrqh2mBh6augiaQkAAYAMAhjQbDq/Sb9W5XlRgcqWnjlhKNGyMpUjEl0vNTr6QnXe429i9GyCv0GRDuWXy+Xyd5k8hcyGRpJXGVRJHLEUZWUjNPqRLW66xm1EHXUE8Uo4/gBDEBAml1fpW/dOvRkMDlOyDNWaNo2RkII420mtZRjwdN5DRAvR8imwiNIo7TrKH5547OCvBEglmhFbOWieF4mRlK6SEWHoM1w7TUYOk2DLg46gRRYo2lR+l6/p00EisCMp2Yp5a0kTxyRuhHG2W1vKuDrvWTSgvSbC6iGtHC73x4fF4WhMBhMPiMbLuyrFIpjwHnvKSVJq5FaxXxOgzc77M6DpIeZm6DFEUXK4g2FR+mb5p3YiCDhyvYisAzVXikheIpHOLOlP053kdPA6T4aiQmRjl7qSxZmZn7+9TvSjAKeDB15VxNJTahVQdBm8G2WowdNTsA9AESOIsABTr0Kv0z5JQGVDhBByvYBguNFLVeCSCSJV26Ou6jUA6TYZMCiFqlw9hUqwZWDDjIs1Ot2aXFCxRkznnnuE9YDrvZtnKOtoYMCxVwCwxcUbF076b8iMDaiIIIPSCwMimgutDLVmrywaAOuvmLrYxUjgqaUXvanYr9vBQqysjhx6tYUIJo6kDBZEt40Wc93OmDrvl9qL0JmziOGOB5uRgwZXj0Kh9N+RgcXLUJBBw9I5Y5mWKxBf7HoU6vW/i9XWHTjanvY6rW8JgNcwGBqzUzR9GKjtqLWtIu4AQcDx2xK8LLHmkjrvl9tLg6LpMtB5HlAAGAKNlaX9O+SJDgxlkjIIIOHpDb4ORTpq1Oz0OD+AsS2RgGbaoDQRox0U6E+2hoH4P8IdBG34dKz5O02YLkikHDnMdpLMh0P8AhvJ9DTBhMluWzgAGADFz4+P075K6DObETAgjCMOI8dzsI0TB0c6Zg62cUKBkS7Prf9PkmCwsMk6MGB68wZoA6HNytpqt0tucAAAwABFmf45+n/IiHBhyN7UBBDAgjqjxXNHA6TnbYHWfABgykmk1v+m56jpMKrOrBgQQchGgL0Oaia69L5AA4HQABi2fHifTvkLDgw9InswMrAggj+GjqOl9tnoOsuADFGmVv+pG46F6GnLYR1YEHCEzQ16MaeDD0vnB0GKBgxi2fH30/wCQ8JGHqjWICrKwI/hpwwZrb7KQdWHSnHsyr/2+SNJlSRaE0sbqcRGTjSl55svtkHCebh6ABQEV2xs2B9P+ROgw9Bk5gmnhZWVgR0UVOgzcz7KUdOXXK0Dvo1L/ALWq+t6begR0aRWFUEKKg5zWpNpR95funzjgABFYno2bA+n/ACPrBkjmLBu+xOJ6tqeF1YEYcgEDBg2732WownmVIYrMu0NK/obv0BorUFedxKKmNkCpnd3bpl2kjN3dz9AMRVyZy5d3+N9T+m7w3DYl4VEidZsbDkclK5YrsrqQcpiMo3dvOTauK/dkkVu5ommU639He21ZY54a87rXDjT05553bJt9OeVOADESeUv1Ip2ds659Lv29XvdqRxxFSlmB4WhMCR07titIjg5QCYmc71fRVRiY2uarpenaJpf9Pdu0bEE8MMsRsJRj55LblfTkJJTOOI0RX0saLNoUkHHDLt3VoJvpXyJq6hVjj4RZWkDA4cOc0789WWJ105FPIO6XqqGDR5o9LRdI/q6/si5UnrzxwzunPJNsx9DkQyKIBIhH2xSGTUtL4IZfjjVfpLNrlwtAgGMWxg4YHDhJNLUDHLSCBiym6FOLkR2TS/r69oW59oXa9eO4/dy76SvPPdSi/HxLFXEAg9cVkriG9XwjbF8H6RuSy5QIOk4OHHDh8bDhKmF11ZZ87jJpSrgK5tHRK1f+vbAbde3NK2dYk7u7VptEiwdILj30uLqA1H8l+SGpfkl1K7L0OaHZ+kfIUrZAFwdFEtZgcYMrKQwI96I1FPTUZtvxqVMQ29X/ALFhQVzdtt27gddlpR4MJSUP3B+/vDdwEUTdDj5sKX6R8kYTWxegwYMMc1cwmA1jVNRqb01gTOWOtZp8KBU0Sko/syqub81F27y6ImAjKFf8UNK/EjSfxA0aPRDpc4kJw4cfPjj6T8k4crYuDBi4uLk/Ujs7OzxCL1xp81fUIBBBSj0/Z+jf2rMOr6lcjkrSnb1WGOCqtEUYYuee7vDB1d3lL4ehxz8b/SfknGNbFwYMXFwZP0UFArfwiQF12tX9IQBf7hU120/f8fx5mvJWEAbBgPOA9BjZJkmHDhx8+N/pG4dR1zWjlXFwdFxcXJ8GAOf41MTGzav+H8hZ8fZr+Vsrs+DB/EYuPkmPhw42SZt3cmjX/o28ctPJlUrgIxcXBkxwZyR14rBMObV/w/kHPj7NfytkZzjoD0GLjCTHw4cYynToNiH6NvTJ2lNYqRgxSuDJ+nIBbrzXaLGzan+H8h58fHX8rYuDAP4cjFxslyTDhJMp0C5sub6NvUvO7VipUgoQQZuofnnnkGuIy52n/h7/ANo7Ioa/lbI8X+I6rjZLkhOEsZTBL8eL9Cd31T8zuPe+4d9iPmsVxSChUgyHnpzznPMRix22pcFuOX+2zNfOpatuSpvy5cvJuM1jCVPUHpyuSGYyFicbHzb22NIyKx9B1uVqS07+gxbP3bUCVwoUKFxcGS5znPP8FavBKdI1SGWlPDL/AGtbMVwGyI60Goz6freVTCsVI0jXFUVfU9X1lrvDNHIDhDK67EmkjMdPVIJ/9/dzeaKx7mq7k1PUpa3iXFlWYT+z7TTiTv7+/v7+8SGUS9+nXtD1EQxTRSf2daEYQRI9PUoLO45IzkTIRKJxaF4aj+S/JjVjqj3jYMvkLkbPg3Hraa/omtpNW1uKT/d3WmrW4NS9+ezwYzH4+zs7PH4/F4vH2dnj7AgULgACZDJLNsW1/Z1UyCLNdepaKpAawjABwZ29vZ4vF4PAKpqen6XoihFHYHqV4famt7Y1X/d1i3ucQ/wP8j/y4/gMiybNif2r1dwke846OAwhx0H8AevPQYevI/guWs21/vbyk1uxD/EZxxxxx28dvbxwqhApXtII4GJk+bF/5km0tj/rdriXdtqmcixugxOg6cdR1Gc9QOq5azbX+9uMX2gtibv7+/v7w3d39/f3iTyeTvEneJPJ3+Tv8gkElqbYEf8Aw1HXbu5JbnvDVq246G7Kl7/lI13NbpVHEkbs5IIbuDFu/u7xIJfN5fL5fKJRMto3Pb9sW57O0mrWv9zcWXFjgNcQeua3r+t63q+r6vp+n6fp+n6Xp+mKfqCmagqCr6kdOar8e/zkk1nWppp7bZ29vGVL+ib0hl/4X5hrHdeiggFKOm1Y1RU9P1BU9P0/R9D0BQ9AUPQ9EUPR9D0BRNH0hQlj+NX/ANzdYs5Hh/iP+g6DocA6R5ZzS9TG/jv/APfhv398/e7u8JdYs2fKZzN5fKJRIJNG10b7O/Bvw7+/fz8hH5Fl+Sda3ZBerbhL18izkjoOgwdAeeeg6DoP4ccAX8+Nf93d+j2Vjw/8OP5cfwPUYMTLGTFs4UABQnj8YiMfiMXj7PH2dvb28cEMHVw/RMTK2V8jznqOi4Oo6jrx/AHBhy8PjQf7usbO1HbTXhbSaV0trtW/Gbhve+L3tm4bp1Mal+SGo/khqZ1MaqNQW81ldS9mdUQQR1VpR0loGgNPGmfjjpv49qB080fS9M1GqGs0BiaN4nikjaONBBHLBqMeoR3pLx1f8v8AmPzP5n82ddGvjXV1b8l+SGqpqNXRP1yTQLFsauNVOpJqVXSJdr7T0X/eI+SIa0YjsL8cVc+SqTYc4eSCSJbIamtIUjUev4fFaWhWFadYooo/UmqLKk621uLe9/3xeGo/kTfa97zWvb9r2TOZS5whg6SRlUTj1njiMVZ4GgNU1DX8JgeERxmhgjnbNJrRJ03vGKtWqa3x9B9C3LthNj/pn6VtzQc3pRkUjiYVlCToyLC8c+Ts8nlZqV9LZxYlWtqFqYR9gQIE7AgTs8fZ2ePxiMJ4vF4vF4vCYWrvWkrNGX8/mjkh1Z9SzubHwr2So+JlNQsiBPjyh11KjLsCPYqbJ0TRvpDruTTyMkFYASLZMM4eWCeo9T1Vqw1YoEiEZj7OwJ2CN8dOwJ2+NIwAgQKHClS4LhUYeLxGCSrJQm096Aox0I6Pqep6vqmt67RWQ2IKOASKE2fpv1Hf+isDnbVUBxbKlZO9mZfGYkwN3h/IXL94fySyTyPKZPKJIpYJO8SpKDDIJfKJfIs8cqzmfz+Y2C7sGWRJWmMoPLYxnN1xiZpxCyDaOkgfUWXeegMKkVIcSZakR1YNnAAXAOAOAO3jjiQAKpDABRCOAq4AmDO/k4DwM444cIJAAAhkBxX7jI8liW1IjIdMxQ0egaV/YB/2tToa7pWj1aWET5O6urpIrpiIyAgg88g89OWyFlPMhBBg6HIsLqQezju58gbvL9/eX8vm84sewGFaaB5JJrEszo8Z0ox4chb6pu+pXhrDLOSurhxILsc9a37zHu7gQ3IPPIPK9ASRnI6cg5zhzk5yc7u7vLl+8uXL96M9j3XvyTSSzSSOjRnSmjDjSX+ps+rzExDi7ksglE0tiKaviYhdwwYHnnnnkf8AHnBnPOc88893PPPcWLlyxcu80D5PjSySyys8bRnSnjxs2y/03nnnnnncbxrNg6XRYjZ2nkmrGqU6Fg4fuDhw4fv7xJ5BJ5PJ5PL5fL5fL5fKJPJ39/f39/eXLlu8uXaR5g9RhkoleSR3AXIzpxgLjZknIPPPPP0LnnnnuLlzJ5DKZdflrtbLdLmJTnrzQmKuKeJnhNAUhT9MVBUFQVPU9X1fV9bwOhfy+Xy+Xy+QOiit6nq+p6nqep6hpmm9Y05YWR0iWBFxx6U9cxmIZHlQ1s42TMJRL5BIHDd3dzzzz/uE889xYuXMhlac2NTanLZyUZcx7MrzY2QCgUzZzix7HsewJxOJhMJfL5RL5PJcr6psy5pJh8XiENPSNN2Xp1ASeTy+Xy+XzGc2PYM3savVia3DV0fVYkw5PLO7ZyuR5CaDZtCZZ1mEolWQMG7ueQRg/wBw9DhxizMzO7vIZHePJ8n6XMkTx+E10rQYkm3bIu+77oui6t4XheF0XRcFv2vZ9j2TYmqtpC6RFXFgWBY9k2jbNw3TdN03DcNz3PcNyacPYzWIo5fJPjQmu1U1lSJdLIzbTpIkiurIyEYMAwAYP9s4cOHGxsctj42Ph6ajC2XgbMsskfh8Xh8SoF0uVlHUAAYMGDBi4MGcAfw4AA4OHDhw4cOHDnPI6MteeWRmnpANGYvF4fF4o4abtZ27IuJiYmR4mDF6DB/vHDhw42MGVkaNoTD4p4CuomOp6H446bLXjae1A4pRUKUApCkKIoiitEURRFIU/UFT1RW9b1vW9f1/WFb1vW9U1TUNM0jRNE0TRNE0fS9L03oW9N70y5Wi0/8AEyafUyzMJItKTTPQqRabcWJY1RFUDBgODBg/3CCCCpUoYzEYTAa5rmDXNMuSL0BU6kXknfSyMjzb6CsK3riuK4gEAh8Ii8Xj8fj8fj8Xj8fjEfj8fj8fi8Xh8JhMBrmv63rGt63rNV1bQJI5zWKm/kLW3qmkS/d3aHiwCERCMIF4AwAYP9zjjt7e3s7PH4/H4zEYng3bpqdBi5agl24+16tQZENqR+IReLxeMR+Ps7Ozs7Ozs7Ozs7OzsCdnZ29vZ2dnZ4/GYvF4/H4vF4vEYdU0TWNMr4pu4mgTbUbbtLrztwLH2dnb29vbxxwB9C4444zgr22qh2v+sHbA2zS0L1/Wv6ENrJtutV444/58ccccccZxxxxxxxx044444I7e3ttUTtX9Yh20tf15qR2v+sjbH61p+jD69x/rD/xWM4/+KD/zuSDgP3L/xABTEAABAwEEBQUMCAIJAwQABwABAAIDEQQSITEFECJBURMgYXGSFCMwMlJTVGBygZGxBiQlJkBCc6FQYhUzQ2OCorLB0Qc04URwdJMWFzVVkKDx/9oACAEBAAM/AP8A2Ew/9lsP/wCjpZbIK2iZjOjerMzCzwvk6TgFbH4RxQtxotIX6bCtg8eGNys78J4Xs6W4qy2wVs8zX9G//wBjrFYatvcrIPysVvtVRE7kI+DM057i57i5xzJNTqNa71I7JpJ4gKRnjMcCeI1PjffjcQRkQVaoKNtI5Zg7SitcDZoHhzHeucULb00jWDi40WirNnahIeEYqrKP6mySv6yAnbrCO2pt1iZ21P6HGp/Q41P6ExT+hRqf0ONT+hxqf0ONT+hxqf0ONT+hxq0ehxqf0ONT+hxqf0SNT+hxqf0ONT+hRqX0FnbT/QW9tWy2wXABAw5hmZRdmgnOdRoJK3vcB0NxP/CgiFC0H23pjMI7o9l5TSLshf8AGqhk8R7SfgVcdqn0fKXQmoObDkVbGGhscStXocStXokKtXokStPocStPocStPocStPocStPocStPocStHocan9DjU/oUan9CjU/oUfbKn9CZ2ypt9hZ21J6A3tqT0BvbTfz2H4SLR78JY5o1oy1YRWtleDtlAioNR6tWDRtWvl5Sbzcat9pq2y0szOjEq0Wl5fPM+Q8XOrqqnHcjwR4J3BO4I8EeCdwTuCdwTuCfwT+CcncE/wAlPrkn8FJwVwVOFFWWgyWyKjE5Lghg+TBg/dSMhL38nZIfKlNP2WimEtl0jPaOiEXWrQrz/UT9paKnNIrTJE7+ZWyFofZ7Ry0fQapjzcn72/juV5gZNtMOTgg3J15hyOoS9eoc0o6in+QjvYV0JzsmJ3BHgjTJOAyRA1EK32E1s9oe0cMwsmaRh/xxqzW2ES2WZsjej1Ws1gs5mtUoYwfEq1W4uhsdYIP3KJNSa6sUE3DBR5KK7koq4BMzoo6EphbmFHlQKPgo+GCjIqAFHjWiiJ8VRiouhMrkFHUKMYXQQoXHIKLgFGQDdGCYLQ6GLdmrz6nU2jpJcImCrinwSyR2CzRcGSvVs0hKZLXO+U9JwHMnsrqwyFqmllvyBqoAxzr0XDe1VbeYatPwKB2fhqJrUYcUFxCGFU1NLcAmplck3eExMyCYVHwCju0UYbuUajUeRTKUAogCUAUNVosUwls0ro3jeCorTSDSNIpd0m4oEVBqPVOz6Hgq/bnd4katWlLUZrTIXcG7m8wAoM1UOevGuoohFXsFSqoFTeiuJ1FUWNQU1kTkZJy5xqSalUjb0ovkAG9GyWJmjLMcZNqU6yOa+GQPYhOwQk+x/wAI3wAekIOYHDIotbeG5DmU3o0VSSSjTNYo8VQI8UU5HinI70XDNVQPNm0c8We2Ey2X92KO0QtlheHxvFQ4eqMWhrDfNHTvwiYp7danz2l5fI81JPPxRqijXAoopycnDeiincUUUUUUU4K9Wu5X5+soNYOpd8veTinWzS1onJJBeadSIqVTnmzWxoBzNR1rlImvBzAcE24WtOA/3WyUWkjhhzCiUUUUUUUUUUUUfASaJnEM5LrG89jpCZNE2SJwcxwqCN49T4dGWCS1Wg7LMhxPBT6Vt77TaDi7IbmjgFVNagqp0mRCtL21YLytEfjxuCIzHgaGmstOKJwoiQVvWGos8U0VIyr1rYOlY1To9GWmRtPFQc7PUAs1TmlkjXDcarlbBH0YIVcDwCwPWq3iOKJGquoo+C4BUTQVXAKU5MKfkQiOYYJho21v70896cdx9TzpLSZssDq2azntO3nUdZLkXBSAVClYMcRwdioJAeViFeLcE1pJhfeahkcCnN2qVCpmhmEFXUOU9wVEKFUAyxONFWNu+rKGqF/iNnAoYqjUFgcV3hfXI0S4oO0ZPGRs0JTGB16rT0YoZUCFTVAEqnOI0YOtG8egf7rNNEb8cb3+yzA3oatocCm1qgghRBClAU0IJiCFMU1ikldRoVSHWl9BwCghZSKJvtOxTj/4RWfMINRgRkV/SuigJCO6IdmT1NGi9EuZG6lpnF1n+5VTXnNwvKOhomuJFRkhtZdJCuk0JCqa70Gm5IKtKEkXL2Y1ansJa4YhFFEhVoQUUUUUUeKdxR4okKkRCuTtdwKq1p8sKscsed5uCa15DqtockBJeJq7JAbwUM1nzb8rGjeVyVghbxFVQOd0qmKIvAcVwzRXTq6UeK6V/MulV36ulFFOdg1VO0mgABY3XIXaIBuVaqoyVHHp5p0PpVs2cTtmQdCZNC2WJwcx4q0jePUuKyWWS0TuuxxtvOKl0tpOW0yZE0Y3gOddTWYFC9RA7wUSCTkUHZa32WShNWHMKK1Rd1Wf3tQHhAqNKo5ctYBU4xmiuThCGcTsGxIPgVwKOoHmutlvZhsgoYAYBuQV1gamCGQkV2aDr3eHJQGohDA70LuavJxHikBV5xEh0XaT0w+pfiaLh9uXXTm0TmolV5r4tmpulBxJCunwlQVijFPQnYdgVQpukdHvsspx/I7gVNY7Q+GYEOajznzShjBmu4rKCB316a1wpiRn16toRDdifwVFREIoqvMOp9nnZNE4tewggpuk9EwWtub27Q4O3+pLYYXyyGjGNLiegJ+kNKWi1SZyvJ8GecWq+1EeCCCxV+SiN24/x25HimwON+MPYc2lWXTtlJZsTtyO9WrRs5itMZHB248wlTSPAuEV4hckWvLCXpsLboAc7eeHUsKBcjZzK8GlMOkoucXONSfCjwR5+3abA/8AUZ6k9xfRa08ZqRDwd4qqIGSp4EolFFFFFFONaAorvjXDdmmTxAjB24pwJDsHjdxTmuBaaKC0sMOkIRPGeOYVgt1X6MtgiJyY9aa/sjC9abJ24FPZ8Z5oYerbctG2DEyF795OJUYN2FlxvxJ1Njj5R5oN5UtqIaSRE3xW+CGoLDwZLcNVOb3H9J7G/c59w9Rw9SaWCxQcZC7wXFBbCAb/ALIAkcyo1OkyTzuKk8kp9cWlP3Md7gpiKiGTsFT+Yk7BT98buyj5BTvIPwT/ADTvgpd0MnZU8gp3LL2CrXBJfZBL74yp7PhJG5o6QmytAOe4hFvjdoaiwgtKtMZ2Jnj3q2+kPVplBvzPPv1y0q2J7hTcFb7WamKS6MhdKtfo0vYKtI8aCQH2SphnE/sqXzb/AIKfzTlaTlBJ2VbHZWWbsFW/0OfsFW1gq6yyjrYVanZQv7KtMYq+N4HSE+M5FU1V8BVyJK2QMkauWfNMdqjkGbXAq/Cx4yc0H1IPddgZwY7wRqsVQYrY6lt82+5RPsrLVbBUHFjFDE27FExg4BqbwCb5IQGQGsIcAhwQ5jJGFkjA5pzBFQVDKTJYX8i7yDi1WuwOu2mFzR5WbSmOyN0p266eop/klO8lyleaNjcStI2vF7BCzi9WKy0fNW0PHl5fBNaAGgADIDmBDghzmuFHAEKxaTsz3QRCG07i3Ip9ktUkMgLXscQQVVYeByVGrE/NZ6jr21XRdlPGFny9SPr9h/SPgsUb4KowcUSw0W2VieZftkbTvcAmxxNYwUa0AD8A17S17Q5pzBWjJ87MGHiw3VYDlJM1WAZzT9oLRkX9k9/tPKstm/qII2dTfwLIfpKSz+0iDz4PAdK2TivkqFU5m0F9lWT9Fny9SPr9i/S8FVd8WwqtCq4lY8z7Rh9sfwf7xM/Qb4LFZHcqMJIWJ1ZjmbQVNGWX9Fny9SPtOx/o+C2hwVHZLCiAjqto836/D7Y/g/2+z9FvgdrVUCuRVBR2J1Ynm7YVNG2b9Jvy9SPtSx/o+DpIKLYagRiqV5v2jB7Y/g/3hZ+i1Y+Ax1YNVW501Yu5u2F9n2f9Jvy9SPtOx/o6sfA7YVGAqrDTfqxPM+0oPbH8H+8Lf0W+DwaCqN6EFiebthfULP8ApN+XqR9fsP6R1Y+AqVjVbAKcW4rF3N+1IPbH8H+8Q/Rb4LFYBG7gFXBbR5u2F9Qg/Tb8vUYaDsUcoh5V8jroVulfcEEDEdOzRSW5mMQujk8FZCcOUHCrqpkYLmVw46qHwGxgqsGa2naqEqiCEFqjmIqGOBohusXxeprbG6RlnY1o4uKtX5YI/eSprVaDBPC1jg28Cw/j42eO9resqz+fj7YVnsljlmEjJHsFQwPFSrdU3WQoabtndNsG3dDdhWWN9G3/AIqOOMuYXYeAxWSAV4rF2uiAQvhSmCMQ2RgYGgC8VafRYlIzOxs7aEs7I5bFQPIFQ/1CtlhFlhss5iZKHF9xWmYjlnPk6zVOE7SY39lE/ld8E9pDjE+nG6nOY5oBqnDNOciFTXK/xYyVbDlZ3q2sx7mcrTDHt2VwUobtQOonPeTcIRO5Ep+4KUJzWYp3ku+CjghcJrwxJAulQZgPw33CoH6Ve98rGN5OgvYKGcEwyskAzuuB/Gui0baZGGjmROcDwIBVontDnzTPe8nEudUqTyirTHpBkszJAwA4oeS8/wCEo+bk7JUjnXxE+lBjdKc+NwulP3pycE5PUuYYSpcxG5StzjcpI6d7KLGkcmUX/kV/drN1PDipS7BpKDbMxpDrwHBVyDuyVI55LY3kU8kqmkIL+yBI2pPWrHO8MgtUUj+DXgn+Pw6JmZBcMsrhezyCtJnLIbLEpNPmE2trG8jW7c6VE15Ie/FNb4r3oOze5EwcnewHQFFvc5QcXKEV8ZRDIOViptib3UTNwKA8pGIEC8pmZEgKZSP4ouGN5MdnfUZ8tRcHKF+Dg5WbyXqyl94GX4hRMykl+I/4UZwL5SR0j/hBouB8naUbxW/J8Qp9AwkWYMcJcTfCk0rPLDOxocxt4Fv4uGzRGSeVsbBvcaKEwPgsLBMyRpY55wz4Kzl1XPlVgv0752lA7bvPqom4AyH3hMeKl0g+Cc6F0PLm43iFGC4mZxTWA0vO61wY4ot/I5ObuUnD9lN8ehWgG9UA9StWO0rRjtBTPzICkINZGI+cHZVPzA+5V3qoxePgo3ZyKPdOQrmU6kc3Cf8Ayq1wR3Y7SAOmMKZ5q60mvsBWnRtuFrgeHyRg+PlitKQePBZ0dMzS2eeJrJmC8CzIj+Ovn+kNse8/nIHUMFS3ye1ros8FSuumPMwVBkqEhY6s9WKCGSwBVWnoFVl1qjwepVmJzxTHu3ZK5GzgqaccOMR/Fvm07MxziWR0DQqRjq1EW544FYahdd8FQT9dAnE1J8Uf7FVNHIY0wVGLburGnFChQQodVEQUDqNdVDS7WpVTqpksNVW4HFEs6kR9KGDjE7+O39KTu4vJX16Tr1HWQaahVFxNAiKhFbOGraVXu4I7K+erOm5UezoKuyXdypKCeKwIRWY3q8R1rZbh1oh9LxAqr1kvKunz+kfxd/T1q/UoqMHVq+0JPaVGKvWrzqBF4cdznfugGddULxBTRmmmPqQEiDgNxACoz3LIHUV+6xz6lma9CNFnXcc1UarqbTNYAE5cVXoBKyanNoVsOX3sg9h/8du6QmBzDih3dJ7XMAB4hVKzV4ojDijgqVG5XVtBZqhxyIRun2VQOpxRoUdWNSsx79WySF+9CqO6iECzBXXVTTo5241VdNSHhEfxZGnbX+oVS7hmEbyA0jJXylWP3LFAsvqllDjuxWDVS70rE6gTRbJ6AAsCF3yvSSqtAQuqnwKogRmhTIoIJoJxUZrtAJhyciN66UMihRAxupuV76TXvIid/HWWdlr0lPK3MuDSaBPma+dsUcO/ZqSrYWg91nshWwNqbY7shWjAd1v7IVp9Kf2QrUThan9kK1ekv7IVqz7pf2QrSD/3T/gP+FP6S/4D/hWj0p/ZarR6S7stVo9Kd2QrT6U7shWnLup3ZCtIH/dO7IVq9Jd2QrR6SeyFOMBaD2QrRkZz2QrT589kK1bpz2QrSM53dkK1CtLS7sBW70k9kK3AEm0ZfyhW1h2Z1pB+Vq/YK2GES2i1Eg5IWDSghftMnNy8cwfxQAJOQUDtN2mWJ/KsL6gsV6jW2Zx6XvA+VVKGOkETGUzNCVHaZ3PMsgJO6iLWVZaHqZ2VpPvaFbWMJba8t1wLShw7vdSmGwFpI1rbzh/KFpBzQ7uw9gLSApS2nH+QLSFaC2nsBaQJ/wC9PYC0lk23nsBaUJxt57AWlHOoLaeyFpWpb3ccP5QtL1IFtPZC0qzO359AWlQK92u7IWlZD/39PcFpO9Q6SPZWkMv6QJw8kK3FpP8ASH+UK2b7cewFa8T3aewFbA4Ut3wYFbhlbj2Qrd6eewFbHOp3ceyFbYWVGkX9kKeO0tdapTaYa7UbsK+8LR1gAtNiguPlYKkuJ/jtfo8RxlaizRpPV81sNWwigR0o7sE7iUeJ8EeZRV6x8kG48P3CuCMxnBzQU8nFORNtjFK1OSaLl0UqF9s2P9Znz/FEaLtN3PkymBqjDG7DcAvs21eytpbB6lecQiWHi1UfgcEHPBK2PaCF3qC2qHIqlQc1h+yyI61GW4Xh1EpoLs/imdI35lMudSbXegzJVKBBrwQ+VVu5goB0IB+KoyTV9mWX9Jvy/jvKaJa0+dHyKDdFEDo+a2GrZVQVVEfg+8RdWs93xXcwUbzB0L7asX6zPn+Kpou0ewVfv9aFV9l2voYVtLZK78i01BVRRUcOG5AwsocgqVKqsQTvVCg4KgAWPPxHOAVYn45rNUsFnHCJvy/jobothPnR8irTadGvMEL3NFDgE9rQHMcCOLUA3FAZlN4hMO8FDw5194i6ir2oRWoOlyUU728k4GgTpdN2MMBJEzT8D+K+x7X+k7VIJXC6TQqeXRdrDInkmI7kGPxTCMHBBrzeNAo3jxwmh10uBUb3jEINbdJTPKTMwRkmcfy1TOOICjCjCi4qJRcUzAqClS8Kz+WFAPzqDe9WYfmVn8pWdQcVCoqYJk6fNKGRMc9zjQBoqSiyzRNOYYB/HWStuyMDxnQhQyV25GDyWEAKzDJz/wBlbLHYWP0XEZnl9H7N6g6gtP2qExTWCYs/u4HNKtukTILVFPYgzIyg7SFmtQlmtnKtH5CxWXyIf/qVkGcEB64lZtDiER2GKk2Tgi84XgnkeO5SecKm3SlWnc8q1D85Vr8sq1+WVax+cq1+WrX5atflBWxWocFauDVazuC0pp6zSWiOdkEbHXavVrFqYbZb2mEZhmZWiWZwvd1vK0SzKxRnrxVlsxrZ7NFEeLGAfimTQvikFWPBDhxCs8ONlayM9LarSMn9XbwzqZRaUfG9/dYmcBg0vK0uJDXQVp6wwFW6KNrpYZbKXZNey6VboYi5sjyrURg5ytrcnK2jgrararard5StvlK3H8ytp/OVbPLVs84VahnKVOP7Vyn845f0hpeCxPhvmU0qXKxbgzsLSNgtcbdF2AWuJ4qXNhOC+kJH/wCjH/6CprXY3yaXs/c771GAMHyIUE1hkZZJSyf8jitKsztcLlbWW+KW1WpnIg1eGklaMZlY41YWeLY4OwFDH/VxMZ7LQPUetisHtuVJAsNZurGhWGoBN1NUcbqHFR5UUcckmw1wcy7juUbfyArSWirILNYJWwxlxfgwZlfSE4m3vWnm52uvWwLSL8JzGUzSjC1wDZm/i+5BciAMx45Bacs8ZfFaG9gL6RMJHdnwjavpIH1Gkn9kLTekoomWqSJ4iNQbitNjkAmjYWvUZc55a0kklR+aaoCMIBioHvDSwBRvBcWkjcmcEzghwTeCA3IDcsMlSusf/i+Dqd6o/U7B7bl34LBY6tlbWoKmoBbQAQc46sdQ5IPLqUXKWcai1wRsmkoJDkDR3UgQCMQfxIiic92TRVOmnfI81LimzwujeKhwouStMjCMWlUOoTQObvzCpJdduwOsgrl7K8jHLWQ1Vcs1U0IQDSgNf3tg6neqL/6XsrKm4IV30LALa1bBWJ1013p29ep8gJaFJwUvkqR2jpqtyFU6djmDMKQojNCMghd06Hs0m+4AesYfibllEYzeVjq77FN5YulBycpE6xaV2hRsmKfJGHAKTyVJ5CMD5rPJgJBVqGrAoagTwVAQsTrMX0osjuL6eqP21Z/0P9yu+BYLHVsFYnm1KdyjXEGlULTa2Qt2i80AT7KHQyMLTmK8CghwTXxOYRg4UXIacfD0lqGu/oyWLfG/5/iQbWamjWABYoNBJwCZbdCuliN4AX28ytkhtIzjdQoT2BhQQTXsLeKjstnjhtL6SMFFDO4tifUgKocFR+WBz1AvCAYtrXT6R2Lpmb8x6o/bUH6A+ZXfFhr72ebVVKa2xN2W48QrR3bZ5YInHk33qtarVPM6S0tLcAACNWyiMgUbF9K5Du5WqJVNUWjppu6KiN7FZ7fBytmfUbxvH4eyWFhda7RHEAK7Rxou7LaY7Ae8MfeL/LUVtsjHvIDgNpGeI2exm7G7OTiEbDA6yWsl0T8ugFBzMDWiqqNXdOh7TGM7lR1jFZwlFG6DTDVFanSG0yNa6rbt45jJQQwmWKSMuY4DA1wOvBbapC5besH6S2P9QfP1R+3IP0Au+LBY6tgrHmvlkDI2lzjuCMUDBPm3cjZXVglMXG6nTWnaJcXZuK36sCuR09e40K5WwwycWAraOq6Kk0UVieYpzWOWm3wTXtDmEFpxBH4WOCF8szwyNgq5zjQAIaV01PPA48hg2M8QE+ZwYwVJTdG2YtLyXOzCPct6tDj8kZH55K/BqBYUC0tKNl09JFweWqoB4iqNKbtQjmYCKtJxUD46RRFr9zg5V2JzQ7na6vCpBjxW2df3lsn6g+Y9UftyH9ELviw197Kx5ktskuxjDe7cFBYIsBV29xRNQMAswEXWqNpRc1oARaLxa4bsQqtzxRFts8nEEK9oSz9DaKp1X2kKaC62QXXtzCNnpBaDWA/FibIwPYQ5pFQR+Dis0D5p3hkbBVzjkAptLSGGGrLCMmb5Okp1ploPeeCisENxmMidLIXPKeYmtYDlRPghMs+D3gUbwRuFp3DUViV3N9KXHIOcHfFXrMw9Go1WwDwOqooU6GjJMWfJB4vMNQUb4HFd5IG5qrIdf3msntj1RkGlrPKRsOioF3zmd7KxOt9rdffsxceKhskIZG0ABFxqdV4q7bYz0oOLMxgEeQwOKN4sdd6yVdggfwcr+hWjg4o01ttDHEHbBKks8tMiFyRbZLS/vL/EPkH8EGgkmgGZKOk7SbJZX0sUR7ZT5pPmUyyRXI/GRe4kmpTpJA1gLnHIBMsjBJNjLw4K+6irOG8dd1yuaSs83lM+RXLaOjO+iKJNQvq7uY6zuocYymyPY9hq0hHudzuK74dbpPpPZQwZOr6o4WD/ABrvqw197Kx1Gaks4ozcOKbCy4wAURNSTzKWhh4ORfDG+hK5SG64D4oh4p8lTQ4PB4QlsM8fAqjUQciiHFFryQmWlh3PT7PLdcu7rJ3NO6s0Q7TfwJssH9F2V9JZRWY8GoyPo0Jtmgwz+ZRe4kqSeQMibVxUWjorzqPnOZ4IveSVtlXLYzrVSUUWtaV9Vgk8h5Cv2Cie0kEYFPLRiiYXDHJUJHMdZJgc464hRz2ASRuq0rvh1/euDqd8vVH/ALDqeu+rDX3sqrir1JrQOpqDG3W86j29aEmjQ/gU28MGprwHhvvQOgHlX7Za4uLELiJBqgrk728CUQmzs4OUtgt0crcHxn4hR2qyxzxGrJG1Hh49GaMntkuUTfidwUtttctpndeklcXEq42+7M/JF7k+aURsFXFR2GHDGU5u14lXZmHpQkssbuLQdV6yHoKv6GkPkOBVZxEtxYFGGYh1Qo8WFpHvVy0PbwPM3FPs1WZxOzCDnkg6/vZZup3y9UduwdT135Ya6xlXnctMPZag0XW+AJ0OS0ZEJ9MboXeheoepCX6MWpU+kRZxjK2NWGS5PStob/PrqQ9uYR0VYHQSxCRgNWqH0RQei/5z/wAKD0Yds/8ACg9FHbUHoZ7as3ob+2rL6Ie2rJ6Ke2rFvgK0Yf66OVi0Pwn7C0NPLcLpYul7FFaIGTQPEkbxVrm5HVV8OjGHLbkV+ToGaPi7ziUXOoMSULLDfeKyuR1YnViEJtCwO6FwQfZ5G9BXKaHtTf7slclpVvtBAlXmEokk4VRZbpObUUKOv712b/F8vVH6zYfYcu/czlDeeNlBout8DyujZWfyBEmjnpoYAHXlf0BbGUzYrn0sgQIodQxVzTcvSAeZsFMiZV+CcRW6UUUVdmZlg4KQlxbuOYTgAaEFODVQGqleCW5Z0ripHZORP0PsR9r/AFFBjC5xoAKlG36YtNqOUjyR1blcj68SrxJKvO5d+Qy5l1y2sNQk0C3oKAOSCq20wHeHNVzTUYRJC705Eg4YlXbYDzrwot41feuye/5eqIOm4hfJIiXftZld0BCNt1vgv3Yo799twHgnVzqE12hbUOLFyX0vsn6tFiFQCnBYlU0ow8YxzKtTy5roxV0b63ULWL7YDC9wAkGTVU6rmJVWElSQSh8TqEKK0wNniF0nBzeB1VY4BXIy0xVfxRFAvuZYep3+oo2P6M26UZmO4PfgqvWx1ozTBiEUIaOdUBV0ZIzgdY7tmdxcVyGn6cJSFytjgfWhLArsR6VUkblV0b+feCulfeax7Zj28/VF7dPh5yfGF33UZJKBNhjAaianwMkm6g4lCCW6g7NlenAptKtqCeCv6MnF7OMrkPpbD0Wn/dAOpwOrFyraoXfycyoIRbOdRVUTabvBUZqLZJIzkRVZ6qraX3MsHU7/AFFCKwQWAZym+7qCxVXU4LN559YwqRTs1BVtk1eJXIfSSXolqg7RFn9hVa1G+4bk19nbQ1IPgA5qfJ9JrIxvnAfVFn9PjiIheXfVWgCEMVTmi4knwD34u2WqGHIXnJ7ugKkpyz3oBu5OOXzROj5G/wB2QfguS+lPVaP90OPBVOeCoTirwgPNwD9W1U5FbBqcFftjz0rvZWCu2o9XM2gvuXYOp3+oru36VWrhD3pvuUYucmS476hEuqBguTs45+wrk8oQQVLbN7RXJ/SGQ8bpV/QsBBxoroo/jmr5woFSzP8AAxw/SWxySbn0+PqjT6Q+1GF3xUF9yvHnvlxybxKig8UXncU9+sxzuKqcSAqx5nrBTTA8E7irn0mk/VBWyD0DUASjNAymJBRHMvQFf+BqDIHFVlJVWHqVWqk9dR1bSEP0BskpyZG8/AlQzaQkktDnnlHlxLekqMyVhqWbid6JDWqlAN3Nc/IKgq80TR4iMVscj5SPFVtUh4q7phjuMYR/oVixxNUDdOGW5GSJ9E1+YoeIRbzyKOaaOaagoaU0WyQnvzNmTr9UIbfpsvgAuMFy/wCVRX5KnIKgujnOeaNFSmRC9Ji7gnPwGA5kjvFY4+5SWeF08gu1K6cVfb/WDqK2SK/BXPpHJ1hHkm9QRoF0mqdEGStoXMdUVWjbeKWiLueXy2ZKVjeUszxPFxYnsNHAg6qxO1Yk0CuwFq21Vq2FSXmbS5D/AKQGThZpP3cQqykoSWQNObTRfWB/KOY9+QTG+OalUFGCiJVHFXLb1jXWQlUtdmfxYQq6JCIqU0ws2d2aa5pF1PaTskotwIW9vP8A6N0sy+aQTbD/AFPGi9CuYx1J56sb0DeVy3JUGJCEMVBznSHg3imxi7H8UXZ6q5BOPjbIUMRBu16SjAwsikY0cKAp1ocS6W8VU4uRbXNObVU0+48QFWzxni0KjUSKK82hQOIVpsj70UjmqyWzY0jCA/zrEJGmSwStnZwBxT4g9r2kEDeEQ4qiqtvVVhVHjmbSFl/6LQcZ9gdslXnE9KpK9nELv0q95T34nAJjOkonmUeqW1uurlsWV/S4IDRiqVWMaooqte4+5ys8x7y3DfU1W9h9xRrlQ86oou79ENa81mg2HepoAJOQR0ppmV7DWFmxH1BUha524K8Tzb22/BqrgzAai7ABUxeVHCNzU0YMFekp7s3KmRRdK8E7kQ7A0V5gNKjiUGuwVNNNPFgVbFD7DdVFgdQKa5TWaQPhkLSOCltdi5GZjS/zm9GFzQVslXyVSTViRVUdVCuSB3IKhR//AC9+j1m/Vce0vmi21tI4o7RfkU1mDBVOOoczvqpa4+ZWwQnhIibA8cHasKKpRZa5G1yKc01BW6QJkjdkhyrixUwPN7g0yyppDPsP9Tf6O0E9rDSWfYHVvRtNqxyzKDGBg5o8eTLcEXlEnALe9Rwt3BHJmCLsXGuurlW1PH8ixRDAA1VdmqaTiPFiro6z/pjXgdZR1ACNxQLUKlCYg8q1nQURHfFojPRjVBjsXVKpXVTcm+QE0Y8mFJa7FY7M+lyzNc1vvcSm1AauTtPUQqt59VI6QEtoFdt0Y4OHMrogHhIF3mUa8UW2+XW5hqDRB2EnxTZG1zG4hFnSOZUYZhf0joaCYmsgF1/WPUzuzTL4mGscGwFyMBccyqknmDx35J0jsETi9RwN3BE1DMAiTUnXVS2t5bCKkBTwm9I0UC5G1QyBtBLFVUR1UtcB4tVdE2Y/ya8CmsdQlNcMwU079fdELQDShqnH84TnfnClhd47aIRAXhU78E9zqtIGKfLW84Yo+dCf5wJ/nAnn+1CcNptpb1KWLJzSjfNc1WBh4tHN4kBRDM1UUWV0IyYAlHumI/zDmV0HL0OCxmGvFSOl5aME1NCp7pJidQcx0ZwOHBMlHA8FvbzLlsmsTspRfZ1j1LGjtD2m1b2N2PaOSNqtxc4k41KoLg5l83neKE+Z2GDUyJtSgMI/ii7EmvOF6QH8ybLC7Z3KrbC0ZCMgawnPfAWDIFfZFm9nmS8o7YJoUW5FPbnig7IprjQ4I2mdsQzcrWfEZUK37oFb/Mq3+ZVv80rcxhcYlbHNBEatg/IrZ5pWt2bFaw0khoTqEmVq5HSD467wrsLW8NQGaYMBijxR3J7s3anzeIK8SpGWmMPBFDrrVGTQs7WCpoPmnMtUrHjmNdY+5y0VEl697kOCEdqlYMg405hCvbMnxQOKocdRsdvs9qb/AGbwT1IPY17TUOFQfUq5BBYme29XIS85lXiSdZlfQKoHkhRwtoMSnyZnwFS/oCLbO+h3IfUoxmGu5nJsi6Qr2h4NeKJrl4g+ZWJ1ELc5EAPB2mq12V9YpSRwKtitatatatrwRfVrjYGh2StitvFWxWy0kmWQkcNyePHJXKWtr1ykFU2JvE8E95q884NschIGKMlsBJxvGqwGrNXLDM+gN1taFCTSknMpfO8UKeQNpBukZgPK51zZfi1BzajEKhVW0XdegICcXRC4fd6lf0j9Ip7pqwPujqCuRho1ukddYEImVcsLsa48978mlPOZAQgmcyuYXenDoV+12f2OZSGylV0JDzA2d7eGH+ZPqU/gnjciNyrgVdVSaY6gm1zTKYFQlxocFEcpGppycm18cJrIgXYEhNjVZmq5ZnDpRe4opyciijwK5OxPaUDaD7SrE09A1Zr7JtX6Tvkr2lNeCLXkcaar2kJjuvc8swOLSgRUGoORVDRUfa7J1SD1J7i0RarRvZGade5X7VJId2CvPOp0r6D3lR2eOpRf0DnPfkEc3mihjyF4o5NACJzKMDKgkGqle0i+SvtGBnBvM+p2dV0Iz2jrxCivnIGqZxCahwTeCuyCivGhVzUOC4J5w+ZonY0Feo1RG7PVwJ+KcWOa4k9ZVXld+Yi+J6qcQm8E3gmKMbgom8FGEzgmuffVbOw/yjVUuX2Xav0nfJXdKs5jrOxro3UcaK1k/wBZVNe3bFShmw0T4yajnUNx+RVHU+C5D6Q2bhJWM+pPIaCZDvmk/YLk7OTvOoyvoEyzxVPuCLzUnmvkUceLzUqgoxoAROfMCZTBoTzpTlS4ANWGulgsqroYe2Vjq22omRxbxRCKcnKpqtrmcm0OPVXgdyvOrkCQcERJsnCgxVXOa8XmhUxBN05KgqiPcr0hXf2IPMrTwQYTQanJ53lORTnGgBKnfmLg6VHBHcBqd5X1SP2RqvPcOhE2C0fpu+SppePr1YLBV5JxZVtEwZMA92sOzQOLMCnxmjhzb7LpO0MkYp4Z24GN4d7wUHsa4ZEV9SL9vgs25kfzV1oajI+6E2CKpRe6pPMfIcAo4vHxcicBgFXWT4rSVO/8id+d4CgHjEuVjjo4wNkHByhFOQs7YON3egG6/syyKuhz7ZVXasR1oR2mQEVF4qOT/gobjRFpoVXVXVXJFxc45VoPcotGx3A3lJjk1W976k3VQgW2K9Gfzx4EKJzGyxytkhkyeP8AdZj4a6vVbSwDgqTyg8Oa5+QTM3mvQFBAMA1qaTRoJRcalVsLPZGqsj+pE2aUfyFU0xH7WrZ1ctGWkgDpVkdGaGNjmjcDio6mmCIyNU8Zt1NkFCjHiMRzC1wIQJwyeP3XL6DscnGJvqPYLNIY57XEx4zBKZbdPzzxPD4iQGEcAEXzEBCNl53vV52suyQG1ItzBQKqJTjuTfzuULMm1TGeS1Rj8xKH5Wp+6gUrfzqR9pax5reQQJ1fY9l9tV0XJ7arVFYjrVLdMODzqc3M1Ca8IZt5pdZwBxcf3K5LSlojfC2QNdm5RS5QBpTcgnRzPsjjWN2LQr0Ra41LcFRZrbK+tsPQVW1HqOuqb1prM8TwCe7AbIROuthZ7Orvz/ZVYn+yVTTLPa1bOp9msjGM/OcVN5wqTfRyZ+ZpCjdk/wBxQdm0FMORIT6HCoRaS5vMJaW7xiFZH6Ds8BnYJhXYJ9Rn2bQ9qmi8dkZIU0r3Oe9xJKc1jnmpqi+QuKHiNyGsvOCEY4lSS5NJTnZlMZuqo4+AUYyqU7cAE92bjzDq+vRdequr7Esvtr7Om9tbZW0sl9p2j2zroig8dO8LFYrBNdE9rgMnfNGHTMz/AMsh1XUXaUiRvSdQRzWaxKramdRVLVqDUTmjzLqvFVsDOrUayOAWBVNND2zqoz36u8Q817PFcQnfnAcon5OunpQe3aAIRqXwH3FPjdR7SDquyAqRjwY3EXSpNJWB9ntT708G/i31FczRjY2/2z7h6iCqglPAuHJCGGgzOsyOoE+7jRoTBk28eJUcY2iEMmNUjvzU6vBHu2L2tWGr7Bs/tr6jN7a2z1LF2oDSlo9s82jx8EDJXX3PapGOAo6pHUf/ACgbVICN6pkNTsbS8UqKNRDfaNfcsCqhYlfW2+yVdtQKACqbzvAVsDOrUKSu6lsOKpp136p+awCyC2ar6nD1+AkjOy5NdhIKdITJmYgPanMNYtocN6IKLrvS1Gy/SKEbpgWH1FqbA3jOf9JWfQ5Fji4hEuKJOCe/GU3Rw3psTNhoYOJUbMtsp7hnQcBzTuT3ZNKkPAI73pm9xUQ3KMZMCaMmrv7MPzBZ66/R6D9RfU5/aW04rPUJrdO8HN5TtykCeNyIzCo5b1gdV4BzfHbkoNJNIlvMl8oZq2yPpDPE8dJDUbLKH21zD0XgmNaGtFGjdvKpidWyVivrjfZKpIr767hqKedykO5O3kBDe5R8CU0OwFFSwtWz8ltyMRMZpmRRXfpBIOEzvmqALFd7CDoIK44lRH8gCj3EhHc5SDpTm5g817MWuTJxQ7L/ANig95vCknHijwIIRg0nZZvIlaf3WHqJW16OH94//SqOlHBxVBQhBxJohmB7ymR+KLx4pzvGOsnJPK8pyjG6qaMgjzKI6u/M6xzPu5D+oq2ScdK2iFRleOod2z+2dfEBMOYTb2ARawnCjU+nihSNpsg1UpwuhOJ8VpI4blOG4OFCpjgaOUgPiD4p9KUCdWgAJRoatog7JfW29RRdMAN5VwmoTeCG4DnbSBsQW7ghDa2OOVaFXmKn0im/Xd81gNWwvq0PtHnNO5eSU5uY5glZycmf5XJj2mKVvfm/JXXA9Kv2aN/lNB9RHX7HKzON5wO+uC75N7TvmhdCDcX5DIInDIa3OTQgMgifB0lb16sNX3ah/UXeLQqPIWSNCh3ZN7Z5tSqwyBC43CuCoQAE0voMCq3gcHNzarriKYLFwpitg0bTL5onO4OiqDw26Rhgcd6u3moUVbY32SqWuLr8DiqWUdZ1h1hiJ4L7zT/ru+ZRc0U4KjSUQAu8RHpPgAQsKtRBodZ5dkm9ooV9SimAwcVf0VZHcYW/L1Ea0We95df3CDpZiMQXu+aDxnipLxT09Bvh++t61lTgsNQP0ag9td5tCo6usd2zUyvnnUjcVFdycOpQ4UvqLEG8ozveSoxleTL+zewTBH+atMQVE38xPUVFcJq4O3BRXcC6/wBIwWC5O2MdwVbaw8XeBxX1ZvWdf2exfeaf9d3zK70FUDHeEKr6tF1nwJpUIyKTcp1M94bmSmWbRMEIzacVXQli/Rb8vURz4YyBW7VUkm9p3zTiqJt7aKaSaOH4DvrOsLAIk0QGAQP0ci9td6mQ3cNVfcvrkvtHnAsdjvC2TgKhEtY65g5BznC4NnD5oFoo2hu1K2SepNDsg0g0KaAXNIeN9aqocAwOJNN6FcaDCu9by2l4LBfWW9SpaouvwO0vqo6zrabM+MkVaVX6STnjO75qjAsusaq2aL2j4EUW0VY2lwtZpwNKrRO6dvYKsg/7aeNzuFKIzRAHivsOxfot+XqIDDTeWu+S77J7R+aNKNTuKJ3pwXHw/fWdaDXgubUcEOU2WBNLiOSAX2Az20GwTdaDnm7dA4XVR4JDXDhdXfbwwqV9al9rnUaUA4ZEKooRinCne6+7rRBwYKCrcQi3NoFcCCOhENvhlW1xqMESfEwJ4ZBPNSWDs70++TRoqeC2DVYK9am9SraYx0+BxX1QdZTYXEviZKCMnKzhlHWKMu4hxUJs7mR2NrHEUDw4r7df+qfmqwiiAOLQrI2AtNjvPpS+XnNfV4x0nwJVwUBxTzm5P8pPaS68i6EXld0PYxwhZ8vUQG36OrkXvH+VUfJ7R+fPCPhO/M6xrwAK+xB7YX1OfrC2tWKrPJ7XOqHIDPKqAbdDsv8AZAC6XYHG9RVBN4Y7OW9Nyv4UrRNGTq0BBFPim47eIHBNqSSPcmAnawpgsFgvrTepVtMXX4HFfU/edYbZwLoX3hl/Wd81SIIn4hYKsMfWfA3OtEmpVEUXWmNu7MrIBXLDAzhG0ft6iUn0c7hMR/kKqHHiU1gqUQ5MnbhzS1NODgmnxHJ41Dn9+Z1rHX9it/UX1OfmVmf16yiqJgBD2kgncaKAjxHdr/woNzHdpRltYwWkDjVUBxpkae9Ag1eKAHcU1hrfvA1JoEw4B1cFCXvbKxxbhg11OKsu5j/c9WPzMvbH/CiOETHN9p1V9aHslUnb0FBww1HnYr6mOvXSMBfb8n6zvmqRjXWKPr57WjEonxeZvQjN4jF2XUuWtUMY/O8BUaB6iONjgkaP6qa+eq6VSIIvZRUcU6J95hoUySjX4OQcjzCMipG76pp8dqidk5cCnc3v7OtA6/sdn6iHcU/tcxpkcSd6jamDIaq5uV5YrAkLFbNOhZkjILac0CmND1IVAoLpzxyCDN1CAu/SD2f90MSsDq+tD2Srr0M3VLd4BooM4myR9ZqgMnV6wm7wojvomHJw1HV9SGqqcG5r7df+qfmsAtyC72zr5gCaMsU52XOaAS7xRn09CvygngjadP2NvB98+71Fe/Q1rEYq4xOonNo2mSawEU3K8y+NckWD9pqZMKtNUHCo57hkSpG704ZsBTD4zCFCd9FEcnhAWhlDvWGrAqmio/1F9nTe3qx1EWmRo4pxTjv105gZmwPB3OTKf1MY+KZugjTcuRjCb6PF8D/yg5xe1jWVAqG6sDq+t/4SrziAix5DlUV1UTgnBSNycVIMzVOkFVWwM1YhCirpt36p+awWOoNYwlMR3BPO9E8wqq3lNYM0ZHcGjILMp0d+3SjMXGeo2jbTKZCxzCeCisGlH2eFhAZvJzBCo90TstyLHHW6N9WlXs81HJhn80142CiOcRrC+vRdevBfZsXtr7Lk9vmVtcnXzcaLArDmBAZqea6ImZtvY8FbHipfEAn2aYxvzCwOr63/AIV3xArG49VHO2V9QZrow9SrpnrkOsLvTOvwLWCpQGy1F5qVVyslo0Sy2WwF5e80amxsDGNDWjAAeo9y3xT7pI/ki2YkITR9KLXEHWWOqEJMHYO4qaE7W2Ex+DjToconjKiB8U1RGbdZRRGr6/H1rZ1/UIfbX2Q/2+ZWd/XzA1UCwPMHuQAKaJBf8UHLj0Jge1kYIa6gaM8NwVyMAqO1wvIF2YCoOrEqtsp/KiJXVFNVVTZk+KDhUcyqumi+oM10jPUq6Wb7fM2GajzCmtzUcQzT5TwGsl4XI/R6xM/ugfUjltDCYZwv/YrYB4YIsdVCVl5uaproi3ZfiE14rGa9CkiOy4hOGEja9IUb8L/uco3ipb7wmO8VydwVdTm7kRbmFYa/qUHtr7G63nmAknWBgMSsbzsTqz5mfDVxFU3lYp4sWsIqmWiFj2OqCMFHBZpHyFbTitoottjSEXzvJ5hjOBw4KOboKI10KpYWa6RO6lXSrPa5lbnMomNTWbwE52DESak1PML5WAZ1Qhs0UQyYwN+A9SBbNG2iz+Wwgda8dhzGrk3UPioSNMjPeOa6M4FMkwkGPFVFWGoVFJH4jiFI3xwHKI5ktUcmRY5RO3EKBwrK83f5RVQRyl0dSOkczvFm6yqaDb7Z5oai7LmAO6+cDgV3NPfuB7HYPYciEYSeQa+IV8W/UKW0vrKfdq2yvrK2nno1UQCCLDgU3xXn4qGYVaaIjLEKhX1NmulnkPBpVdJs61hrFoIGR3K7GXVyzxCa00pVNjG0Q1NyYC5SOGLqDgEXZc2pquX01Y4f7wOPux9Su49OzgCjJNtquynp1GN112SEjeUi94VOa+M4FMfhIEHCsbqojMaiFNH4shUzBixrl3dK6KRgjeOZhZVTQUXWeY7xWpzsSdRCoM1jrGOaaRmq12gjuoqK7rqFtFfWPcjGwkK+aFA70Dqcnjcp4TVhIThhK33hNl22GoK+px9Wu7YZzwjd8lXSTFhrFljrm45BTkUuBTyZvIHAYLeUBgMUXHHm1Ku48Ff0rJP5qP8Ac+pXK2CK2MziNHKsd7hrMbrr8k2Zl+LPeEQec9hq0ojB4qoZMjdKNKtNQnOYSnxIjS/MxsnvVNBQa8ExpIpqPFfzKpQdDf3hw5oQQQQ1BcFW0+5VZgnNxCcDmQnj8yenbwFxYmb2KB2YQvUb4qpZI+rXd0Xajwid8lXSI5l+VvQEEBgMSnO8BkBkuR0RJOc5n/sPUpttsE1mflI0hOhnls8oo5ji0hXXkazC6jjsplpZebS8nMNCPASM8VxRmgdezV2JzqKmkweLeYDPZvZKpoSz9WvAq9eL3uxBp7ijx11KDIw1N4pvFM4pnFM4pu4pvFN4pvFDimoKtoeelVbVU9lb+YUU2eySSOzGARbPd6iqWdg6Br+yLV+mUBpE8wkk8AE7ySqSvHTz95VEXOAGZwCFj0ZZ7P5tgB69/qX3Pb2W1g2J8H9ayeOYYXUPilMnZtIsyxHgAY3uQFglAAxCuW9nTEOZW1wewqaHs3sa9h3UqndQX/nrLirgTWRYuAqcyo/PsUfpMfaTPSYu0mekxdsJvpMXbC/v4u2E30mHthM9Kh7YUW+1RdtQ+lxdpRMAJtEeOIUeN2QEo3yTvKowag7Fmfk/8KqI5lLCely+sM6gqRN6tdND2n2FW2v5h5V3DAavrcvtc7eVTV3ZpyAEbEZ5R3u9TBpTRE1n/PSrOsJ/JvZI0ihLSTxCLSQeYYzcceooO2XIPBLEWmh5w7meF9Teq29g4RDmEaShA82qaLs/sDXSJ/slOjJAPH90SaBOdi7BMjbXIcSq4NFAr0bQeKCCCFEE1NTU1BCz6F0BOB49j+Tyr7r3AqshFa0VGjXXx8enemuHFXhVpT2Zt1OjjDBlWqcbSyuWAWwNdzQNoPV81V8juYC8+0hRX53u4nmEoDF2upXJWGa2OzmN1vUPU1thtN+MbFrff9khfnHv5uFyQ9RTosH4t4qOdm4hOAJbiNZGqaAERSFvUppZA2WVxHSVytve7g2nMvacYODAqWCAfyDX3l/Ug91So4huCA8UfFOcakknUb7R0aiiinJ3FP4lO4px3oly5X/pzoG1+aZc+KpGetEzHnEJjsHtVnmyoCnREkGoX1iP2gsNdNAvHF7V3mU8yt6EZh6dxPMLjggwVKrlrfPOyKMVfI4NATLHYYbNHlG0D1NJsVlPSUJIy0oseWnm0FyTLinRm8woHB+B4qOYVyJ3hObuqENRC2wqSyHmX/pH7mqlliHBg1jknVR/LhqJVAS7JPZZTKYiG8UDObprQDwW2ELT/wBIouMcDHq4CEHzAON0FABwkpsmia8VaU5vOc3M1CBtkdN5B5lNERjjKFSxPPMra5HfzcwnNMjai7oGupXLW99seNiDL2j6nEaCjHGVXXq+y+MxzjHsuxag9t5hqE+I7J9yY7B2yVHIKjApzd1QqoseqcpzL/0nf7QCpEwdA194ciVUqWQXiBGzynrR1hyHdEo45Ke1WSSMgNizuhOgndIC69WrgcjXwBDCm8mXFzhIibtV3V/04ZB5diVJSOlERFyIDutOZkVueo5BgURlzK4Ktrj5n1Wys4vJVNG8ysjj06nPQamtwZiUXGpxPMLnANFSTQBDRuiooPz5v9o+px/oWHgJVtarj8MjznRGrSo5v5XotUkXiuw4FMOD9lRTf8hNcaEVTbMDd5nKfSp/6oVG6y8XQK1V1t+0PbCz+bNWWyvrZW33ZXn/AOwU9o8d5oqlfVHnjgn4X3F9MgU7yU47k7yU7gn+Sn+QVIfyqQ5tU1KblJUUajH9GLBHIKEQCoUcWnp7NfZCWSubV+ApuKFkPc1Gm5vaag9KL2EnePlrLcinDxsVFKuBRaquVbYOZWWyM6HFXdFs5jpSSHsGO80Xc8pYXNeRvaahRxYZngE+ToHAc2pXdWke6nisVn/d3qeBoqzspiZFisEJGFpRY4g88jZlF4JsgrGaojNSRGrHkK0N8koWmAPyO/mcp9LD0z8x9mPKx+MFNO+sjy7rOrjqitlvhjnjD4g1z3tPwC0SMrBB2Foz0Cz/AP1BaL9As/YC0Uc7BB2FoZ8rZDYY6tWiRlYIOwtFegWfsBaK9As3/wBQWiDno6z9gLQ//wC3QdhaPg/qrFAzqjGoMt8FtGU+DusIhzBupQJtzCmyQSqOI5r2dIUcgo7AoNIu70XWlx4DmX9LQs4Rq7o2PXQEpzKkmg3N3qWWtXYc+62qDtDSuDCCZM/U89yWTrcsVgqFCVl5vjDwDmGrTQprsJh70HCsbgQiF9WPXro0rlPpSP1ieZsaqai54A3lXWT2jqjb7vC39ARy+amCNwHyXfNcnOL2VaOWR9x55LVtSHmcp9JC3gGhXLBEP5dZZZnlpoUSanE8+p1S/wBBv81fN31PB0DG/eJgsddDQr87ff4FzDVpIVcJW+9AWToJ10heeAKv/SJp6SefgsTKfy/PchZNGQQ7w0F3Wc/C92/Ry2w7+TvDrGKvX2cQfiqOXL2am8YHrHPowLvTzzOW+lE36lPgrkEbeDRrpBTifABraaiNAF97OQ+p/wB3R+sFuHM/KVcNR4p8FSwRjoVBqu2Gd3BhV/TF7g08zHXVX7fZLIBUlwdJ4YEEFHRmnLRBuY+rOo4hXXGmWY6iuRmxNGuwK/MN/wA1Tm0CpZCeJ5ndH0ie7ypSf3VAsdWDRz7jb3wVdX3aH6rvU/7uj9YasOYHgseiw8RuPgbtjjH8oWGq5oi1H+7K+tSv4N5mJOsNa60SjZZ4o8p24IyWue2yY3NkHi45+H5WyxaSizi2JOpB8HTH/pKuOIQmhMLztBFpIOY1l7w0b0L9BlquWJuu5BK8/laSr+lmu6dWOqsg5285BVOv7st/Vd6n/d0frBYrDmFuIzCZaYyx/jIxu59XBAQNGu5oK0LYndzKRNPGvz1OtEoYMBmSdwXKyNigBLGYMHE8ULBoyKD8wFX9Z8PHa7LLZ5hejkaWuCl0PpeWzSioae20oAktNQMjxCdE8PaaEJtrgD2eOMxqxVZvcdVXBXbMwdGvk9D2p392Qq2ov1VOqsh5t5XsB4o5n3YZ+o71PicItHQmpa6/Im51xTSNbRmUwtoHBFj7zDQhR2yItPjDMIxnnVnYP5gqR66aHI4vC+oSu4u5lIY/ZT5pRHGKkpkEfctnN9x8dw/MeA6EZbV3VKKshy6X/gRpmwViA7rhxjPHoTg91nlaRIwmgP7hcm/oKdZ5rwy3hMtMXKw+8KhX1gLEq9M0KjAOjXc0K8eW4BUie/XVVceYXGijLxGXhjN5ULcI31CBQQUL7BJYCQJmPL29IPqczRFiMMLgbVIOyE+eZ0sji57jUk6iU8bynAYFOriSiinxSB8ZIcFHbYrj8H7wizqR5lbVH1qjNf1KFvF6uaKHS5HW+dsccQqckyyA2Wx7TspJR8gpbVaWMA77J/kbxUdjsscEQoxgp+C7uDrfYGfWRjIwfnQmY5rxR48b/lGN5aQn2aSrct4Udri5SHxt4VLS0dKo4q9IXHcOZ3izxcXkrk7B168eYXGgCbDGQP8A/U6R5JRRRVVNYrUyezvLJGGoIUWm9HCXKdmEjPUyOwWGW0zGjI21U2ktIS2mY4vOq8UANY5jo3h7DRwyITZ28nPS981dxbi3mVtQ6Atka6vs0fWVyWioB0KoRqiZGDpTxZBZbMAx1KPfvT5Xse9hc55pGze4pujrNV2M78Xu/wBvwgtt636LAZa83s3SIS3mPaWSMNC05tKdE8hwT4JLzD1hRTubI3B4zCo/rVyzXt7nDmcrpOGLyWfMrk7DGOcXmgQJ5NhpxcoZHVdaVCcG2lqtTG3owJR/JinxPLXtII5kuhtKsnYe9nCRvEJlogZNE68x7Q5p9S8WaNi9uRXqvcrzkANVUGYDFyLiSebRf2c+W4oObfixHBEaiJnE8OZyumo2cGgIMskTODQiFUql53ktJRtT2Pe0yPcaNjG8pthZy01H2lwxO5vQPwz7Za7TbrFOBK/bawp1+SC0xGOaPBzTmE6J91yls87XRPOX+y7qYC7AilQuSs8DeLS/4n/xzO6PpDLwDg34K7E0cBrqdTpDwG8q8DHEMN54otbqKmidWNxBUNtj5O2xAndIBiE+xm8NuJ2TtdEZ7DJo6U7cGLPZ9Sg1hccgKlG3aXtNpP53lbAaOZcj6T4F8Dwx+0xRWpl6Mi8U4SDZyKuPxFOZ3R9JT7YCwA1uDCRvIaoY9FC1XKzPNL34izaRtUU5YBKxwJPFqZcNp0cyrM3w/wC4RZRrxiCFSZtN4CPLUP5Gtb8BrDWOccgKozW50p3uqqDUU+Z4YwEuKliZyloFxiNoNxtGRBQhgDSFHxCZvIUO8hRcVHxCglsz4JaFpC7mtckXA6zo3T9mmyYXXH9RVRX1J7k0BbJd9wgdZWJV53MNxrvAi+K8UYqOZnRHleTMYJu1XdAva7rSeAXL6ZdL0k8yyW3RL5bZFfrLs+5RWWAQwMDI25NH4jJ2qO1xPt1ljBmbjIzywrJDYG2y3eOxt+gV+Z7vKcTr5LRspGbhdHvV1pdzHREhorRSSCjk5tACpB+ZP4p/FO8pO4qTyk/ykZpbxzpruuqu7NCWOfe+IV9Sbn0cu+XKBqq7mAihThizEIjMc065Q0ANar5MpwJarkDdfJaOnfwYQqB79dXLuXQdlj33Lx9+P4m9C7ox1YYoWTQEoGcuwFjrvmKAe0VycDRqzV1pdwTsUTz3ORbtOVTXmX/otD/I5zfUn7Ch/W8Ax+YTQU1NTU1N4JqCuQMCoENV+yCKtL5XIWaiqdTrZpOCHy3ivUg1oAyH4rk5S3UJrayysNWwqmqi7r0hyubKV6ujmRWmcRT1uEEmisIyc5WLzjlYvOlWLfK5WLzpVi885WDfM9aNi3veoGEiGMNHO+7jv1j6k/Y8PtrFYnwwKjLaPaoeDlGYr20mN8WqnntkNYiI20DMMwnZBuakOTSpj+Qp9mvWydl0kUjH4vlW1HjBDR9mfgTLuCmtE75X1LnGqkAyKuYFQ6Ut89mk3Wd7/eEYCYC2ha5NeASSo1Cci5Mh8RuPFHWUUdRIoef93H/rH1J+w4v1fB1qiFRtac0PNEO5y2gFCVUlQv8Ao1Y78TH4HMdJVkz7lh7AUTfFiYOpoTRkB+NacwCoHZwxnraFYn+NY4D1xhNh+lNtZC0MYH5NX2xa+iyPQZp+UAU2Wf6RqYYnBwyOaAdRUGoUx5uCw52C+7R/WPqRJovQ8tqhYHvaQACrdpiW/a34Nwa0ZDwdU00bhWi6ebtlfV1ivu3ZOo/M/wAD+9du/UX2tbP/AIj194Zupn+katp4W1VYeGwWkNEwcjZyww1rdc1DSei4bWBdLxiOB9R/u9MOJb80XlzSALmAoFR/gnLAl2aoK15vjngF9WWJX3bsnUfmf4H96rd+ovtS2/8AxHr7xT/4f9I1Uc5VpzR4LZT5thgJNKpw0PLC/OOY+o/2C/2ghyr1trHwOCddJGSJAB5oDZK7xgu8IjPKq+7dk6j8z/A6fSy3e2q6St3/AMRy+8do62/IasTz8PA4LuS0mW6HbFEbRDbpcr0w+XqOGfRe1P6vmqzHFVcsfBANIu87Ek5BUhW0V92rJ1H5n+B2236Tnt9ho/jGprE/SdptQMdyExXDmSV947T1j5DVW9qp4bBXIgQUf6BfJ5cx9Q2saXPcGgbyVYWZ2mNaP9JHwKj0ZamQWRglwq5z1JpTRcthkhYA/exSCzstFQWSEgY4inEfgjcdTIZqsB6FtFRM+jlka/gfmVB5aZKKxuDh0fjA0EuNAFAMnF3UFF5DlDo2xPnMD5aZNan2udkTbFcL8g5TM74IRJGc6ZhaO0viS6CcZSN+RQP0othGV9YLadzcObXnYJ+lrIJzbGQRV3tJK0forRsVihnBEYzOZKhm/qpWP6neoVLRDDJ/VSNPuKYJKFxwKjoLpVmtkge9tVYXZxVUFiZBZYIRHdJJR3Khx1nnlHmFOGSe+yOeCKDNYlWOzaHs0Vpn5J10kVYaZneEybvkEjJGcWmqks819jag5hNmiD25H8WRo51CRUjJUcI5vc4ZFNe0Fqa5t1wBaVBFJykcLGv4gAFNv3C7HhxUdqBls45KU7txTh9IbWH5iShQIV55bWhOSmLq4KbgE8G6VJkApd7VL5KlyAxUoxulS+QcVIM2lPG7nMfomSD87JEHPNW0cnto9rqEZEFTRUZae+N8reop2Xonhw9QC2OIhPLI5Wg4gVT7tbidvjCfYHxsigBL95TtLMpKyJvUxBmSKe3IqQJ3BEflTvJCk4BOOeooooooop+4KcR5YFSmXEbKtNmgZHHEXR/EK0W2cxS2Ext8topqdZSCAXN3hNljD2ZH8VXRzutCtHp8LCG7cZ3bx1IzAlkt4cDuUjm4FSRTLSGjmMhhaTXN7lHNLJLI5we43j0lOY8BlacVL5wqUjG1OCtAynerQCTeNetWkZOcrWMnuVsH53K2eW5Wzy3q2jKR6tbs3uUzs6lPOYKPBO4J6JUDLByuTycSpNHck2ylrnvz6laZgQ+RSWu3chOWubTMIwODmPLXJlQy09sJkrA+Nwc07x/HjKGMGd1WiyWSBkMpZfdRxb0AK2AU7oerVdry7v2XdFOXaX04lQeZ/wAxVmOdmHaKs26zM+J/5UG6zs+J/wCVD5hn7qLzTFH5tqZ5ITfJCZwTPJCZ5KZ5ATPICbwTeATeATeAQ4IXMgm72hM8gINyBHUUbwrLKB/K8hTtJDbRNT2yrSdMSQvme+N0ZJDvxQFgeVRzn5UOp0Fke9ri13EK0u/9VN2ynStrJLK49Lyo3E3wSoG/2aiGULVGMoWJl4d7j7Kj8zH2VF5iL4KH0eJWf0aMqzE/9sz91ZfRm/EqyejN7RVk8x/mKse+A9pWPzJ7SsXmj2lYvNOVi825WLyHKw0yeooWgQzSsHAFMnDRNLJIG5XsVZ65fso7NJfgkLHcQ0K0HDut/ZCtLdkWhxVtj0/ZoRO4xTPuvZ/HrM7ScUBfSWhBBQHI/qEau9/hcOZtBYr7bf8ApH8UbVY5IQ66XDAp1ntTobSMQaVTaq5oS/xdTVsFYHmbSx8PXmYqjivvNYf1f49O3T0ssP8AZFd02CxS+W5x+WrZ/B586jwsV9tv/SPgwBUkAKzjOePthQPwZNG7qcPDNtFqlDhvwPBS2KcxzAlm47imz6ALBnfBWS72tnViNW3qx/B4r7zWH9X+PA6ctgPFCzUhkf3prrzHcK7irOcpmKMjB7VH5bfio/ON+Kj8434qPzjfio/Lb8VH5bfio/Lb8VH5bfimeW34pnlt+KZ5bO0o/OM7Sj84ztKM/nZ8VH5bO0o/Lb8Uzy29pM8tvxTPLb8U3yx8U3ym/FMr47fio/ON7Sj843tJlQQ9vxUL31a9rBwquV0nNM14IZH4Gy2KrQTLJwYrbNURvELeDP8AlTS4yyl3S41UbM5Qohk8+4Iw5Tvao34SvY/9irPaxWJ+PA+DHdcntFMdVudDQhPlsTxC0udmAg15DiAQaEJhbQPCjDcZGD3phykZ2lHX+sb8VH51nxUZxMrO0FF51naCi86ztBR+dZ2govOs7QUfnWdoKLzzO0ofPR9oKHz0faUPn4+0FD56PtBQ+eZ2lD56PtKLzzO0ofOs7Si8434qFte+R+8qHzkah84z4qDzjO0oPOM+KiJJ5RvxQf8ASKCVm2IaueobUHGF4ddwcOH8dppq2darpJ3UPkoj40TD1tUAyhj7IUBzhj7AUHmIuwFB5iLsBQeYi7AUHmI+wFB5mPsBQeZj7IVn8yzshQeZj7IVn8yzsqzeYZ2VZvMR9kKzeZj7IVm8xH2QrN5iPsqzeZj7Ks/mGdlWbzDOyrNugj7IVnp/UR9lWfzLOyoPMs7Ks/mI+yFZ/MR9kKz3v6mPsqBrzdiaFS32lg81z2RRl8jg1ozJUklWQEsi6M3I9QRqRHipHmsjiVTWQcFa7I8Oglc1NfSLSYu8JQo54hJC9r2HJzTUeBbZ5ZZHZBxwUc+kTHeFJG/uE041qmHS05ujGQqJ2cbfgrOX/wBSz4KzNH/bxdgKz+Yi7AUHmIuwFZ/MR9gKzeYj7AVny7ni7AVl32eLsBWXfZ4uwFZPR4uwrLvs8XZCshys8XZVj9Gi7Cse+zRdhWP0aLsBWT0aPsBWP0aLsBWT0WLsBWP0WLsBWT0WLsBWOn/axdgKyD/0sXYCsno0XYCsno0PYCsno0XYCsZ/9NF2ULLeMHe/ZwTnx2+/xZ/HS3T1q6bvyVdIv93y/DYahzNtTaMtRms77r6UU48ezRFSeixqX0aJS+jRKX0aJT+YhU1qwfGwN4BCTxmJ01aPw4KUZOHwU3EKbiFLxCl4qXipeIUhzcFbNFurDNVhzYcWq074IFPus8Km3wQp/o0al8xEn+YiRGdmjT90DFPb7268pI7QJA7EIllHrlbW9/E11d81U8CRrxWHganVsrC39bP46+0sFugxfG2j28QiLa/8ZiVREoooopyciiiiiiiiiiinIoooooo6yFVwOrb/AAJ5p5mwqWW3e23+PNtVqfPYpWRX843K26NsUlqluPijzuEkqAYEu7JUJyJ+CY991t4nqQh/rGvaoHmgLuytKPiDxEMRXxwnaPtLoLUDHK3MKAfmcrP5blZ/KcoTkXKL+ZRcXKEYkuVm3ucrPxcrPxcrNvc74Ky+UfgrN5R7Ksvln4KynJ7uyoDlf7Khy2+ymsbVzXgdLCrOMy7slNmxYDishRVdRVKqqqqJfQNKumhCDty6iFXNt1YnNBU3Zoot1BU5gPOKusxTI8wUwYUJUVa3XqOR11kcjj1JkRo+OQdbVZhnfVl/mVl/mVm/mVl4PVl4PVk/nVj/AJ1ZXeUoHZByi4OUYxLXKHgVCThVaTtVnZPFZdh4q0l4C0t6L/natKxsc82Mmg3PaShZSRPDNGR5UZCsx8vsqA4AP7Kibmx4/wAKheaAPr1K32yztnhsryx2RJDVpafDuYN63tUuhrDMyctL5X3sP4+CKEVCZBpqHkGCKsIJuDpKkkgD+UIKkBqJnA8aBSlrSZnuUT320zxskc25dLxquWqC1jKRt09Y1hGM1Ca+IEqJ0oD6BqrMQMqol1UUUQFTWWXbpIQfZI3lzqkcUB+d/aT71OVky3uVXYklAFPtDxcGCdZ3bWeoBAGq6VxyVcuCI3/BcDVN4IZhBxJzJ1UKBHRzOjwAqog2oV/U5mTiFLfaOVfSqjZQtvVpneKY+peC/wBpxTnSuBcefQFUemOa2+rNvkau+uaPF3anWvSUEDc5HhqbFEyNgo1oAGsSfRS1j2f9QTa7x70y47F1a53ih5Tu0VGBayWAkU9Q4dOFsvLGKdjbvEEK2xR3GzQOWkOMSt7yAXwtCGhY5ay8pJLSvDV3d9HZroq+HvgVHHXgV3sLFZFBtwbyFkmDFMFcUEQiuUpVGKJsZYCGprty5Qg1ATAMXpoNS5CzREDFyfaHk0prJR1Eoooooooo6iNRyRxGo4onUUdRaTgUGDApyJOKFQeBTSNqEfFNLdiMJoJdvKFUNR14HV3puoV1cvpk2k+JA3mRaRsEtknJDJMy3Neat/ajU8eVti7BT/z20dhRaIjkDJHSOkzJ9SQ9ha4VBFCEdG6ZtFm3B1WdRy14O6l3kaqogRnfdR3oOBxBKY5b6I8EeCPDVTchQoBZlEHVUZasKhBoZh42AXJNF8cFdaXHIAlEtDyMCEA15pg1pK2A/cVfcAg9pc0YBA1Krktt7QBsGhqeCaXUogetbbWBtXUr+6AJFAg40ACuSBjW1N2udN9EMRShREzo2sBoASappQJQoU01wCDq4V1PG5HgiNyOZGoIcEAChUmiAqgAQQroWDliqwjVtKpX9HaDjviks22/1S7rsIt8ArNB4/S3Xev9SrZgitpeJ1BcNVcyq0rihwQJTXOqMAEBqxrTfrq6urBUQVY2HyXBB5h6SD8BRB1jeBmCR8V3lrQtiT2CEe5mqjxiu8v6VljqAmmPFxUb8HD35EIOhIf4zcAar60HfyhNFcE1VtLfYp+5Qcw0zaVS1uJyIHyWaGqp6ShS81woOOSv720OWBTW50yTXEioqEAHFpZgr25gRafydGaOVGnoqhnuIQAQ4oAkrArNbB1d7RWIR0npZt8d4h2nqgoPVIPaWuAIIoQUdFW8ywtPcs2LOjoWK5QS9SrZhqo5VLerwJ8D3h3Riq3D5KF2RvBywC70/wBkrvaoUeTKOo33+0dWDiOJR5bqATjjhinKswH8oVx4cMijyztR1HEImlBlT5pxJobuKONaHAUqiQQRRxcSnXafy0PFOJIa4AVzAzRDiTtVbgjfa7eQalG4BrGqtVVh11ARe9rRmSAEzRGjWwDGQ4yO4n8TX+NwaTsElltLasePgeKm0VpGSzTbsjuIRMJcR41V9X1UBKq7XRVVU5xoASU5njDmjnAtcOIXemnoVJnjiAdVIZPZWyFiqx0WGYWFVVzwN7iqYBUAAVZXdQVBQkKgzCHLf4Qg6Oiqyu8BCiHOoKUC6G/BZ4NRJyb8EZTdAFSpjWgU0Qq9qosCjq2Tq2lgFde1w3FX4WO4tB9VWWy1ujdnuKbDyTOAKo144OKzXezzKIMT3Y1U0Ut5j6FG0VErR7ggHEeDAyQrXfzQNVEDiOdXUPBjMoBw5Mqfc8q0GK6+QkK9U66nX3xbIOrlNEWR/GFvy9VGsYXOIDQKklR2i3cpG68wvwKPdTOsqkk44PPz1Uhd1LHXdaUXzBVYFiqErH+D0ar1dXeieZjq2VSdVYVsq/8ARyxHhGB6qfYdoAOYA/cK9DD0OCuWsAcSqWy0j+8Pz1d5ciKlEKiLiu+NWwsVT8KPAnwRc0rbKwVYnIgkaiiVRbCpOFVq2c1e+jcHQXD1UroiUez8wiWN9pHukO41VLfafb1bBQnY4kgUCAcUFQqkgVYwsVJIKRMLzwaKq1g42eXsFWrzEnZKtPmJOyVafMydkq0eZf2VaPMv7KtHmn9lWjzL+yrR5l/ZKtHmX9kq0eaf2VaPMv7Kn80/sqfzT+ypGeOwjrHgXvwY0nqCtByif2VaPMydlWjzL+yVafMydlWjzL+yVafMv7JVo8y/slWnzL+yVafMSdkq1eYk7JVoGcL+yVavMSdkqeN118bweBCm82/4KUYuY4dYTnNoM06N4vBYLYKMgc4K5mquV1tdWCuzNKvMHUqxqmhHs8mY+qnL2GWPiEeSHSQu+hU0lOOr5aqRFG7RqcU/gU6qxVYgsUI47Tx2V0rp5gQQQQQQQQVltsJitULJGdIQxfo6X/A9WuyOpPA9ipu18ArXazSCzvf0gYJxo+3zBg8hmJVi0dHdskDWcXZuPvQ1DUEEENbSgmWoNmbhNHw/MOCaGZKO0wESMBBBBBTLPNI92W4IRljmimKwWCdE5wCLkap2vbCrAxbBVyx2hnCT1Uq0hcjaZYuLrw96qWr7Tk6h8tVYyEb2roTTm0JoyCLBgjVBkc3uXSjq6V06jx8ECKFWGbGSyxE+zRaKP/pR2itFtysre0VYocY7LCD7A8B0rpXSuldKPFHijxXSuTlL82O8YD5oSQsIxB3oVTnREtBNCt2oPKCGotRG5Eld5agWlXRaW9LT8/VM6iC2UflKvBpCkjtz3hpu4YrDxAi85IONQNfQujXyFpBPinA9Sc1xCKKKcnJycnJycnJycnJyKKKcnJycnJyenJycnpycnJycnIoubRSWKXGpi3jguUYXxuBVWSgolrpI8xuW456yjwRQObU0bkInDAkJprRhTjbXimy5lPVAooo6r8ZBCMMzon7supMEBdmrwq6RrehcJWr+8aj5xqNnoHhQCzOJYb6a0UuozSgN3lHywnnEPanSwhsg74zA9I3FdC6F0LoXQuhdHM6OZ0c7o5o1dGroXQuhdC6F0Lo1BwoQp4C59nyObDkUDeF0tcHUc05hYEIOrI3AhF7Qb4VRjKjZxeLg4FQhry9lSBgo2g7OKEj6BPcAeUblVSNUqmsz6szQtLzDI27MBXocPVV74TLAO+sxA49CvWDpqqtGsrCNUhcEXOX1lvXrvvlHQNfRrCGsIIIeBCGoIc0eADhQhMtIvs2JBk4KezTGG0tuybuDukLvT+pHk2rBfV1SOQokld+CIjrVO4o8UeKLtNR+wfVQIOaQQnWQ8tEzvUjtvoKN0cya0viigjL5DkAtL3CG2YdsLTfof+dqnsekORtURjkbmCiscUTPPwuj+EAqC3QFj2dRGYVq0Y8smF+I4NlA+apE3q1Vsy0pyNRZvGGReFpl+VlHbC0rYRy9ospEbcyCCu88yunovYPqtHaYiyRgcDmCFYPRYlYfRmKxeYYrH5hqstlm5RkLQ5R+SovJVktpBmgY+mVQrAMrMxWEY9zMUdnYGRsDW8AP4QEFFaGFr2gg5ghWCuFmYrH5gfEqxRyB3IMwUYbSij8kKKWNzCwEEUIIVhDiW2dg6qhWPzP7lWMZQ/uVY/MBQWKYyRRBp/8A5zgEDqB9c//EADERAAICAAQFAwUBAQACAwEBAAABAhEDEBIxBBMgITJBUWAUIjAzYUBQI4FCcaBScP/aAAgBAgEBPwD/AAr/APL7LEjDdkuLitkPi5n1WILi5+xHi4+qIYkZ7P8A/wAOnjRgT4qctux3e+VChJjhJbrJdtiHEyj5dyMlNWvmbaRLHw47sfGQPrl7H138Prv4fXfw+u/h9d/D63+H138Prv4fXfw+u/h9d/D67+H138Prv4fXfw+u/h9d/D63+H1r9j61+xLiJNew5ORRGDlsQ4b3FhRXoJJFJksCL2J4UoZYeI4PsPjGt0fWv2PrX7H138Prv4fW/wAPrf4fW/w+t/h9b/D63+H1v8Prf4fW/wAPrf4fW/w+t/h9avY+tXsLjY+wuLgyOLCWz+NzxoQ3J8a34ksWUt2W2KNnLZy2ctnLZy2LDZyzls5bOUzlM5TOUzlM5bOUzlM5TOUxwogu5J2xGDgau72Eks7RSyqzGwNPeOTjqNBoNByzQzlnLOWaBwNDNLNJpbNDNDNDNDHFo7ik0Q4icNmYfGJ+RGSkrXxaUlFWzG4u+0SUm8kiECMEKKNKNKNKNKNKNKNKNKNKNKNCNCNCNCNKNKNKNKMeSukbZYGFqdvYororOk0Y2HpdrqSRRSKRSNKNKKRpRpRpRpRpRpRKCJRGqyshiyg+xg8Up9pfFMTEWGrZjYzmx9yhIikhOhTOYc05xzTmnNFjHOOcc45xzznnPOcc45xzjnUjd28oxtkIaI1+FGLDVGsmajUajWazWazmHMOYcw5hzjmnNHjHOOaOY2PNSowOJce0thNNWviOJiKCtmLiObtlFZ2ajWajWamazUamamamamamamamamamamWxN5R7jEcPG5/j3RjRUZvLZ9FvO3l36WUUymV1KRw+Pp7PYTtWvh8moq2Y2K5vK0OaE7I4dn08nsSwZoaZRpKKKKKKHGiihxo0s0lFCQuxFEtsuEW7/GjiPPJxKNJpNJpNJpNBpNJRoNBoNBpNA6RuaJMcGaXkiLOGxq+1/D+JxL7IbGx5QIIUGfciVPdEsNbxEOL36nvlEsstZIWUdiW2XCPdfjRxPnX+KU0jVKWxDA9ZCjGK7IkmSiNZogzAxNcfhuJPREnK3lQ0UQRgioonEn2JSMLFrsyWCprVEcWimUUxoplMplMoSEmJFFUh7FnCyqVfjWxiSubeVllllllll5XlZY5EpsSshSIiykSG80zAxdLE01a+Ft0Y2JqYyihooXYw50RxTmoniE5DZZgY7gyaWItSK/EhCJbZxlUrItSjf4sWeiLybF+VjzTIyFMcyUyTHmiLo4bF/wDi/hePOlQyuhjLFNo5jZKY30YWK4mqxyo1ms1Fll5WJiYmX2yae+XDY9fbLb8KMfE1ypFjZZZZZZZZZZZZYh9NmtmpjY+hCMN0QlqjfwrGlqllXSxrJI0lDiPOM2XqRXRZZYmJiYmWMgjEhoZh4mh90YGPGaorq2MbiL+2OXpf5LFk3+HSaSs0ROGl6fCZuosk+tjygrFAcCUSWcRCNLNJpNBoZpNLKEiisojipxpk4OLLMLipQ37ohj4c/UVPZlFpbslxEImLjyxBMQ7fY0mk0Gk0Gk0s0mk0lFEV+FEYigOI1miJgyqS+E4/iPqrJiRhxIwJQJoks0KJDCsWCclnJZy2cpnKfscpnKOWctnLZyn7HJaNLRKKkqZPCccthTktmc6fuOcnuyxGHgzl6EeHaOQzkP2OQ/Y5DOQz6dn05yD6d+x9Mz6ZksGhwKH0vOKIoSJIkPJESAtvhGPt0PpYjCIEyZMeUV3MONmDhJK3+eWEnsODjuSwYyHgP0HgzXocqfsLAxH6EODm9yHDwh+btLtIxYaWNDyS6WQIiJEh5IRAh4r4RxG3Q+hjEYZEmyZMYhGB/icIs5UTlRFCIkl/hx9kSH1PKJARIkPJESBDxXwjiPHofQxiMPcgTJkhiImB/wAfG8USH1PKBARIkSeSIkCHivhHEbLofSxbmERJbEyYxETA/wCPjbIkPqeUSBEmTHkiJAh4r4RxGyyWT6ZCMMiTJks4nD7/APHxtkSH1PKJDKZMkIQiBHxXwjiNlksn0siYZAmTJ5xOH3/4+NsiQ+p5QICJEiQhESBHxXwjiPHofTIj3MMiSJk84nD7r/j4+yGPqYyBARImPKJEiR8V8GxJ6FZPiZJWifFynujmtmHjOTp5NlizkQMMRImTPUihRMN6e59X/CHESl6Dxpexh4jl2f8AvtFockkT4qadE+InPcnjSTIYzlKmvwRICZMmPKJGIkfVNdkj6t+x9Y16EeMTdNfAsaTSJMm+xTFGS70YXads1pku52LSHJFmiT2QsHE9iGHiL0NU47oeJL2JTbexKTfoUQFRqVGow8RRfc5qMDFi5bikns/9r2H33EkNxRN3JstmJGTd0YaanbNRq6NLNEvYUZL0FJr0OdXoPGHijleUNyDQ2s2m32IeSsUovZ/9+c1EnxLTpInxEpeg8Ucx1I1dqHGAmki0XEUsP1TLQmLEoXEM+oY8dseJaLRUSoiUUJotHY0plJEcNP0FiywF2OHx3i9n/rbSJ4teJTJYkojk33ZrXsKKkKParHgxQo0aX7FP2LfsKUhORcz7xqRokzkv3OQclHJQsGhQo02aDl2RThsOF7mHcHaHxU16GDj8zs/+7N3Nk9x9SQ1mvwIS6luQimcQqRwXn/rxfIeWKu+SSRFl75RyeaZfRZfRfUiexw3n/wB1+TJeQ+h57jQkJFHqbsooooXobOjZ5J5xIMx/A4Pz/wBeJ5Z43kMTE7ErGixSG+2S7lfkTFnRM4bzX/dfkye+VFFDyqxKhlC7ZoWTFeSyvNIiRMZrQcEu/wDrn5MeWNuPbLDVqyK+0aGhZxGhb5roZZqLRaLLLEyZw3l/3ZYSVyZJRqy4+xa9jWvYc17GpP0NS9i17GpexqXsa17Gpexrj7GtexrW1HMS9B4q9jWrqjWtqFNexqXsKX8LS9DWhHc1NEZtkourZwmLUq/1zasstk5tsZGmKBci5FSZokaGaGKDNMhQZpkaZGlopig2aGaDQcs5ZyzQaTSNJEZq9iEYpWl/3cbwJeObNRrZqZbLf47FIU6NQnROVJNHMZrZgt60cT2Rw/mv9T2KEh7Me+WEu4thZRF0SKRSKRRQl+FGIQ3IeK/7uJ4mJ2gPJlFFFFddZVlXTJfYissDzRxRw/mv9T2Fk9mPfLB8hiEQ26l+azEZDch4r/u4mxOLlHsOLNLNDNDKZpNLNLNJpNLKKGmUymV0U8//AIIULOWYMdMrZjyUtjh4PWv9T2FlTaZJU8odmKSYmhNEGWi0aka0a0a0cxHMRzEcw5iOYjmRObE5sTmo5qOac05hKTkRhK9iPZf97SjQjEi0vtGsRqqIcPJ79iPDpO2zlo5cfYx1GC2OzNKNCNCOXE0I0I5aNCOWjlo5aOWjlL3OUvcWEvcwuGU1YuEin3FgYa9DlwXoJJf69CNL9x4d+o8H+D4ePqh8PAeEkLCicpe5yl7nKXucpe5yl7nKRy0ctHKicuJoiaYlIhGLaVHIgTwa8UaJ/wD8mHhX5IeDFo+ml7i4d33Zoj7FL4RxWy6WzUajUajUXnpkVJI+4wG9A8RixWRmKV/65SolOaVixpP1OY/ceI32ZB3sYieoSZTKF1Xk2NlmA7xF8R4vZdFjL7lmo1FiIws0pDkayzAlujE7SNQpCxakn/ru2bmInCdCkWQnpkYq7WhMssw92V1yGxHD/sXxHit10vYe/QlZGIkMnJJmtGtGBiLWjiXp7nNRzkcyzBlqw0/9M9skcZH7dQsSjmo5qMOXMwhzp0cxHNRhYicuqspjEYPaa+I8TuuljzSIoihGI3XYx1JPuXlFtNM4pasG80cFK4Nf6ZPPESxMNjz4GfdwOIi1Niyi3F2QnpjUiM4ydLoRInkjC3XxHiOl7DQkJCiJZYkpKXYhN13MaWurNAodh4Z54H/o0MoRweKoN2QnGauP+dtLcxMS+yITtdzFx77RMLF0dmYi+9igaDAejETOMj3TNBJdssJxl5MjGK7p9CJbE8kYPkviOP0soSEhIQ2kiX3Oy9L+10SurYllZw/3YbRVDQkR3OHxlhtpid/5sWWqQ+x3kNWzTRiKpCy2OIWvC1CyYlbHprsjDxq7TFnLYnkjB8l8Rx8nnRQkJZSkojd92Od9kQiYniJ9srODfaSMR6ZtF5Ql9yH2bMDHcOz2E7/yTnfZEmkhLV3YkJJOyU+/Yxe9MssZh/fw4n2LGREVZDFeG6ewmmrWUieUTB8viOOh9NCWUp1sN0NuRGIkT8SCJw+0ivRnDqps4lVisSKKK3EcPjaXpe3+PEn6IbpG/diNjyZIxFcSiiaODd4bRJfc0JDVlZ7kJvCf8IyUu6JE8omAvuXxHH8ehdEpeiJSo3FEisnsYcbRptDh3MJacVGPBPENBoocBDRZw2LqWl7r/DOWlZN6hDaRbmJUS3JeDFA0GJDscG6k0YqrFZTEmNdMJvDf8NSlG0TyiYHmviPEePQs5S9ESlWSQkLPhlaeUkn3KqcWYkbmjR2NByyqyaIScWmiMlJWv8Enqdk36CG6O8nYlWUtyriYauKNBiwuBw3bFMZVjJjgaENLboWUZvDf8Ju8omB5L4jj+PQspSJSySEhLo4bZivKfp/9k/KIkUKJiqpvNdmcPjaU4s56OfE5yOejnI5yOcjmo5qOYjWjWs8R0qG6zb1MWSHuLY4f9ZQ49jD7YiOK7OLNI4WTgS6Ebm3bKJgeS+I8Rss0IxJ+iJzrJISEhdHC7yWeJ4k//i/7khHEKsR5vuhSp2RVlZ2X1w2yxHciXtlN32Quhrvlwvi1/Ssn2xDiVcLFlMnv0LJqxexEwPNfEeIPXOU6JyzSEhdPCusRkorLF8GT8U//AKE8kcV59E/cwWtNMeaysTtZ1lDYuu43k3SF79NZcM+8lnidptmL3wTCdwTPQmzE6FkmNWRMLyXxHiM3KiUs0hCyWaizC+3FLyxO8GS/VYnlZxXkuhq0YT7dEV2yRDphsYjqA3lJ26/BgOsR5WYvmyPfB/8ARgP/AMaJGIybWSzTz2dmErkviPEZXRKQxIrJZoUTshyE6xBMkyT+xoXfB/8AQpFmo4p7F9EO0/8A7zRVIeUemGxjv0JUPayPuLqZB1ilmoxvNmB3wjA8aJSMXuONdaywJaZpP4jxGU5Dd9CEIUTshsvK6mmayxtUYffB/wDQpCkajGeqjYTEIfZp5JEUSHvlDpw9jEpsm1fYc7jWS6pDdTTNZrMV/ecK7wzDdOSHMnMc7Y3e5XUssDE5kP78Qx5XlJ9CKEhJIciy8rG+9iYpWtxswP0moUhSJYlUxYuHieSolgdri7GmhE/Ej3QiBIe+WH04exIxOzFv0I7IcjVZZiejNRqJvucC/tY3WJIlIclpQ2hPrQjBxOXP+fD5y0omSedZpWWlsXlZYzW49oslqfdsSEqGzhe+ES7Nil2NY2aSM5w2I8TGfaaOUpd4MmmkYTuCERJEt8sPph2g2SMXayO+SLSHM1FiY9ib7FjkSOBe6MZ6cRkmX9ozD+1dyaV3E1dKFlw2JrhXt8OxJWybHklmkN5WbnZGopsjhmJ2RGQu6JnB/rMTtNliZYpFjSYpSg7TJcU5Q0swJ3EjIiyWxLcRh79K/WSMTtEg77mtDmaiyyxM9CWwy8uA8mcX+zJPtkoUhxHE0tCaeSzRgT0T+GzlSJMfSkNjKZ2Q2yhRIxNNIxV2ERl2JSOC8GY3m+qxt5cL6oiiJiYmkeIpEMRSZDyysssv7aJGIrjRqZqL6EItEmN5I4L9hxfaY2IW5h94DRRQ4DTXQssCeuCfwzFlbok+lIby2HJvbLSKIkOSjua4vYlbj3Kos3OB8WY/7HmjlsepbnZlPLhmoydnNgLHiYs4z2IyiiLUZXYsWCZ9TA+ogc+Bz4HPhRzYk2PfOislbFHKS7Et8onCftOO3WSywp0u5ri/UoooocPYWaOElUnH3+Ft0rJMeVFCQ3k5ZJCQoiRRxKujCTUkT2aztHBOkziP2PKhLuRHFMlhIcHHLCjrnR9JP0PpMU+jxT6PEPo8Q+jxBcJNn0kz6WYuFmfTSFgDj2onGpCRWVNiwxQFEm6JbEt8oI4ftiI47ullAoUvscBWRdrOihxs/jyRGWmSkJ2r+FYr7USfS2XQ1ZQkJCQkVlxBgXrMTZ9HCepxP7WLJblSyY0Sgd4O0R4qcGLj5H10j66R9bI+tmxcZJH1cz6uYuKkPGnLcjiM1WifeV5KNighISKKMTzH4sZRhmF5o4zwWUMoR75LbqaNuxVZcPLVhr4Viu2PpbKKNIkJdPEowPIxV2fRwe7OL/ayOSK7HMNaLRaGrRKBGPcUDQaDQaUKCOWhYaFBUOKRHxJL7hL0Es08u5NPXZLxHlhmF5o41fZlERHJbdNZNG3Z5cJLu4/CW6RN5vLcSEhISKLys1Fko2yEEjE7xb6OD3Zxn7SOS3FJFrNmtjEtPQkVkhEWSI+LJ9mQ7sQkJFI7FotE6HlhogqkjjO+GIQjDSbHCKVmvv2I43uRkn1NZYEtOIvhOI6iPfJusqEhREsrL9zVmkJFZY1KNZXlwvl/6ON/YQy9STV5WzUzUxvv0qJRQihCGxeLJeQkLorJ4kfQnKV2x5YStkV3Rxf6xISEjEtLsOUnu81JohitbiaeVFFDLruJ2vhGK898lEorJl5JCRaQ8WKOd7IeLNjnLezFm5eh2SyRwvl/6ON8yCEihxTkxwa7ov3KzW2SKMLAcu7I4UUPDi/QxMCu8clkyPiS3F0SnGG48dvxO77tlpEneeCu4l3OK/WxCEOOocE1l2KyjJohO82s8F3hr4RjZPuxISKybG8uyNcVsOchtvdlLKjSaCcKQy+4jhvI4zzMPNqpPJwTHBxKvKO2SRhQ1SroaMWOmXRHYn6CEOSirZPGb7RFHv3EIa6MDyPU4j9bEhIR/DSOCHhoeEypITEQl0NGBJaK+DubsbtjfoJCRWTY5ITb2R3KReVFFFZz8STIKxI4fyRxvkYOct30SjXdDXqQ2EI4fyfQzH3XRHYkIlPSh3J9xREhIo02ONHq8uHXd5Y/62IQhdTimcv2FqRDEWzyoaHKkcNi61T+DMuhIURIboliDbZSLyooooS6Z7EtyAjA8kcb5IwSsmu/S1RDJGHLTKxZsxJapdEfEk+wpofdiQiuie56vLh13eWN+tiEL8TimLVEjKySJnDTrEXwaW4yIqJYqWw25b50V00yiiiiia7EkRQmYHkjjPIwc5eTLfQyGSEzCxnHsyOJCXqOcV6mJjauy6Y7GJsJIWaTFFmk0k4oaqTy4d96GY362IixbE9kWazWi111YnfZmIYbqSfwX1RieTIsstsrqor8M9iW4ssDdHFr7jCzxPJ5WyzUTl7GEriUVQixXnQkaco7GJsRT9SskumZLdjMOWmViMb9bygIxH9pqLLLFJimKRfR/GYiER7pfBEYnkyqzoor80/EluIRgbo4ryIbjyxPJ9LRh7dNl5IQ8o7E8lkumZN/dkjC7wRjeDFEURGJ452WLJCYnkspvsV9tmF3gvgiMXyYnZYmIZf5p7E98kzAfdHFeRAssn5PqhsdjsJotFosQqFmvEnkuuZPzzwH9hi+DIZ43h0oTFmpURmmaiXclHTBIwf1r4IjF8mJCExMeVi/JLYm+5G26LSOGdyRxPkJjeU/J9Dyhtks07zQs14khdC6JGJ5Dy4eXajF8GR3ysxfDpQhZtmE0ty8P3LitmTlZg/rXwRGJ5MooXSvwLonsNpS7kGr7Icl7HCeSOJaU+5rRqRq72T8uh5Q26kIXQvEl0IXRIxPIw5Rj5KxYuHXiLEhXaJP9X/ojQqIzw1HujF8OhZIWaViijSSiIwvBfBPUl5Moar/ABS2JruRWXC+Zxu5EvKW/VDboWaEJ5rYnuIXXMxN88N9jE/WxZ4nj0oQnkkJCLHlDxXwV7ssascWiy+i/wAc/ExPIWXC+Zxm5HN750zSSdMw9i0XEdemSKK7CFmtjE9xNPJdUzE3zjsYv6hZ4nj0rJIURZtkuyErfwZjEzclh+x3Qn02yy11z8SStlUI4X9hxXmVkijsjUjWxt+4o6mQVJLoQhCeSzWxir7SHZmqPoajUakds5mJvmrMX9TzRiePQk2KDFChLobIr1Y3Zgxua+DOI0LJEoJkoNHcTLLyXRbLLRZPxHkjhP2HFeeVCQ0UdsmyG5HNOjUvY1CkKWSzXiTXYmmiMhPJIoVmpk5MlvkhIxv1PJZTVo0MUBQQl02aiEb7sbIqzh4d9XwZxTMVJMYmIsascEOA4stoUiyxMvpmuw8kcL5nFeZRRQ8mNjMNd7I9SIYXa2VH3GqzXiMcbRiQcH2FIUhSExPKe4yskzH/AFP8VlmuhNsjCty8sDDTjb+EYyGbPO8mhoaHA0sto1CkKQnnPbo4XzOJ/Z0N982hRsSoXQijCScjEl3LIvNbDylFSRiYTgKQpMUxSNdDlfQjiP1MXXZZrLbI4bYopdGCqgvhGIriMYiyxPKhxzcRwQ4tFtCmLEIysntksuF82cR+x9DfcvKKbFGsl0ISIdmYivJZvwLvOrMTA9YlNCtCkXayWa3OI/W+lsscjWd5EcJ+pGKRfQiKpJfCZKnk+tlZ0OA4Gg8d2a09ujhPJmP+x9En3KZHD9xJLbNdCEJkZD0liyfiIXRKCkPDaNA4tCdCzjujifDoqx4UqPubFhP1FCKF/BLqgrml8KxlTHk87LL6NPRpTHhInh6OjhPJmN+x9CgrsorJibsroXReSyfiSdEG3lRpNLKY8Ox4bRNd+iHkjivDptGrJRYlRZfQvc4ZXO/hWKrQ86ysssTL6KRpJT0uiL1GI+3Rwm7MXzfRrSNZcmVI0kIrorNdCyfiS3I9hHc1CkzUajUjEpu+jD80cV4rotI1CiJJFll9C7jZwsai38LmqY82hrKxPJMXRj7mDO2T26OF9TF830dpOjSVlIh+KxMTJeJ6jT3RGQnksqJSqVEvHowvNHFPsuhqymLbr37I27CMOOmKXwvGj6j6Wh2iyxMsWeLuYe49ujhtmT8n0JvNujzZgx1SOUvc5P8ATkf05H9OQzkM5ByDk/05JyXVmiiWxfcslC+6FNrsxSQmLKfeZLx6MLzRxPp0Ls8rL6bNssCOqa+GNWqJRrqaHEuhMTE88TyMIe3Rwq+1k/J9CG0h4i9DS5d3sNqqRw0dyiijSaSiiisn+tEieViY6luOLjsRxa3IyLJLvZJ/b0cOrxEcV5LqfTd9keOSOFhScvhuMqQ+pjjY00KQpCZZSe41S7Etujhv1se/Rr9hRlLcVR2G2xHD+PVfSv1kye2VdDimaWhMfdDfRwvmcV59F9ultIVyNtskRVuiEdMUvhuN4/garKUCmjURmRlY32J9HD/q6Ukth5KNkoNK0YUdMF+OHiSJK3RpptGkorp0ko0r6OF8mcR59DfQ5ewot92WtkLPhoapX7fDsXw/C1m4DTQpilZLowf09SRHCbP/ABw/pPFcotC2/AhSyw9kSRRLcTFIpM0ldE9ujhPUxnc30PfLULVLYSUBu8lkjBhohXw7G8B5PpsQ1ZVFDQ4Jn3J9h9EO2B0JEcL1ZrhDYeJKWSV/gorOGxNLUTejsPv36NR2ZpKynt0cIuzMTvN9ChKROEk6FhpeQ5dqXVw2Hqlfovh+N4dV9Njj7FFGhDVdD7YHRGWnuOcpPJLLAhckaUUjSjSjSjSjSikUikVliq42TRJdupMTTJE+jhe0LJd2+iktzUPoWSOFVR+H43gPNl9dlpmkon0YnbB6UhZI4aO7/LNXFj2yeS6WS6MHthdCH36aFlH3OFvT8PxfB9LWd9d+5Pfo4j9WdFduiCMOOmKX5d+2c0V0oY+hfbgi2z9H+FM4bx+H4vg+lokhMv8ABLNbpHFeK/BhxuSX58RVIe4+tj3z9TG7YWTyexfXfoROG8Ph+J4MfQiUSqLFlfS84K5o4p7ZLJ5xXqzho23L881cR7ZSXQhvJ5xVyRxTqNZ0T7LoSzbEQOG8Ph+J4MfQhjQ1Qsl0y3zwFeIjivJZLJ7ZRjbyw4aI1/gnHSxqsmqzW565PPBV4iOKeyzTJ5rPcYiBw3h8PxnUaGsqKEPJoaaL6WPJHDL/AMhxPnksnsJNukOl9qMDDt2/T/DJalRJZ7ZR3H08MrxDiX9+ViY+lK3RKlkkRRw0u1fDpSUVZOdjeSyb6HEaoTFm8llwquTMd3iPoab7IpQiQg5MjFRVL/FOGoksmrKoW49xbN52cIrbZju5sbGyPRRv2Q3mhEJUyE1JfDcWdsbzssdGo1GosY1QmXlPYWfCLdmI7m3kiKFP7SMXJ2YWHoX9/wAkoaiSccmryluPthvo4ZVhtmI++cc6Kvsjlr3OUvceDL0Gmt8kxMwp0xO/heLKlWaysliUarEyyyyxoaE8p7Cz4bthtj3yRCrFHUzDwtPd7/5pw1kouLLy0uzH7QSzZH7MBEnbFkskhK+yEqzTaLUlUieFp7rYQmJmBO1XwvGlchvtmycqRvkumxoaIzp0zFYskR+3AyWUdmzh4rTf+iUVIxML1iPsyzHfdLNK3RxDqFFZq20jltd2RWr/AOhRSKRpNKNKNKFFVTMSOiVCYmYMqfwrZEnbzeU1mhMTzWWoUNXcmms1uY324VZIRw+GnF2JJdl/oeWLhqXdbmHgpq2Yr1SbzwI6sRHEyt10VRqFiUc05pzTmnNOac0xJWxEWQIO4r4TiOoMfU4ZIQixZ8pDjSVEu7KEYMdU0cS9kUJCiYSqC/0vPFlpg3lWXCxpORiS1SzSJb5V0VmlW42IiRZg+HwnG8BiyfRVk1RdGs1mtnMYsRixDV2GUUcNDvqMbvMURRIQt/7EcTK/tKKNNjSw8NLoW9DhE5cTlxOXE5cTlR9xYUPc5cEOlsSyREiYHj8JxvDJZPpxOhZWajmdqFJGm1Yn3MJrRaHAUBQIQp3/AK2iUtKJJydmko4fDuasx33aKNJQlXXY2MlkhETh/H4Tj+AxZPpxMomli7IfRBJifahbmB+tFL/gUjSvYnWpmE/vRxW+Sof4mPJCImB4/CJy0RsniueUcn04mSQ6qurCFsLcwP1r/h4j+9mH5I4nfJPuP8THmiJDH0djDmpxv4PjfrYnksn04mSsv3HVZpFGELYW5gfrX/DxPNmH5HE75ev42PJESI7bo4W9LXwfH/WxZLJ9OJmk6sttZpZYb3I7C3MD9a/4c197MPyOJ8svX8bHmiJF6ZWcNLVb+D4362JCEMY+iZQ0J0uhZYaFsLcwP1r/AIc4Nu0Q8jiN81+BFDHmiIzhF9j+CWjUieLpfYliOaqiUaWSH1TKKKK6KMNOmR2FucP+tF/7rLHJJHOLO0jivLL1F+JjzQjDwdfcw4qEaRfwKTKKJQTFhIxUoqsl1zGyyzUWWWWKVEFaPUwZxWGrLT2F/sexedITKOJ8svUUWaGaTSzSymaWaWUyhokPJCOGaprNSE/gGIWJiZiYuhbE+Icu1EZaiijvlR3PuHGTOSzlM5TOUzlf05f9OV/TQvc0neiKRDEnFUkYWJKTpoo2/wBctuihqiWK4LsTlKbtjRQov3KfuVL3NMvcWv3Pv9z/AMnuf+T3E8Q+8amaJM5RyhYZoMCKSsxsTQux9RMwsaUpUy6FP/vzMSTjHsc2a9Tn4i9SXESluc3+HM/hzv4c1+xzWc1nNZzJGuQ5MtllmrqvsJikKciOLL3HjTXqcJiSlKm/9T2zn2RzZ+48SfuKcnuxstFo1fwtexa9i17Fr2LXsWvYtexaLRaLRaOx2yU2tmSm3uyyMmnaZzp+48efuYGNNzSb/wC9JqzH2GPb8FZ1+aIzgvP/AFNWjZ0xGN4PJkej1/ys4f8AYv8AvYr+4xncUMfRWdFFFFFFFFCiUKJRRQ1miRwXl+S0X+aStluLMWV4byexEea3/wAVFFEjh/2L/vT8mTVqhpjRpZpZpNJpNJRRRRRRRRRRWVPKmUzSyimSVvscHF6r/C5JDxGOQ8SK3Zz4i4iPuRx4v1E7/G9yRONp0U0NCRpHBigzSzSaWaTSaTSzSzSzSzSzSaSiihJlMoplMcWzh4NTQmn/AN2fkzE3GNsU5GpmuRrZqkWy2Wy2Wy2W/ct+5b9y37lv3LZbLZ3LZbLeSbJN2cE221+CTGTxUthzlLfKs44k47Mw+JT7SE7/AAy7Gu5CMa9Tyd2Jstmpls1M1MUmWWy2Wy2WzUy2Wy2Wy2Wy2Wy2Wy37ltPc4R3f/dxPIxPIZL/SiRws9MhY8TnROdE50TnxOej6hDxyc9RRRRRSKNJpMOej1Fjo56OfE50TnROajnI5xOd5KZiO5Zev+pnCev8A3cWF90T3GS/0okYcqFM1s1mo1M1Go1M1M1M1Mtmpmo1M1M1s1ms1ms1imKZqLLybHvl6/wCpnC+v/enganaJcNJIeEx4bOWcpiwmz6PEHw00ciRyJHIkciY4NGkWGzQzQzQzls0M0M5bOWzls0GhiVDiOokX3EhRNJpNBoOWcs5ZoNBoHA0Gg0Gk0lFFZJiYnnQ4lGkSsarKiisqZpZoY4yzUZM0M5chcPiNH02IfTYg8KSNLKKNLFgTatI+mmzBw3hp38Ax0icqZqNbZwtO8sRdr6G6RKaFKLIyTZpNBoNBpFE0jVDeWJiNOkQm5PvlLD1HLcBTNZzEcxHNOajmHMOajmo5qOac1DxEcxHMRzDWazUaiyyxMUjtlpJ/aaxVvQ2nuiKs0mg0mkoURIY5Ky4kZ3aQmR7iVZz8HlJ0ajhO9/A5wUyfCyb3PpJC4SRg4XLWUlaHnLYm+5ZhsT7Z2WWahyysnh6u9kcLSURmoonLUysqKKKKKKKZTKKNLKNJpNJpNJpNJpNJpKEJ0ayX3GgSKF2LLLLyWTJvuWYL7iMGNy6GrVD4b2Y+Ev1FwX9MLBWF8KxFTzexNFENy+yFIvroorqorsVW4kV6iRXaxIooo0loSKyRtlXehISKKNJRpKKKKK6KyQhk13eWHuIwY1H4lONoeSJrKO432RYmWWWWWWWaiyyyyyzUX2G7Qn2L7EZF/aKQpblmob+4TTF2tMvuhstDexqNX3IUjUahSLLLLLLLLLLyvJZMnuUQ8jBhqfxTFjXfKJLJbkmJll53+KihLJ5JFdhLOh7lD36vXoWVllllll9hMsssTExDJ7iI+Rhw0r4o0mqJxcXRhkt3kjcossvKivxroXXeT6F12WajUajUXZpZTRZYmJZMnuI2ZF2viuKrEqJ7vJCRRRQlQ2xTlEWLr3/02X+aKsn22OZIWLPYtsSIoSGMmhZYbuC+Kzds9TE3eS3FE0mkjAlGkS3yisqKKKKKKKKKKKKKKKKyooorKiiiiis6KKyoooguxPKO4oiiRiJDQyYssB3hr4pPxeT3MTyeS3INM0CgKFE12J75QRRRRRRRRWdZUUV0UUUUUUUUUUUUUUUJFFFFCiJUYmUdyKFESyYyYsuGd4a+KT8WIZi+Ty9TDiJkbynsTXfLBjZy2ctnLZyzlnLZyzlnLZy2ctnLZy2ctnLZoNBpNJoNJoNByzlM5TOUzlM5bFhs5bOWzls5Zy2aGOBpKGjEXYYiFsQnmya7Z8J4fFJK1QmMxfLL1FGhIiWSZPcRwXr+dpPcngexKEo79KhJ7EMB+pGCjt+WcbQsq7mL4jyw9hCzY8+D8X8VmtMmMxd8vUTLLLGxws0HCqm/8ThF7o5MPY5MPYUIrZf4Zw9UWMnsOHc5ZhpxExM1FjHtnwT3XxXFjav2yxU7yoTLLNRZeWHPTK/+POHqhsslGxCYmWakaiyQouyjhbjP4rRKOiVE32LLLLJMhVCH2RrE2zh56o0/T/j4mFfdDTWU42ahSZJtow2nlsamWy2RxHEwMdYnZ7/FcWGpdtyew+iRhrsUT2yRwe7/AORPDUyUXB0x7ZyfYwhIlsIsss4X9q+LcTh19yGmUUUQw3N0iPCzRyJmNBx7MoSOD3f/ACZRUlTMXCcD1yashwmIj6aZiYM1G6F0cL+1fFmjlQ9jlQ9jlQ9jlQ9hQjHZZuKlujlQ9jlQ9hJLb/l8qHscnD9hYUFsujk4fscjD9jkYfscjD9iGHCDuK//AAif/8QANxEAAgECBAQDBgYDAQADAQAAAAECAxEEEBIxEyAhUQUyQRQiMDNgYRVAUFJxgSNCYjRDoLFw/9oACAEDAQE/AP8A6x1OjUq+RXKfhdR+d2F4XSW7Z+G0B+F0ns2T8LqLyO5Uo1KXnVv/AOHUcHVrfZFHw6lDrLqxJJWWWpGuC3ZGpB7MuNKSsyt4bTl1h0ZUpypy0y+s1Fy2IYOvPaJHwuq92Lwl/uPwn/o/Cf8Ao/Cf+j8J/wCj8J/6Pwn/AKPwn/o/CV+4/Cf+j8I/6PwlfuPwhfuPwj/o/CF+4/CP+j8IX7j8JX7j8JX7j8JX7j8JX7ilgacJd2RioouTrQgrtlTH+kCWJqS9Ryct2KTj1RTxs4+YpYmFTbKvhoV1aR+FR/cfhK/cfhK/cfhK/cfhK/cfhS/cfhK/cfhK/cfhK/cfhK/cfhX/AEfhX/R+E/8AR+E/9H4T/wBH4V/0fhX/AEfhT/cfhT/cPwqfpIl4bWRPDVYbx+m6OFq1tkUfC4Lz9SFCENkWRdI1o4iOIjio4qOIjiI4qOKjixOLE4sTixOKjio4sTio4iOIhSuTfQhHShmKxah7sdyU3Lq2XL5XLkZuLujC4rX0luXFKxrOIazWjWjiI4iOIjWhTRqRqRc1JGuJqRqRqQmjoOCZVwdKpuit4ZJdabJ05U3aSt9LQpyqPTFGG8OjHrPqyMElkyUyVVjqSOJI1yOIzXI4jNcjXI4jOIziSOLI4shVWcaRxJHEkjiyOLIw6ei7LXGYzE8KNluOV+rL86bRhK/FjZ75M65XZdjkzUzUzUxzYpsU5I4kjiSNcjiSFUkcSRCoyExO+Vith4VVaSMTgJUusOq+lKGHlXlZGHw0KKskJZyY43HSOCcA9nOAcA9nPZz2Y9msezHsx7Keynsh7Keynsp7OezXZsrC6InJQi2ytVdWbk/hYeq6U0xO6y0Gg0DpnCOEcI4JwTgnBOCLDnAR7OLDo4AqNhUxLkcUzGYBT96G404uz+kcPQdadkUaMacdMRCzsaTSaTSaUaDSjSjSjSjSaUaUWLIUUaUWQxjN3lj56aTXw1uYWbnSXIkWLFixZc6+E0Y3BqotUdxpp2f0fTg6klFGGoRpRshCQkMdRRPaqa3I14S2ZdZXyuX5E75ppl0XzbGNkdxHib2XxMA708lIuajUajUajWazUajUazWazWKZrE28tcV6imi+ckY/C/8AyR+j8BQ0rW92JCQspFSVidSJ7rKbnHyshVb6SWSkuaO2TLFizyY8vUhvl4nHqn8TAK1PLrfl6/A6lmWIwbNMYLqVMV6QHOc37zINEJifJON1YxdDg1Pt9G4ai6s/sQjZC5JmITJtibRSmQ6iRUpt9UU8TZ6ZiafIudsbGxEejy8RpuVO/b4mGhopJCQkWLFixYsWLFixYsaRRIwQ3Yndk16oYrsp9CAs2jGYfiwGmnZ/RcU5OyMLRVOFiOayaKtNyJ4c9nZTokI2FlXw6qL7lGcqb0T+GyQxb5J3KtPXBonHRJxfwsPTdSokJWQk/jovk0SgOmKmQpojEXJJHiGHt/kj9F4Gld62RFkuSxoTOGiMbCXJOmpdRISLFixblYxrOD9Msfhb+/H4WCoOnC8t2JC5rFi2VixbnsOCNCFESFyMqwUo2ZVhw5uP0Suph4aIJCEIXIsmXL8rRa3wWMaLCGynNSRXoupHo7MxWFnRd+dddjCYK3vzy+3PfO+bfw7l+RkjxCnZqf0Th46qiIiEIWSyWUnYcyNRMixZvJs1Go1Go1Go1Fy5fORGTg7kJqSuhpNdSvgIVFePRlXC1aW6GrbnQUXLZFLBVahh8HCir7ssNisnc1Go1Go1Go1Go1GpFy42XFmuWUhzFIWbJGMhqpv6JwSvNsiIQuVFypInV6kKvUpsWbGyc7DrI4yOMcVHGj3ONHucZdzirucVdzirucWPc48F/sj2im/VGqMtmQk4O6IVFLJq+5KhTluj2Oj+0jQpx2QllUrwh69SWKi92e0x7ntEe57Qu57Qj2ldz2mPc9qh3Pa4dxYqPcWJj3FiY9yNZMUzUJizWSykTkORCRFiyYysrxY1Z/RGB3YhCEIWSEMrFUpPqUtiObZOVjF4t6tEByb3fxU2ndFLHSj0n1KdeFTysjXlEjXi9xVYdziQ7jr047sq+J0YbdSt4jWq9F0XxqdaUGUqmpXE7kRcqykTGUiGbGVNifmf0RgPUiIQuRCGVtiorspLqUtiObK7smxu7v8AkE7EcTVj6ix1Q9uqdkPF1X6kqk5bv8jg37pAiLlWUioMpEBZMZUJ+Z/RGB9RCEIWaEMrbFTcpFLYjnIxHlf6Pg9mRIi5VlIqD3KZAWTGVCfmf0RgPUQhCFmhDKxUKW5SI5yMT5X+j4PZkCIuVZSKpMpEBZMZU2J+Z/RGA9RZIQhZoZV2Km5SKYs2YnyP9HwflIiELkWTKg2USAsmMqbE/M/ojAeoskLJZoZW2Km5SdmUiObMV5H+j4PykSIhcqJFUaKRAWTJFQn5n9EYD1FkhZLOIyqVF1KaRSI5yMV5H+j4PykRCYnkuSRUGikQFkyRPYn5n9DYehx5WuQ8Npt2bZR8PhRvpYqCRUoqKuhERLkQyqS3Ke5SI5NkplSPETiLwrvIqeHwp9NTFgqfrJmIw0aUdUX+fSb2NEuxTpSnJIh4XRcbtspYKlSVkRw8GidFRjdCYmLlkVENFJEBZSJSJO4vC1Lq5H4VH9x+EL9xU8KlCLkpfQWCpxm22inFLYpq0i6HKL6XKnWDSFFoihFs+JBbseIpfuKlei+mo005PpIhSjfoyCSW5FouVLjTIRdyzK9KU9hUJmNoTVNJIlCUfMrfnYK80hWj0Q2RhJvYh0ii6ISjtcq9YNIUGKIll0NcTiQ7jnB+o4xl6jw9/UjQt6kYWLZTRUTEmR2yTSXUr9aUrdiVOcPMrfr9Gg6quU/D4NdWUMFClez3FRtsxRYro09bilIaky0hRkNVvSxZjiyVHUPCQZ7HEWGSI0bGhotI94eo0stIvI1SRdvqTqyRPDQxb957GNwccOk4v83GLk7IoYN31T6GmJGhBiglsaX3HOURztLVYjWbG77s6dzp3NKfqOCHCJogaaYtBxIo467CxH2FXOOOtcc4sTSOIcWxJqb6nEfcrwjWjpkxeGUntJmKwnA2d1+u0YqNJFPYjyouLmuLkY3khjLdCT90qTaMFK7Z4r8r+/zeFilTQsqWwhkoiWx0JZRfXNxZpLZ6blhZpZpZsgeI/L/XYL3EU9hc+wmXG81sXLiZcfVM3Q9hDWcyrHoYPpM8V+T+bw/SkhZUtiJYasN2E7s0DgyKaYl0H0L9R525LZ2uaRly5qTIbHiHy/12HlRT2FlcvyMXOhHTnbJOzJ9UYaLVQ8V+UvzdD5aI5UX0I5TfWxKXv2ISEx5IkRY9skPJlxI0mg0ssxxNJpGiB4i/c/XaOJu1BIi57CjPuKE/3HDn+4VOX7jRL9xol+40S/caJfuNEv3GiX7jRP8AcaJfuNEv3HDl+44cu4qUu5ol3FCXc0S7mmXc0vuaX3ND7mn7i/kcL7jppEHHVpieI4dzp6k9vzdBNU0mJMS+5TppLcS+5JNepqv6miBoge6uhria4nEihziXgOUUi8C8N7icX6icX6kpRia42uKaOLY4y7nE+5xb+pxPuavuNvuR1S9SdOaj0kVatSbtN/ruD+aQ8whCLGksixb4lhxLFiEb3uaEKCK6SpswN2ncxfyZfx+ap+dGscmQ86ICKjsj1N0PYmuo11yYyBdo1M1MUncUmN3yuXyvySfQw5U8rJ+Z/ruFdqhRd5kRCyv+Tj5nniflswGzMX8mX8fmqXnQ1lTfvoiIreUVmNj2Km+Vsmi9h8i+Hh0VPKyfmf67hvOUpxjLqyM0KaOIjWhSRqRqNRc1Fy5cv8JeZjkazEXlCyMJTlTT1GNmlRlft+apfMRJFlYhKKmupHbKfVDhJDi2rjTSKkbjTNLNDNDOEzhSOCzgyODIVFjoyFRkcGRwJHAkezyOAzgM4H3FSXcpRUFuVJxS6sl1b/XU2thVZI9pmYaqpytN2FKlF31Iq42nTXTqT8Q1RtGNj2uY8XU7swdSdW92XZdmqRrl3Ncu5rl3NcjiSOJI4sjizOLM4sjjyPaJdjjz7FfH8J2aJ+KO3uofiGIf+w8ZXf8AsyVSc/M7/mk2ndHHk/MKtTW8SOKgtkLF/wDRDFyezI4qfqxV5NXHXme0S7HHl2HiJdj2iXY9on2OPM48zjTONMdafc4sxTn3NU+5VqyhByue31Sji9a952/s48P3lfGaHaDuQx01L3tj8QpdmVMfBxelO48RVf8AsOrN+rHJvd/Q/hu8uSwomk0mk0mksWHJI4kTXFs1xMZFOqLDx7Hs0OxLCx9CrRdNX9PzdKlr6vYp4ehPpYlg4R9D2aHYjQUOsSb0vqUpx0jnE1xOIuSxYsaTSKIomkxitRl9I+G7yzSLCRY0lixYZOdhO5pNBpMXTtaRR96BoNBUoa6bj+aSu7CSjGyFJxd0QtUgmOBoKtLXFow8utmaTQaCZcvzJCWWO+Q/pHw63Df85LNC5GyUyTIkIOSOEzhMxNFukzALVeIqLFQOHYxUOHWlH8zQjeV8meHzveA6SY6LOEytDg1/5IwbVzhM4TKtNpc18oLJmLV6Ul9I+H/Lf8iFms2NkpDZZlO1+pQ02ds5JSi0YB6MQ45s8Vhaspd1+ZoL3Rs3MM3SrpMWfidP3FNehhJqdJPOUVJWZOm27x2JQlHfkbIMhkzE+R/w/pHAeR/yRELO5cbJSG8qUYuBOHVdCjHQmahzFUGuHjP7NRfLxHDSrxWjdFWjOjLTNfl405T8qKFHRHruVKPYw2F0+9PcrUNT1Lcg7xTHI1mKjxKMonhs+jiaiLWVVyh5V0JVZSdmuR7FPchkzFeR/SOB8jIiFncchsbySbdkQWmNiUIzVpK5HT5UOWVjG+7XTFK6E+mePwjrxUoboaadn+VSbdkYenw6aTEm+iLRgvuRn7pruyk7xGeuWEfDxDhlYQ9hKV+r6FWh6xzZT3IZMxfkf0jgfIyIs7lxsbGQpuZCCjsWsTmUn7w42F16ZeJLrGRQWqnFlsmuhG9upjcCqy1w83/6NNOz/JpN9EYegqau9yCcnZF1BWQ5DbasiFNpXZR9UWLCK3+PGXLX6lhD2yTsVKSn1juNWGUtyGTMX5H9I4FrS0IWTZcuN5U6WrqyMbKyNichsg7SJuyZTre8Tl0ujHe9ST7MwUr0ENl+TH4PWuJDf8nhqGhanuRTk7I6RVkNi6sSUF13Iu5TdpFzURd0eIq1WMinK8Ey6IsvfNOxUpqorrcnFq6ZT3IbZSMY0oP6RwHmYhFxsuXyp0r9WJZSY3lHcqytdGvTK6I1ehiXroP7MwU7Uf7OJ1NYp3XJ4hheFPXHZ/kcLR1PUxIS0IbLOTsiMVTX3GyOxHzDqHEKM73PEleEZGHlqoIumdERfXlqU1Vj9yMXCdmQ2yZj/lv6RwG7Fm2Xyp0/ViWTY3mjGStJfwXZCTXQvqozRh5WpP8Ak4vU4hxyLvFPOtSVWDgypB05OL3Xx4Rc5JIpwUI2RTVveGyzbsiMVBDyWwn1KkrSaNZh6lpmNV6DMLK+Hl9hVfucSVyLlvzTpqp19SKtkzH/AC39I4Ddi2zeVOn6sSybGx5oxltUb9hpHr0KV7SX2ZR8khvqaxzMM9VKLzfUx2D4slUR7BI9gkewSPYJdz2CXc9hn3PYZdz2KXc9jl3HhJnstQeHmhpp2eWDp9dZFXY8oR0IY8lsMxTtVZqITtJMrLVRkjBdYTj9hTaRGrZlOouZOw1frkzH/Kf0jgPUWbKVO/VkVk2NjfLjuigxvshXuUOsiltJfYkNjMA70FmnYlG6aJSURPkfI1nW87yow0QSIKyvlTj6seTyTyx/zE/sXExPVTv9jBO1RoYtyiQ8q5k7DGY75T+kcB5WemcIamRWTY2Pmx6vQRTnK1mKxhk+Mil5pL7MkhoaPDPk2+/IjEQeq6It2zbsXGrj6MuM1C2K3nZRjqqJCWSWp2Nugxjyvlj15WWEUlemkUPcxNvuVY6ZyQl1KSKXl50S6GL+W+l/pHAtaLHplGOpkY5t5PluYlaqBZr1Ff1KHSrH+SC/ztfyOPQaHE8N+W19+RFZdeScuthZS5a3nZg4Xnq7EV1ypqyvk8nnHYxivTizSKJR8iJ+5in/ACYmP+WRDdlFdCCaXwH1Rinppu/0jgb6D0F1IRss2+a5uJE1ekxwIRRTjaqn9x9MV/Y4GgcDAKyly1VeN82y95C2ylvyMredmDjaBEXVm3TJ8sNistVI0Ggo/LRi/dxJiPOQhfYo9BO6+DjoOdFtfSOBd4ZU4eubY3yN5JFsmtVNo4RpsyKakVemK/scBwNBhlpb5Wrxazm7IhuLbKpy1fOUvdjZFOLsQhZ3zY82QLaoNHDOGikvcSPEFasmVI6lF/YjTKcGRjZFvgNJ9GYuhwKrXo9vpDB03CHUirsWV+V9RREixYsxLpYcSUbPYiupiumJLXHA0EaaleLOHXo+V3RDFRfSXRiaeSJdGMqshuRyq8s+tUiUuqFk89zSaRIgtzSaSCsjxNWnFlNaqUWRiRi1JiTGvg42hxqXTdfR+HpcSf2RC5FWLly/KkJCQoiSQ6cZ9ZEVFK0UXHK4kmY7piCOyGjSRVi5OnCorSQ8NUpdaT/pkcTp6VFZkZJ9UVFaTGVGQ3IZVuW16xB3KO9s2WbFAUSw0JdSG4kWEeKLpFmE96giKLdcpq5C9rSHHnRj6HBqu2z+jsPT4cClH1+AkJCWbkSqFJ3JRH0ZTbZ4j0rIp+RZWFmmShGatJFPCqnPVF9OxXhaRJEyO5DbKsunLFf5WUyk/eGrGm5oNJbNi3I75Wy8TX+NP7mA+Tk8p1PeFMjVFJSGubHUeLRdt19G4anrn/BCN3YRcvyJCQllc1DkSmaupQfXKUbshE8S+Yv4KPy18HE+hJkyjRc+tyNFxJ03FFVe7yx87IFLzFixbN5M0u5Dfk8RV6P9nh3y2JZ1+lRiYpClYjVOkubFUuDVcfozC09ML9ymrK5fO+SQllcuOQ5DkRhKpsOlKLuymkp9OTxPzRMP8qPJxFsJp8mIg5R6HAmPCzZRpSghpk05KxKjOSHg5nskz2SZ7LM9kmncVCcSmuotudysarkX1FkzH9aDPDH0eTyr0nJ3RwppbCZcuRm0RmmNcnilK8FUXp9F04a5qJCPpzJCWVxyHIchsuYR7laziyla8X9uTxNdYmE+THkkm2KTRGsxTUsqk1CLkz22it2e3UO57fQ7nt9Due30O4sdQb3HjaKe57ZR7ntlHuPGUu57XTY6yIu0yDvEuXLjkkOqOoxyIJy2I9JCyZjOtGR4Zu1k8tFqmvKatJrK4mJkKh/GdanxabgNWdn9FYOHVyIL1Ll84xyuXHIchsuXywhX8jKO8V9uTH7IwfyVydMkxMjM6TVmTwdKouqPwykfhtI/DqR+HUheH0V6DwNJs9hpHsNIeCpEaFOC6IlBDjaZDyjY5DmOQ5DkORQ8h/uhZTK/y5Hh/wAx5PKb6CSKvneaYmJkZ2N+qFljafDrv7/RWGhpppG2e5GNi45DkNjkNl8rM0mF6MxD91lB3kv45PEdomBd6CHmpK9jhM4bNLEmQ3ExsuajUajUzWzWOY6hqZLzog/dG2NjuNDRpLFJpQF5xZTKvy5fweHP/K8mMkJlXzPluJkJWL+qy8Uh0jP6Jpx1SSKayuJNiVhschsbHIuaWxQNJbJT0onVbRQ86XJ4l5F/JgPkIeeksWLCRpEN35LrkZJCJedEFdEhjY3l1LFmUxbZVHYqO8H/AAeHO1XJjK8moXIVqjdkcNNdSVDsSg1yXEyEssdDXQf2+icJG879iPRZJXFaKHIchyL5KIoluRvPDxbnq5PEflr+Tw75BLL0EunKuS45FzUNlxjyl8xFN9GN5PkSuKjJ79CEUlYWVV2Q5XTMB84Y2NkEpbijFbLNxTJ0V6Di1yJkJX6MktUHFjVnb6IwUejYzcS0jkORcuJXFEtyJN7CpTYqHdiowRwoNWXQpQUfW/Jj/l/2eG/JJsuXFKyFNMtzMbMRjVTemO5PFVZPchiakXuYfGcTpLfJjyl50UvUlyRpynsRw6XmOkdskL0yruyRcwXzxj3GKaiyMpJ7ZNPOUUydNxzTIsT9TER01pL7/Q6hJ7IwitGw2RWlEpDkXyURLJJvYVNipr1Eor0LmtDqHFZxWU6rcuTH/L/s8N+UVHYvkn7qyU2hSTze+TZiavDptjbbzpyszD1NcBjyl50UXuSGKLlsRopblxsuLJbLLE+VCMJ0rjGM2TkcRiqNCrMVZMvFjiSiVKfqskxMizG05cVyS6fQ0VeSRGkkhLSiK9WSkNlxK4okYNmhLdiilldGo1DeV86XnQhvLHfLf8nhnyitsXyT6LNEZCYxkjHeRckdzB+V8kvOiluMjDUxJIci42XNQpXFsssU9kRML88Y9xkvJypiqNCq9x6ZFSg94jTQmQZpUmY7DqlLUtn9C0leRFdCxJ2JSGyKcmQov1Iwii5c1Fy5cvy2Ka94RLLG/Lf8o8M+WyttmuqXLFjyZiqeuDQ1bOCMPDRAecvOiluOIuhcbG+SGxHyrLF+iI7mH6YgYxvqPyv4F2Kox6Zk6TiRKZj6eqg32+haPr/BDy5TbHchQb6shGMNjUai/JYtz0/MhMeWN+W/5R4Z8tlbOEfcRpNJbJDyZJGIwin1juToVIbojSnJ9EUMJp6yLDzl50UvNysui5cg+hHrFZYuPRMg+pR/9H95SQ9yOzNKNBoZpZbmU+44+qKWxXjqpyX2+haflkUvKiSHFCikNly5f4FuWHmI7Dyxvkf9Hhvy2VT1ypeRZWNJpFEm7M1Mvm4p7o0pbLK42ai4/OUdx5sfJT2IeVEUV4a4NGzKP/o/vKQ9ynuzSaTSWHFDghwGuRFF+hLqiatJr6Ep+qKPkQ3cY2NmouLkXLbkh5iOw8sZ5Jf0eHeRlTYWVHyLmnvy2LZNDzfnRSHk8nyQKfWOdbpUZS+f/Zew5DKe7zsWLDWTQ45XEU97mrq0VlarL+X9CQKK9xDiOLHFjTEhISzt8KG5HbJoxitCX9Hh3kZMSLFLyLkWUtzqWZZlixbJjzfnRTHk8nyQKfkzxC/yMpfP/skPKju+drKw4XOGxQZFadyjJznJmI+dL+foSm7Mo+VZMaGhL48NyGwyxjvly/lHh3y2WYkNFLyLkWUt+RDzY8350Q5XyQKflzxUHrTRS+f/AGPYZYpbvneaKkZW90/y9i0n5kQgomI+dL+foSJS8iL3LjyWS+G8obiTcehJO24k+54h8p/yjw9N03YVOSXVjg2tzT0sU/KuaW/OxjyY/OiGTzfJApeUqwnPyuw6Na/SZwautNzuiH/o/sZJMlSrud9fQp787zk7DmzUyEsq/WrL+foSHlkQ8iLkXct+RhuR2Hl4h8k8L8jJbZx25pb87GPJj85Tyeb5IFLbOoryKf8A6P7yZYguvwGN2G8kLc9Co7zb+hKe0iPlRYvYjJMsWLfGhuQ8o8vEfk/2eGeRjzjtyXETTuWkJSFf15NXUY835yiND54FHbOW5S/9H952I87Y5DziiFndk5Wi39C0n1EIkjqn0KdVPozoxxLFviQ8wuiz8R+T/Z4b8t5suXeaNVkPkeTHEY835yj0kPqizGhrkuQKW2bSKX/o/vNkeRtDmhyY87CiSfohbGLnooy+hYeZCkRd0PqNDRCtKJCopFkzSWLFvg0/MLPxH5P9nhvynnc1WNWaQ9hiyauafuaRxGn3GPN+cpvqQZKI4j5OhBdCn5cmXKHXEf3yJ2HNHEHJ81hQJz09EJEpWMfWulBfQ0a84qxhpOcLsRJDQ0JuJCqRqCaZpHE0lixbkRT83J4l8pfyeHfJ5GLJLKTHzNlXFKPSPU9on6ohUU1m/OQE7EJKSHEcBxHEtlDYp7ZtGG+fyS25rGkULllEnU9FkjGYmUammI227v6HwMvday3Q0NDRsRkRmxTFMumaRwHEcBotlT83J4l8pfyeHfJ5LFi2Tdsny3MRNqLsRj0uN3Iy0u5e+X+5DJS0shUUhxuSgOA4mhsUbENs2YX5/JPbOxYSFE0pE6qiSqOQlniXetL6IwcrVLd8thq40NFslITExTaFUFM6M0ocB0xwaKa97k8S+WjAfI5EuhbJy52SZUWpNEH0sbCV2JWyXzCKzTaKdf0YmmNDiaSxHbN7Mwfz+SRYURRNB0RKqidRssJFsn0RJ6pN/REJaZJkXdZJ2HEaLFiwm0KQnkmxVGhVRVBSGtWwqdt+TxPyRMB8hci2Ljl8BjQ0To326HBl6sp0lAeX/wAguWNSUSNZM1HTKO2cvKzA/O5HawpxvY6IdRDqMb7jlfYSLZpGJlooyl9FYSeqms4saGhosWyTaFIvyKckLESRQxHEbT5PE/LEwXyFyOTFmh8zLFi2Usv9yC6Esrmo1FyM3HYVdepF9CO2dTyMwHzeSTsamdXk5JDbYkWLZJD7HiU7U1Hv9FYOdpaeRSLFjSNFixYTaFIUi5GGpE04lCNp8nim0TCfJjzXRqQpEnzvNjGL5hHYY8uhYsWLMpt2sR2zrfLl/B4d5+SabaNA52G2ywkJcmyErHiM9VXT2+ioScZJopyUo3XJGVhdRo0jiNDRbO5husSvC0Wykve5PEt4mF+THktZXNTLlyJLohMuXyvzsh52LYRKA1yXKVPVC5HzWI7Z4j5Ujw9Wm+R5SXVlsrciVurN+o+iuVZ66jl9F4Kpdae3Km0Rdyw0aRoaGs8NsV37rIK0uTxDzIofKXOlcS0ory0wOOzjnHOOcc46OOjjnHHXPaEziXKe5HbJO3RjhccRoaLFH5YvOR2zxPypGA3fI1lJdS3KlbqzfLHVOHRf3+jKU+HNMjK65UyMsrDiNDWeG8piPKQ83J4i/wDIkUflrNj6CVxU+5dLoj7sxUuiLly5cuXLl8mQfvyIFMWxYsJtCaluSpEqbRYg3axCL13FnjHaizw7ZvmfJYUbdWPrn4lV1TUO30bgpuX9cyYmKRYcRwHEaFJrYUnJ2ZSd5Pk8Q+cin5VycNF4rYd3uJWGYnzfCZF/5ZIp7FHfmUmjUnuOCYo2ZFCz8Q+SeHeR8kt+VRbLKJvnOahFyZUm6k3J+v0bgd2LmTEKQncshwHAlAS6lHkx3XEEds2dXvm2krsjNN29SrLVN/Buhl7ViHREPd6mq5fnUmRd3yeJfKX8mAVqXI115FAckuiP55PEq2mCgvX6OwSvUF8BSyUhPKVO5KFmUuTFdcULN5XROvGPTdlqtXfoiFCFN6vUlFXb+A9hwF5SbtUIiZDbO5fkRDfk8SfuxRglakuZRLKI5OQlyN2V2Ymtxqrl9HYJpTYhfAUi5cUxSNKaIx08lXri/wC+RjrLaPU4c5+cjShDZZSdk/gXyY9ir52UpNwTKa1dSHNcYilvyeJPrFGFVqS5JNIi01cc+3P4lX0Q0Ld/R+CT1sWS+CpFxSsKrIjLUuRe9jP75JRU1ZkYKKsuTH1XCi7Pfocap3OLPucWfc41Tucep3ONU7nFn3OLPucWfc4s+45yfrlgp7xIEH1+BEpb8niHzUiirQWSycmb87PEW3U+j8D5mLkt8G5cpbclDri/75FvyNniVS7UP7+LhHaoITsbrnhsUuTF+9ibEeiWctvgzfoeIW4n3+j8G/8AILNMRb4VLy8mC64i/IuSb9DEVOJVlL4tKWmaZfKm7q2b5I7FPkn72Kf8izl8FniD/wAtvo/B/NFyRYixb4FLpHObtFnh6vUb5FyYipppyn8ehPXTTN0Rdnm87CIbclD3sTf78nrmuaRj/nfR+D+aLkZCWbXPDy513alL+Dw1dW+SOc3/AKo8Sq2iqa+Pg6lpaX6i7ZU5eg+SwiG2c3aLZgetW/It/gyMf876PwfzRcjERlfJoa5UR2zxbtRkeHL3W+SOU5KCubK8jEVeNVcvjptO6KVRVI3Rv1yi9S5o7Z4l2oyPD1u+SPOspGP+d9H4OD1amReVy5cRchNPJoa5FuLbPHu1E8P+ULODJSUVdivJ65f0eIYjTDhrd/kaFXhS+xCSf8DQnpYnqXKs8c7UWYCPuci5W7IiXG0SZj6b16/o6jRdR/YhBJWEsnyptEKilk0POO4s/EX/AI0jBK1FciahdsvKpMrVY0oOTKlR1JuUvyWGxGj3ZbEXcaIycRPVsMRvJLk8Rl7sUYRaaayQ+VuxvyNFSmpqzK9F0pW9PoyKcnZFGnojYSEsrFhI0Gg0GixGV8mhrKHm5PEn5UYdaaUVm2On7+/X/wDCc4wRicQ68vstvylDEOHSWxFqauhidmboWxDrPkx71VVEoxtFLnbsOUma2vQVaPqJplhoaMTRVSI007P6LwdK/vMSEskiwoGkaLckJ+jylEaKS68mPequkQVopZzv6E6ihF+ndmJxTrPTHy/lqOK0JJkZxmroaHddUKaaKHVt5oqf5MUyKsuZuxe+bSa6mhwd4Mp1FP8AksNElcxtLTLUvovDw0wSIrJCIxyYx8sKlujF1JUymrclT/Ji+R+ZHiNSTqaL9PzFGq4GHxer3ahZNCVmUFaObdk2YVa6rlyXSV2KpFuyJS0fyOTNTNTNTNchzY5SUtSIS1RuNDRi6euD+iqUdU0iCyQhEcmMfI8krDrWdiLUlfNuyuYT367lmzH4mcKiUGSlKbvJ/mIFjC4lx9yWxUxL16IEFaKWeLnooswMLK/ImM0mhGhGhGhHDRw0cNEFZZNFRdCrHTNr6JwivVI5IQsr8jGiw1lxpEXqbbIKyzxE9NJswEd5ZXGzEz11pP8AMx3zwcNdZfbkx0rtQRRjoglmxSed8r8iEsmTRi1aq/onBfMFkhC5EaUaEaEcNHDQ6SHSFDqLPHT91QRhlopmocirV0Qb/Nxd1lgKWmGp+o2XNSRG9as5Czkna6FqLy7F5djVLsapdjXLsOpP0Qp1H6EU/USyYyRjfm/ROB87FkhC5I81jScM0ZSiV03WsxVElYdVDrIxVfUtC/NwlYo0uI/sRkoqyNZqRiqtoWRhYiLFuW2ViwkIWbJGN+Z9E4H5jFkhC5I/Bk7Futx7GNb48jVLuXf5/VLucSfdlJt0otlZe6YTyiyXwUIWbJGN+b9EYekqtRRZSw0KStHNCFyRzXNM9R7GN+fL9DpL/DEqbGF8ghi+ChCzZMqYONZ6mVqTpTcH9D4D56HGw8lkuSPIuWZ6j2Mb8+X6HS+TEqbf2YXyC+Is2MkRskeJJcZPuvofAfORIeSELkjn05pj3GY358v0Ok/8MSexhfIL4iFmyRKGuNjxKOmUV9vofA/PRKQ8kIXJHO3NJj3HsY358v0OjVjw1FkvQwvkF8J5IWbJEWeJyvVX8fQaVxUpv0OBU7FDB61ebsU8NCjLUmRncvkhC5I5XL80h7j2MbBuvI0SGmt/zqgzhshh3N2uPBRS3IQT6XEp0/ujC+RC57l+RMXJIrYnhdErlXiVZubQ4tbr6Coro2hTdhzZGs0h4goNy6iyQhCzjy2LFsmiUkpWHsYilOVVtIcbdGTipKw007fm6XmHD1RsRuam+lyVNiqOPRmGfuoieg5o1o1GpGpGtGpGpDkKSExcjMXFqSYmXROin5Rxcd/oDCbsjFdUx00Omilho1H1ZTwkYbMnDRkmJikJmo1IU0jio4hxDiHEOIcRmtmplxuVypRpzd2zE0YQV4yNQ4qY1Z2/NUfPl0kPpuhTSINSRGhTnL3iEIRXQv16CbNT7Gr7Gr7Gpdi8exePYvHsXj2Pc7CcOxqicRHFRxTiDmYlybsYagpt6z2WkjEYeEI6oDjqJYd7xGmnZ/r2Edm2UYxnJtio03uhYal2I4aMdjhfc4P3OAu5wYioo4SOGjQjQjSjSjSjSjSiyLIsWysWNI4IlSj2QqNPsjxSlCNJSS63/NUvMRGUEm7CpQ7CpQWyOHFbIUTSaTSaTT9zT9zQ+5o+5pZoNJoZoZoNLNLNLNLHTT3RGmo7I0rsSpqSs0ez0/2ioU+xjsPT4Lduq/XqMJaGzCkBb5r8sxHi3yf7/NQlplcjaUbxLswjvOwhD/MoxvyJfr2HX+JFBWmyAvgXLly5cuXLly5cvzo8W+Uv5+Jol2NMl6fGpycYoUY1Y3RhqTjVuIQ+R/lUY35Ev17D/KRTai+pCaezEy5dFy5cuXLly5cuXLly5fK+d0XRcuJ2XVnis1w1H4NOhOZTwsFv1I00tkKlJ+h7NJksFf0KuAktkTpyhv8ADj5SldFKav1E0xMbNSFJGpGpGpGpGo1I1GpGpGpGtGpGpGpGpGpGpF13Lrual3NS7mpGuK9TGTjKk0mThKG/67h/lIpbCS7ChHsaI9jRDsaIdjRHsaI9jRHsaI9jRHsaI9jRHsaI9jRHsaI9jRHsaI9jRHsaI9hxj2NEexoj2NK7GmPY0R7GmPYcY9hRVtjxaKUIvnSuUaKXVkUQoN9WRhGO3LOjCorSRiPDWutIacXZ/BpJySQqWmOVJLSI0q2w4x7GiPY4cexw49jhw7HDj2OHHscOPY4cexw4djhw7HDh2OHDscOHY4cOxw4djhx7HDj2OHDscOHY4cOxw49jhx7HDh2NEexpjbY8T3j+u4TrSRS2Ii/MsjsYujGrGzH4evRnsH3PYPuewfc/D/uLw5dyGAUfUWGsQhoNbNTNTNTNbNcjXI1sr0I1vMh+Hrufh67nsH3PYfuew/c9hXc9iXcWCj6spU4wVkJjhcp7CPTJP8x6Hie8f13CVlD3GUtiIvzLIk0OBoNBpNKNJpLFixpRpNJpNKNKNBoNCNBoOGaBwHAcSwhERHp+b8S80f16hjnTVpK5S8RpydiNeJxUcRHFQ68UfilBMhjaU1dM9pge0QPaIHtEBVEzWh1YnFicWJxInEicWJxYnFicWJxYnERxIjmhSLNjXTK5c1GtCqHEOKjiI4iNZrNaNRqLly/JYsOI4lhLJSFUFURrFNGtHERricSJxInEicWJxoCqw7ikmakOpFHFiceA8fQi7Nn4jh+4vEcP3I16ctma4mpGpDnFEsbQi7Nj8RodzF11XldfQGBk3B3ZSi5K9zQ+5oseJuUdKTywkt0LNblOPQcWTjYuai5qLly5cjHoKJGmhwSNhT07ikpI0mg4ZwzhHBOCcI4BwBUbConCFTOGcM4ZoNJpLFiw0NDjnch1FFml9xRfcl0Lly+bLkSEXY0sqRsNFR6Ytjd3fPB/PQokIJmhHinRx+g6GIdHpboUvE4RXVM/FaPZj8VpdmYvFvEtdLJZUJaZoiLJblPbKoi2dmWLFixGXQTFOw53ynT1shHSvjXL5XLly5cuXLjzaNJH3TWazWS6lixYsWHlEproWKyGY2emFu/JSqOlNTXoLxN+sReKpf6j8X7QMTipYlptWt9FYeeuIsluU9spCiOBpzsWEhL4NxClfY1Gr0GxS62HKxqLlzV0udS5q3eady5fpcuXLly/xXlEpvplV2JGKqa6n8fSWGqaZWYsvUp7LJiRYsWNJpNJpLFixYsWLFixYsRVmW6lutxxLe8OPQcdixpNN4pDTWxZ7o09GaWaS25p6jjeJpNJpNJpNJYsWLFixYsWLDRLKJT2yqeUxdbhw6b/AEpha2tWe+UtynssmJfknksny3Ftks0LmexYSGixYSLCRYsWLDJZRKWVTyMr1XVnf6UhNwlqRRqKpDUie5T2Wbdi/K38ZZP4iH8CxYsaTSbGpCknmyTyiUspK8WNWbX0rhHaI+pS2WT2HIUhSNQ5oQ4p9GcNRXT8zbK3xWyLuaUaIjshyJSJSEyJSy9CsrVJL7/SiVyhFxhZi2KPlWU/KOYpnEJVSE7sjtk2XLly5cuXyuXLly5cuXLly/JcuXLly5cuXLly5cuXLlyUim8pbDmOZKY5CZF9ClvnilavL6Uw/wA1HqyOxR8qyn5WVE0a2OoOoUZdSG2U3Y1Go1Go1Fy5c1Fy5qLly5cuXNRqRqRqNRqLlzUajUajUajUajUajUahzL3KbylsSkOY5ZIhsUt88erYiX0phvmoaILoUPLlLYmSgiaLFLpIp7ZYiagcePc48e5x49zjx7nHj3OOu5x13OPHucePc48e5x49zjx7nHj3OPHuKsmcQ4hrNZrOIaziDrI48e5x49zjx7nHj3OPHucePc48e5x49zjx7nHj3OOu5xo9ziI4hqEUn1EMnFDQ1nDYp9GLLxNWr/SlKWiakNESh5cpeUcyUiWVNdSntl4t/r8eM5Qd4so49bVEQq06nlZZFiyOhOrTh5mVcfHaCuVK06nmfxaFXQ7PYtcjdMTKT6iyquzGx5IiR3I7ZeKq1RP6Vw8+JTQihtk9hxNJpNCIxSIzscQ8U96Cf5KNerHaQsZW7jxld+o69SW8n+Rw9a3uyLWZEh0YpnEKjUhxQ4mg0CIojtl4svK/pXB1NM9L9cqUrI1GoauaTSaTSWyrUuLBxGnF2f6NQxFvcmIsKWTQ0aWaWaWWI2QpGo8SSnS/j6VTt1KNRVYKRTXU0mk0GghEluNiV2cMcUjxCioT1rZ/o+Gxbp+7LYjJSV1lBmi46aIwsySGbsUDSjSSpKS6mNwXB9+O30rha3Cn12ZT3QuRE2NlLfJnivlj/P6RQxEqL6bFKrCqrxFuLbJIkSZHct0NKLIsjxH/AM7+lvD668kntsKSLoui6KtaNKOqWxLxGiz26kYarGfVF0NnirWmK/SadSVKWqJhsVCt9mLbJtLqT8SoDx9FlLFUpzST5fEv/O/pZNp3R7RV/ce0Vv3M9prfuZ7TW/cydWpNWk75wqTh5We0Vf3HtNb9zJSlJ3k/0q9uqFiqy/2Z7XX/AHMeJrSVnJ5p22Paqy/2Z7XX/cz2uv8AuZ7XX/cypXq1Vabv/wDRE//Z"

EXAMPLE_PUMPS = {
    "Wilden PS820": {"caudal_lpm": 409.0, "p_descarga_bar": 1.4, "p_aire_bar": 4.1,
                     "consumo_aire_nm3h": 88.0, "litros_ciclo": 5.6, "color": "#00d4ff"},
    "Sandpiper S20": {"caudal_lpm": 409.0, "p_descarga_bar": 1.4, "p_aire_bar": 4.1,
                      "consumo_aire_nm3h": 102.0, "litros_ciclo": 3.4, "color": "#ff9500"},
    "SAMOA UP20": {"caudal_lpm": 409.0, "p_descarga_bar": 1.4, "p_aire_bar": 5.0,
                   "consumo_aire_nm3h": 140.0, "litros_ciclo": 5.1, "color": "#a855f7"},
}
PUMP_COLORS = ["#00d4ff","#ff9500","#a855f7","#00e676","#ff4f7b","#ffeb3b","#76ff03","#18ffff"]
CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,28,40,0.6)",
    font=dict(family="IBM Plex Mono, monospace", color="#b0cad6", size=11),
    xaxis=dict(gridcolor="#1e3a4f", linecolor="#2a5068"),
    yaxis=dict(gridcolor="#1e3a4f", linecolor="#2a5068"),
    margin=dict(l=60, r=30, t=50, b=50),
    legend=dict(bgcolor="rgba(13,28,40,0.8)", bordercolor="#1e3a4f", borderwidth=1),
)

# ─────────────────────────────────────────────────────────────────────────────
# CALCULATIONS
# ─────────────────────────────────────────────────────────────────────────────
def calcular_kpis(caudal_lpm, consumo_aire_nm3h, litros_ciclo):
    try:
        if float(litros_ciclo) <= 0 or float(caudal_lpm) <= 0:
            return None
        ciclos_min = float(caudal_lpm) / float(litros_ciclo)
        aire_especifico = float(consumo_aire_nm3h) / (float(caudal_lpm) * 60 / 1000)
        aire_por_ciclo = (float(consumo_aire_nm3h) / 60.0) / ciclos_min
        return ciclos_min, aire_especifico, aire_por_ciclo
    except (ZeroDivisionError, TypeError, ValueError):
        return None

def build_dataframe(pumps, lang="es"):
    rows = []
    for p in pumps:
        kpis = calcular_kpis(p["caudal_lpm"], p["consumo_aire_nm3h"], p["litros_ciclo"])
        row = {
            t("col_pump",   lang): p["nombre"],
            t("col_flow",   lang): p["caudal_lpm"],
            t("col_pdesc",  lang): p["p_descarga_bar"],
            t("col_pair",   lang): p["p_aire_bar"],
            t("col_cons",   lang): p["consumo_aire_nm3h"],
            t("col_lpc",    lang): p["litros_ciclo"],
            t("col_cpm",    lang): round(kpis[0], 2) if kpis else None,
            t("col_aesp",   lang): round(kpis[1], 4) if kpis else None,
            t("col_aciclo", lang): f"{kpis[2]:.4e}" if kpis else None,
            "_aire_esp_raw":  kpis[1] if kpis else None,
            "_ciclos_raw":    kpis[0] if kpis else None,
            "_aire_ciclo_raw":kpis[2] if kpis else None,
            "_color": p.get("color", "#888"),
        }
        rows.append(row)
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────────────────────────────────────
# FTMg CSV PARSER
# ─────────────────────────────────────────────────────────────────────────────
def parse_ftmg_csv(uploaded_file):
    try:
        raw = uploaded_file.read()
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        first_line = text.split("\n")[0]
        sep = ";" if first_line.count(";") > first_line.count(",") else ","
        df = pd.read_csv(io.StringIO(text), sep=sep, decimal=",",
                         on_bad_lines="skip", engine="python")
        df.columns = [c.strip() for c in df.columns]
        flow_col = next((c for c in df.columns if any(k in c.lower() for k in
                         ["flow","caudal","flujo","nl/min","l/min","volumen"])), None)
        time_col = next((c for c in df.columns if any(k in c.lower() for k in
                         ["date","time","fecha","hora","timestamp","zeit"])), None)
        if flow_col is None:
            return None, None, None, f"No flow column found. Columns: {list(df.columns)}"
        df[flow_col] = pd.to_numeric(
            df[flow_col].astype(str).str.replace(",", ".").str.extract(r"([-\d.]+)")[0],
            errors="coerce")
        df = df.dropna(subset=[flow_col])
        if time_col:
            try:
                df[time_col] = pd.to_datetime(df[time_col], infer_datetime_format=True, errors="coerce")
            except Exception:
                pass
        return df, flow_col, time_col, "OK"
    except Exception as e:
        return None, None, None, str(e)

def nm3h_from_ftmg(df, flow_col, is_nl_min=True):
    mean_flow = df[flow_col].dropna().mean()
    return mean_flow * 60 / 1000 if is_nl_min else mean_flow

# ─────────────────────────────────────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────────────────────────────────────
def chart_kpi_bars(df, lang):
    kpis = [
        (t("col_cons", lang), t("kpi_air", lang), "Nm3/h"),
        (t("col_cpm",  lang), t("kpi_cycles", lang), "cpm"),
        ("_aire_esp_raw",     t("kpi_specific", lang), "Nm3/m3"),
    ]
    fig = make_subplots(rows=1, cols=3, subplot_titles=[k[1] for k in kpis], horizontal_spacing=0.1)
    colors = df["_color"].tolist()
    for col_i, (field, title, unit) in enumerate(kpis, 1):
        vals = df[field].tolist()
        names = df[t("col_pump",lang)].tolist()
        max_v = max([v for v in vals if v is not None], default=1)
        fig.add_trace(go.Bar(
            x=names, y=vals,
            marker=dict(color=colors, line=dict(color="rgba(255,255,255,0.15)", width=1), opacity=0.9),
            text=[f"{v:.2f} {unit}" if v is not None else "-" for v in vals],
            textposition="outside", textfont=dict(size=10), showlegend=False,
        ), row=1, col=col_i)
        fig.update_yaxes(range=[0, max_v * 1.3], row=1, col=col_i, title_text=unit, gridcolor="#1e3a4f")
        fig.update_xaxes(row=1, col=col_i, gridcolor="#1e3a4f")
    fig.update_layout(**CHART_LAYOUT, height=370,
                      title=dict(text=t("chart_kpi", lang), font=dict(size=13, color="#00d4ff"), x=0.5, xanchor="center"))
    fig.update_annotations(font=dict(color="#8ecfdf", size=11))
    return fig

def chart_radar(df, lang):
    metrics = {t("kpi_air",      lang): t("col_cons", lang),
               t("kpi_cycles",  lang): t("col_cpm",  lang),
               t("kpi_specific",lang): "_aire_esp_raw",
               t("col_lpc",     lang): t("col_lpc",  lang),
               t("col_pair",    lang): t("col_pair", lang)}
    fig = go.Figure()
    for _, row in df.iterrows():
        vals_raw = [row[v] or 0 for v in metrics.values()]
        maxes = [df[v].max() or 1 for v in metrics.values()]
        vals_norm = [v / m for v, m in zip(vals_raw, maxes)]
        labels = list(metrics.keys())
        vals_norm.append(vals_norm[0])
        labels_c = labels + [labels[0]]
        fig.add_trace(go.Scatterpolar(
            r=vals_norm, theta=labels_c, fill="toself", name=row[t("col_pump",lang)],
            line=dict(color=row["_color"], width=2), opacity=0.8,
        ))
    fig.update_layout(**CHART_LAYOUT, height=370,
                      polar=dict(bgcolor="rgba(13,28,40,0.6)",
                                 radialaxis=dict(visible=True, range=[0,1], gridcolor="#1e3a4f",
                                                 tickfont=dict(size=8)),
                                 angularaxis=dict(gridcolor="#1e3a4f", linecolor="#2a5068")),
                      title=dict(text=t("chart_radar", lang), font=dict(size=13, color="#00d4ff"),
                                 x=0.5, xanchor="center"))
    return fig

def chart_scatter(df, lang):
    fig = go.Figure()
    for _, row in df.iterrows():
        ae = row["_aire_esp_raw"]
        cpm = row["_ciclos_raw"]
        if ae is None or cpm is None:
            continue
        fig.add_trace(go.Scatter(
            x=[ae], y=[cpm], mode="markers+text", name=row[t("col_pump",lang)],
            text=[row[t("col_pump",lang)]], textposition="top center",
            marker=dict(size=18, color=row["_color"], line=dict(width=2, color="white"), opacity=0.9),
            textfont=dict(size=10, family="IBM Plex Mono"),
        ))
    fig.update_layout(**CHART_LAYOUT, height=370,
                      xaxis_title=t("chart_scatter_x", lang),
                      yaxis_title=t("kpi_cycles", lang),
                      title=dict(text=t("chart_scatter_title", lang),
                                 font=dict(size=13, color="#00d4ff"), x=0.5, xanchor="center"))
    return fig

def chart_waterfall(df, lang, ref_idx=0):
    ref = df.iloc[ref_idx]
    ref_ae = ref["_aire_esp_raw"]
    if ref_ae is None:
        return None
    others = df[df[t("col_pump",lang)] != ref[t("col_pump",lang)]]
    if others.empty:
        return None
    names, deltas, colors_w = [], [], []
    for _, row in others.iterrows():
        ae = row["_aire_esp_raw"]
        if ae is None:
            continue
        d = ((ae - ref_ae) / ref_ae) * 100
        names.append(row[t("col_pump",lang)])
        deltas.append(d)
        colors_w.append("#ff6060" if d > 0 else "#00c878")
    fig = go.Figure(go.Bar(
        x=names, y=deltas,
        marker=dict(color=colors_w, line=dict(color="rgba(255,255,255,0.2)", width=1)),
        text=[f"{d:+.1f}%" for d in deltas], textposition="outside",
        textfont=dict(family="IBM Plex Mono", size=11),
    ))
    fig.add_hline(y=0, line_color="#00d4ff", line_width=1, line_dash="dot")
    fig.update_layout(**CHART_LAYOUT, height=340,
                      yaxis_title=f"{t('chart_wf_y', lang)} {ref[t('col_pump',lang)]}",
                      title=dict(text=f"{t('chart_wf_title', lang)} {ref[t('col_pump',lang)]}",
                                 font=dict(size=13, color="#00d4ff"), x=0.5, xanchor="center"))
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(pumps, df, lang="es"):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    import tempfile, os

    wb = Workbook()
    C_HDR_BG, C_HDR_FG = "0D3B55", "FFFFFF"
    C_TITLE_BG, C_TITLE_FG = "1A3A4A", "00D4FF"
    C_ALT = "F0F7FA"
    thin = Side(style="thin", color="B0C8D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, ref, text, bg=None, fg=None):
        c = ws[ref]; c.value = text
        c.font = Font(name="Calibri", bold=True, color=fg or C_HDR_FG, size=10)
        c.fill = PatternFill("solid", fgColor=bg or C_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def cell(ws, ref, value, align="center", bold=False, num_fmt=None, bg=None):
        c = ws[ref]; c.value = value
        c.font = Font(name="Calibri", bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if num_fmt: c.number_format = num_fmt
        if bg: c.fill = PatternFill("solid", fgColor=bg)

    pump_col   = t("col_pump",  lang)
    flow_col   = t("col_flow",  lang)
    pdesc_col  = t("col_pdesc", lang)
    pair_col   = t("col_pair",  lang)
    cons_col   = t("col_cons",  lang)
    lpc_col    = t("col_lpc",   lang)
    cpm_col    = t("col_cpm",   lang)
    aesp_col   = t("col_aesp",  lang)
    aciclo_col = t("col_aciclo",lang)
    dcons_col  = t("col_dcons", lang)
    daesp_col  = t("col_daesp", lang)

    ref_consumo = df[cons_col].iloc[0]
    ref_ae      = df["_aire_esp_raw"].iloc[0]

    # ── Sheet 1: Full comparison ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = t("xl_sheet1", lang)[:31]
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:K1")
    c = ws1["A1"]; c.value = t("xl_title1", lang)
    c.font = Font(name="Calibri", bold=True, size=13, color=C_TITLE_FG)
    c.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    headers1 = [pump_col, flow_col, pdesc_col, pair_col, cons_col,
                lpc_col, cpm_col, aesp_col, aciclo_col, dcons_col, daesp_col]
    widths1   = [22, 13, 15, 13, 17, 10, 13, 20, 16, 16, 16]
    for i, (h, w) in enumerate(zip(headers1, widths1), 1):
        hdr(ws1, f"{get_column_letter(i)}2", h)
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[2].height = 32

    for ri, row in df.iterrows():
        r = ri + 3
        bg = C_ALT if ri % 2 == 1 else None
        kpis = calcular_kpis(row[flow_col], row[cons_col], row[lpc_col])
        vals = [row[pump_col], row[flow_col], row[pdesc_col], row[pair_col], row[cons_col], row[lpc_col]]
        for ci, v in enumerate(vals, 1):
            cell(ws1, f"{get_column_letter(ci)}{r}", v,
                 align="left" if ci == 1 else "center", bg=bg,
                 num_fmt="0.00" if isinstance(v, float) else None)
        if kpis:
            ciclos, ae, aciclo = kpis
            cell(ws1, f"G{r}", round(ciclos, 2), num_fmt="0.00", bg=bg)
            cell(ws1, f"H{r}", round(ae, 4), num_fmt="0.0000", bg=bg)
            cell(ws1, f"I{r}", aciclo, num_fmt="0.0000E+00", bg=bg)
            dc   = row[cons_col] - ref_consumo
            dae  = ae - ref_ae if ref_ae else 0
            dpct = (dae / ref_ae * 100) if ref_ae else 0
            for col_l, val, fmt in [("J", dc, "+0.00;-0.00"), ("K", round(dpct,2), '+0.0;-0.0')]:
                bg_d = "FFCCCC" if val > 0.001 else ("CCFFCC" if val < -0.001 else None)
                cell(ws1, f"{col_l}{r}", val, num_fmt=fmt, bg=bg_d or (bg or "FFFFFF"))
        ws1.row_dimensions[r].height = 18

    # ── Sheet 2: KPIs ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(t("xl_sheet3", lang)[:31])
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]; c.value = t("xl_title3", lang)
    c.font = Font(name="Calibri", bold=True, size=13, color=C_TITLE_FG)
    c.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28
    headers2 = [pump_col, cpm_col, aesp_col, aciclo_col, dcons_col, daesp_col]
    widths2   = [22, 14, 22, 18, 18, 16]
    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        hdr(ws2, f"{get_column_letter(i)}2", h)
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 30
    for ri, row in df.iterrows():
        r = ri + 3
        bg = C_ALT if ri % 2 == 1 else None
        kpis = calcular_kpis(row[flow_col], row[cons_col], row[lpc_col])
        if kpis:
            ciclos, ae, aciclo = kpis
            dc   = row[cons_col] - ref_consumo
            dpct = ((ae - ref_ae) / ref_ae * 100) if ref_ae else 0
            for ci, (v, fmt) in enumerate(zip(
                [row[pump_col], round(ciclos,2), round(ae,4), aciclo, round(dc,2), round(dpct,2)],
                [None,"0.00","0.0000","0.0000E+00","+0.00;-0.00","+0.0;-0.0"]), 1):
                cell(ws2, f"{get_column_letter(ci)}{r}", v,
                     align="left" if ci==1 else "center", num_fmt=fmt, bg=bg)
        ws2.row_dimensions[r].height = 18

    # ── Sheet 3: Bar chart — Air consumption ─────────────────────────────────
    ws3 = wb.create_sheet(t("xl_sheet2", lang)[:31])
    ws3.sheet_view.showGridLines = False
    # Write data for chart
    ws3["A1"].value = pump_col; ws3["B1"].value = cons_col
    ws3["C1"].value = cpm_col;  ws3["D1"].value = aesp_col
    for ri, row in df.iterrows():
        r = ri + 2
        ws3[f"A{r}"].value = row[pump_col]
        ws3[f"B{r}"].value = row[cons_col]
        ws3[f"C{r}"].value = row[cpm_col]
        ws3[f"D{r}"].value = row["_aire_esp_raw"]
    n = len(df) + 1
    for col_l in ["A","B","C","D"]:
        ws3.column_dimensions[col_l].width = 20

    def make_bar_chart(ws, data_col, title, row_start=2):
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.style = 10
        chart.grouping = "clustered"
        chart.width = 18; chart.height = 12
        chart.plot_area.spPr = None
        data = Reference(ws, min_col=data_col, min_row=1, max_row=n+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        return chart

    chart1 = make_bar_chart(ws3, 2, t("kpi_air",  lang))
    chart2 = make_bar_chart(ws3, 3, t("kpi_cycles",lang))
    chart3 = make_bar_chart(ws3, 4, t("kpi_specific",lang))
    ws3.add_chart(chart1, "F2")
    ws3.add_chart(chart2, "F20")
    ws3.add_chart(chart3, "F38")

    # ── Sheet 4: Methodology ─────────────────────────────────────────────────
    ws4 = wb.create_sheet(t("xl_method", lang)[:31])
    ws4.column_dimensions["A"].width = 80
    notes = [
        (t("xl_method",   lang), True,  13, C_TITLE_FG, C_TITLE_BG),
        ("", False, 10, "000000", None),
        (t("xl_formulas", lang), True,  11, C_HDR_FG, C_HDR_BG),
        (f"{cpm_col} = {flow_col} / {lpc_col}", False, 10, "222222", None),
        (f"{aesp_col} = {cons_col} / ({flow_col} x 60 / 1000)", False, 10, "222222", None),
        ("", False, 10, "000000", None),
        (t("xl_units",   lang), True, 11, C_HDR_FG, C_HDR_BG),
        ("LPM = litres per minute", False, 10, "222222", None),
        ("bar = gauge pressure (bar g)", False, 10, "222222", None),
        ("Nm3/h = normalised m3/hour (0 C, 1 atm)", False, 10, "222222", None),
        ("Nm3/m3 = air per m3 of liquid pumped", False, 10, "222222", None),
        ("", False, 10, "000000", None),
        (t("xl_sources",  lang), True, 11, C_HDR_FG, C_HDR_BG),
        ("Wilden PS820: EOM example point", False, 10, "222222", None),
        ("Sandpiper S20: datasheet curve read", False, 10, "222222", None),
        ("SAMOA UP20: datasheet curve read", False, 10, "222222", None),
        ("", False, 10, "000000", None),
        (t("xl_warning",  lang), False, 10, "884400", None),
    ]
    for ri, (text, bold, size, fg, bg) in enumerate(notes, 1):
        c2 = ws4[f"A{ri}"]; c2.value = text
        c2.font = Font(name="Calibri", bold=bold, size=size, color=fg)
        if bg: c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        ws4.row_dimensions[ri].height = 22 if bold else 16

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "pumps" not in st.session_state:
    st.session_state.pumps = [
        {**{"nombre": name, "notas": "Example - manufacturer source"},
         **{k: v for k,v in data.items() if k != "color"}}
        for name, data in EXAMPLE_PUMPS.items()
    ]
    for i,(name,data) in enumerate(EXAMPLE_PUMPS.items()):
        st.session_state.pumps[i]["color"] = data["color"]

if "lang_name" not in st.session_state:
    st.session_state["lang_name"] = "Español"

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE SELECTOR
# ─────────────────────────────────────────────────────────────────────────────
lang_cols = st.columns(len(LANGUAGES))
for i, (lname, lcol) in enumerate(zip(LANGUAGES.keys(), lang_cols)):
    active = st.session_state["lang_name"] == lname
    label = LANG_LABELS[lname]
    if active:
        lcol.markdown(f'''<div style="text-align:center; background:rgba(0,212,255,0.15);
            border:1px solid #00d4ff; color:#00d4ff; border-radius:6px; padding:4px 0;
            font-family:IBM Plex Mono,monospace; font-size:0.72rem; font-weight:600;">{label}</div>''',
            unsafe_allow_html=True)
    else:
        if lcol.button(label, key=f"lang_{i}", use_container_width=True):
            st.session_state["lang_name"] = lname
            st.rerun()

LANG = LANGUAGES[st.session_state["lang_name"]]

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f'''
<div class="main-header">
  <div style="display:flex; align-items:center; gap:1.5rem;">
    <img src="data:image/jpeg;base64,{IMG_B64}"
         style="height:85px; width:85px; object-fit:contain; border-radius:8px;
                filter: drop-shadow(0 0 10px rgba(0,212,255,0.4));
                background:rgba(255,255,255,0.04); padding:4px;" alt="AODD Pump"/>
    <div>
      <div class="header-badge">AODD &middot; AIR-OPERATED DOUBLE DIAPHRAGM</div>
      <h1>{t("app_title", LANG)}</h1>
      <p>{t("app_subtitle", LANG)}</p>
    </div>
  </div>
</div>
''', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"### {t('pump_mgmt', LANG)}")
    st.markdown("---")

    with st.expander(f"+ {t('add_pump', LANG)}", expanded=False):
        new_nombre = st.text_input(t("pump_name", LANG), key="new_nombre", placeholder="Ej: ARO 2 SS")
        c1, c2 = st.columns(2)
        new_q    = c1.number_input(t("flow_lpm", LANG),   1.0, 5000.0, 300.0, 1.0, key="new_q")
        new_pd   = c2.number_input(t("disch_press", LANG),0.1, 20.0,   1.4,   0.1, key="new_pd")
        c3, c4 = st.columns(2)
        new_pa   = c3.number_input(t("air_press", LANG),  0.5, 15.0,   4.0,   0.1, key="new_pa")
        new_ca   = c4.number_input(t("air_cons", LANG),   1.0, 2000.0, 100.0, 1.0, key="new_ca")
        new_lc   = st.number_input(t("lpc", LANG),        0.1, 50.0,   5.0,   0.1, key="new_lc")
        new_note = st.text_input(t("notes_opt", LANG), key="new_note", placeholder="Source...")
        col_idx  = len(st.session_state.pumps) % len(PUMP_COLORS)
        if st.button(t("add_btn", LANG), use_container_width=True):
            if new_nombre.strip():
                st.session_state.pumps.append({
                    "nombre": new_nombre.strip(), "caudal_lpm": new_q,
                    "p_descarga_bar": new_pd, "p_aire_bar": new_pa,
                    "consumo_aire_nm3h": new_ca, "litros_ciclo": new_lc,
                    "notas": new_note, "color": PUMP_COLORS[col_idx],
                })
                st.rerun()
            else:
                st.error(t("name_req", LANG))

    st.markdown(f"### {t('active_pumps', LANG)}")
    to_delete = None
    for i, p in enumerate(st.session_state.pumps):
        with st.expander(f"{'o' if i==0 else '-'} {p['nombre']}", expanded=False):
            ca, cb = st.columns(2)
            p["caudal_lpm"]      = ca.number_input(t("flow_lpm",   LANG), 1.0,5000.0, float(p["caudal_lpm"]),      1.0, key=f"q{i}")
            p["p_descarga_bar"]  = cb.number_input(t("disch_press",LANG), 0.1,20.0,   float(p["p_descarga_bar"]),  0.1, key=f"pd{i}")
            cc, cd = st.columns(2)
            p["p_aire_bar"]      = cc.number_input(t("air_press",  LANG), 0.5,15.0,   float(p["p_aire_bar"]),      0.1, key=f"pa{i}")
            p["consumo_aire_nm3h"]= cd.number_input(t("air_cons",  LANG), 1.0,2000.0, float(p["consumo_aire_nm3h"]),0.5, key=f"ca{i}")
            p["litros_ciclo"]    = st.number_input(t("lpc",        LANG), 0.1,50.0,   float(p["litros_ciclo"]),    0.1, key=f"lc{i}")
            if i == 0:
                st.markdown(f'<div class="info-box">{t("ref_pump", LANG)}</div>', unsafe_allow_html=True)
            else:
                if st.button(f"{t('delete_pump', LANG)} {p['nombre']}", key=f"del{i}", use_container_width=True):
                    to_delete = i
    if to_delete is not None:
        st.session_state.pumps.pop(to_delete)
        st.rerun()

    st.markdown("---")
    st.markdown(f"### {t('ftmg_title', LANG)}")
    with st.expander(f"{t('ftmg_load', LANG)}", expanded=False):
        ftmg_file = st.file_uploader("CSV", type=["csv","txt"], key="ftmg_upload",
                                      label_visibility="collapsed")
        if ftmg_file:
            df_ftmg, flow_col, time_col, msg = parse_ftmg_csv(ftmg_file)
            if df_ftmg is None:
                st.error(f"Error: {msg}")
            else:
                st.success(f"{len(df_ftmg)} rows")
                st.markdown(f"Column: `{flow_col}`")
                unit_choice = st.radio(t("ftmg_unit", LANG),
                                        ["Nl/min (FTMg standard)","Nm3/h"], key="ftmg_unit")
                is_nl_min = "Nl/min" in unit_choice
                total = len(df_ftmg)
                r0, r1 = st.slider(t("ftmg_range", LANG), 0, total-1, (0, total-1), key="ftmg_range")
                df_sel = df_ftmg.iloc[r0:r1+1]
                mean_nm3h = nm3h_from_ftmg(df_sel, flow_col, is_nl_min)
                st.markdown(f'''<div class="kpi-card" style="margin:0.5rem 0;">
                  <div class="kpi-label">{t("ftmg_mean", LANG)}</div>
                  <div class="kpi-value">{mean_nm3h:.1f}</div>
                  <div class="kpi-unit">Nm&#xB3;/h</div></div>''', unsafe_allow_html=True)
                pump_names = [p["nombre"] for p in st.session_state.pumps]
                target = st.selectbox(t("ftmg_assign", LANG), pump_names, key="ftmg_target")
                if st.button(t("ftmg_apply", LANG), key="ftmg_apply", use_container_width=True):
                    for p in st.session_state.pumps:
                        if p["nombre"] == target:
                            p["consumo_aire_nm3h"] = round(mean_nm3h, 2)
                    st.success(f"OK: {target} -> {mean_nm3h:.1f} Nm3/h")
                    st.rerun()
                if st.checkbox(t("ftmg_chart", LANG), key="ftmg_chart_cb"):
                    import plotly.express as px
                    fig_f = px.line(df_sel, y=flow_col, title="FTMg data")
                    fig_f.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(13,28,40,0.6)",
                                        font=dict(color="#b0cad6"), height=220,
                                        margin=dict(l=40,r=20,t=40,b=30))
                    fig_f.update_traces(line_color="#00d4ff")
                    st.plotly_chart(fig_f, use_container_width=True,
                                    config={"displayModeBar": False})

    st.markdown("---")
    if st.button(t("restore_ex", LANG), use_container_width=True):
        st.session_state.pumps = [
            {**{"nombre": name, "notas": t("ex_note", LANG)},
             **{k: v for k,v in data.items() if k != "color"}}
            for name, data in EXAMPLE_PUMPS.items()
        ]
        for i,(name,data) in enumerate(EXAMPLE_PUMPS.items()):
            st.session_state.pumps[i]["color"] = data["color"]
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# BUILD DATAFRAME
# ─────────────────────────────────────────────────────────────────────────────
pumps = st.session_state.pumps
df = build_dataframe(pumps, LANG)

# ─────────────────────────────────────────────────────────────────────────────
# KPI CARDS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f'<div class="section-title">{t("kpi_section", LANG)}</div>', unsafe_allow_html=True)
cols_kpi = st.columns(len(pumps))
for i, (col, (_, row)) in enumerate(zip(cols_kpi, df.iterrows())):
    kpis = calcular_kpis(row[t("col_flow",LANG)], row[t("col_cons",LANG)], row[t("col_lpc",LANG)])
    with col:
        pump = pumps[i]
        color = pump.get("color", "#888")
        if kpis:
            ciclos, aire_esp, _ = kpis
            consumo = pump["consumo_aire_nm3h"]
            st.markdown(f'''
            <div class="pump-card">
              <div class="pump-card-title" style="color:{color};">&#9658; {pump["nombre"]}</div>
              <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:6px;">
                <div class="kpi-card" style="border-top-color:{color}">
                  <div class="kpi-label">{t("kpi_air", LANG)}</div>
                  <div class="kpi-value">{consumo:.1f}</div>
                  <div class="kpi-unit">Nm&#xB3;/h</div>
                </div>
                <div class="kpi-card" style="border-top-color:{color}">
                  <div class="kpi-label">{t("kpi_cycles", LANG)}</div>
                  <div class="kpi-value">{ciclos:.1f}</div>
                  <div class="kpi-unit">cpm</div>
                </div>
                <div class="kpi-card" style="border-top-color:{color}">
                  <div class="kpi-label">{t("kpi_specific", LANG)}</div>
                  <div class="kpi-value">{aire_esp:.3f}</div>
                  <div class="kpi-unit">Nm&#xB3;/m&#xB3;</div>
                </div>
              </div>
            </div>
            ''', unsafe_allow_html=True)
        else:
            st.markdown(f'''
            <div class="pump-card">
              <div class="pump-card-title" style="color:{color};">&#9658; {pump["nombre"]}</div>
              <div class="warn-box">{t("kpi_nodata", LANG)}</div>
            </div>
            ''', unsafe_allow_html=True)

st.markdown(f'<div class="info-box">{t("kpi_explain", LANG)}</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    t("tab_charts", LANG), t("tab_table", LANG),
    t("tab_deltas", LANG), t("tab_export", LANG),
    t("tab_manual", LANG),
])

with tab1:
    st.plotly_chart(chart_kpi_bars(df, LANG), use_container_width=True, config={"displayModeBar": False})
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(chart_radar(df, LANG), use_container_width=True, config={"displayModeBar": False})
    with c2:
        st.plotly_chart(chart_scatter(df, LANG), use_container_width=True, config={"displayModeBar": False})
    if len(pumps) >= 2:
        fig_wf = chart_waterfall(df, LANG, ref_idx=0)
        if fig_wf:
            st.plotly_chart(fig_wf, use_container_width=True, config={"displayModeBar": False})

with tab2:
    st.markdown(f'<div class="section-title">{t("table_title", LANG)}</div>', unsafe_allow_html=True)
    display_cols = [t("col_pump", LANG), t("col_flow", LANG), t("col_pdesc", LANG),
                    t("col_pair", LANG), t("col_cons", LANG), t("col_lpc", LANG),
                    t("col_cpm", LANG), t("col_aesp", LANG), t("col_aciclo", LANG)]
    st.dataframe(df[display_cols].set_index(t("col_pump", LANG)), use_container_width=True,
                 height=min(200 + len(pumps)*40, 500))
    st.markdown(f'<div class="info-box">{t("table_note", LANG)}</div>', unsafe_allow_html=True)

with tab3:
    st.markdown(f'<div class="section-title">{t("deltas_title", LANG)}</div>', unsafe_allow_html=True)
    ref_name = df.iloc[0][t("col_pump",LANG)]
    st.markdown(f'<div class="info-box">{t("ref_label", LANG)} <b>{ref_name}</b></div>', unsafe_allow_html=True)
    if len(pumps) < 2:
        st.markdown(f'<div class="warn-box">{t("need_two", LANG)}</div>', unsafe_allow_html=True)
    else:
        ref = df.iloc[0]
        for _, row in df.iloc[1:].iterrows():
            st.markdown(f"**{row[t('col_pump',LANG)]}** vs {ref_name}")
            metrics = [
                (t("col_cons", LANG), "Nm3/h",  True),
                (t("col_cpm",  LANG), "cpm",    False),
                ("_aire_esp_raw",     "Nm3/m3", True),
            ]
            cols_d = st.columns(len(metrics))
            for col_d, (field, unit, lower_better) in zip(cols_d, metrics):
                v = row[field]; r = ref[field]
                if v is None or r is None:
                    col_d.markdown("-"); continue
                delta = v - r
                pct = (delta/r*100) if r else 0
                worse = (delta > 0) if lower_better else (delta < 0)
                chip_bg = "rgba(255,80,80,0.12)" if worse else ("rgba(0,200,120,0.12)" if delta != 0 else "rgba(150,150,150,0.12)")
                chip_col = "#ff6060" if worse else ("#00c878" if delta != 0 else "#aaa")
                chip_brd = "#ff6060" if worse else ("#00c878" if delta != 0 else "#aaa")
                col_d.markdown(f'''<div style="text-align:center; padding:0.4rem;">
                  <div style="font-size:0.68rem; color:#7a9db5; font-family:IBM Plex Mono,monospace;
                       text-transform:uppercase; letter-spacing:1px; margin-bottom:4px;">{unit}</div>
                  <div style="font-size:1.25rem; font-weight:700; font-family:IBM Plex Mono,monospace;
                       color:#e8f4f8;">{v:.3f}</div>
                  <div style="margin-top:4px;"><span style="background:{chip_bg}; color:{chip_col};
                       border:1px solid {chip_brd}; padding:2px 10px; border-radius:20px;
                       font-size:0.77rem; font-family:IBM Plex Mono,monospace; font-weight:600;">
                    {delta:+.3f} &nbsp; {pct:+.1f}%</span></div>
                </div>''', unsafe_allow_html=True)
            st.markdown("---")

with tab4:
    st.markdown(f'<div class="section-title">{t("export_title", LANG)}</div>', unsafe_allow_html=True)
    xlsx_bytes = export_excel(pumps, df, LANG)
    st.download_button(
        label=t("export_btn", LANG),
        data=xlsx_bytes,
        file_name="AODD_Comparison.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="section-title">{t("preview_title", LANG)}</div>', unsafe_allow_html=True)
    st.dataframe(df[[c for c in df.columns if not c.startswith("_")]].set_index(t("col_pump",LANG)),
                 use_container_width=True)

with tab5:
    st.markdown('<div class="section-title">' + t("man_title", LANG) + '</div>', unsafe_allow_html=True)
    st.markdown('<div class="info-box">' + t("man_objective", LANG) + '</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    def sec_header(text, color="#ffbe00"):
        st.markdown(
            '<div style="font-family:IBM Plex Mono,monospace; font-size:0.78rem; color:' + color +
            '; text-transform:uppercase; letter-spacing:1.5px; margin:1rem 0 0.5rem 0;">' + text + '</div>',
            unsafe_allow_html=True)

    def sec_card(icon, title, items=None, body=None):
        html = ('<div style="background:linear-gradient(145deg,#1a2c3d,#142230);'
                'border:1px solid #1e3a4f; border-left:4px solid #00d4ff;'
                'border-radius:8px; padding:1rem 1.2rem; margin-bottom:1rem;">'
                '<div style="font-family:IBM Plex Mono,monospace; font-weight:600; color:#00d4ff;'
                'font-size:0.85rem; margin-bottom:0.6rem;">' + icon + " " + title + '</div>')
        if body:
            html += '<div style="color:#b0cad6; font-size:0.84rem;">' + body + '</div>'
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)
        if items:
            for item in items:
                st.markdown(
                    '<div style="display:flex; gap:0.6rem; margin:0.3rem 0 0.3rem 1rem;">'
                    '<span style="color:#00d4ff; flex-shrink:0;">&#9658;</span>'
                    '<span style="color:#b0cad6; font-size:0.84rem;">' + item + '</span></div>',
                    unsafe_allow_html=True)

    # A. Equal conditions
    sec_header(t("man_A_title", LANG))
    col_a1, col_a2 = st.columns(2)
    with col_a1:
        sec_card("💧", t("man_A1", LANG), t("man_A1_items", LANG))
        sec_card("⚙️", t("man_A3", LANG), t("man_A3_items", LANG))
    with col_a2:
        sec_card("💨", t("man_A2", LANG), t("man_A2_items", LANG))

    # B. FTMg Installation
    sec_header(t("man_B_title", LANG))
    sec_card("📡", "SICK FTMg", t("man_B_items", LANG))

    # C. Configuration
    sec_header(t("man_C_title", LANG))
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        sec_card("🔌", t("man_C1", LANG), t("man_C1_items", LANG))
    with col_c2:
        sec_card("📊", t("man_C2", LANG), t("man_C2_items", LANG))

    # E. Test protocol
    sec_header(t("man_E_title", LANG))
    sec_card("📋", t("man_E_matrix", LANG)[:40] + "...", body=t("man_E_matrix", LANG))
    sec_card("📝", t("man_E_record", LANG), t("man_E_record_items", LANG))
    sec_card("📐", t("man_E_kpi", LANG), t("man_kpi_items", LANG))

    # Cycles/min
    sec_header(t("man_cycles_title", LANG))
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        sec_card("🔄", t("man_cycles_title", LANG), body=t("man_cycles_def", LANG))
    with col_d2:
        sec_card("⏱️", t("man_cycles_title", LANG), body=t("man_cycles_method", LANG))

    # F. Golden rules
    sec_header(t("man_F_title", LANG), color="#ff6060")
    for rule in t("man_F_items", LANG):
        is_neg = any(rule.startswith(w) for w in ["NO ", "DO NOT", "NICHT", "NE PAS", "NAO", "NU ", "NEPO"])
        color  = "#ff6060" if is_neg else "#b0cad6"
        border = "#ff6060" if is_neg else "#00c878"
        icon   = "🚫" if is_neg else "✅"
        bg     = "rgba(255,96,96,0.04)" if is_neg else "rgba(0,200,120,0.04)"
        st.markdown(
            '<div style="display:flex; gap:0.6rem; margin:0.4rem 0; background:' + bg +
            '; border-left:3px solid ' + border + '; padding:0.4rem 0.8rem; border-radius:0 6px 6px 0;">'
            '<span style="flex-shrink:0;">' + icon + '</span>'
            '<span style="color:' + color + '; font-size:0.84rem;">' + rule + '</span></div>',
            unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(f'''
<div style="text-align:center; color:#3a5a70; font-family:IBM Plex Mono,monospace;
     font-size:0.7rem; border-top:1px solid #1e3a4f; padding-top:1rem;">
  {t("footer_txt", LANG)}
</div>
''', unsafe_allow_html=True)
